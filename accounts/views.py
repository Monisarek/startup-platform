import collections
import json
import logging
import os
import uuid
from decimal import Decimal
from random import choice, shuffle
import time
import datetime
from datetime import datetime as dt
import boto3
import requests
from boto3 import client
from dateutil.relativedelta import relativedelta
from django import forms
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.staticfiles import finders
from django.core.files.storage import default_storage
from django.core.paginator import Paginator
from django.db.models import Q
from django.core.serializers.json import DjangoJSONEncoder
from django.db import (
    models,
    transaction,
)
from django.db.models import (
    Avg,
    Case,
    Count,
    DecimalField,
    Exists,
    ExpressionWrapper,
    F,
    FloatField,
    Max,
    Min,
    OuterRef,
    Q,
    Subquery,
    Sum,
    Value,
    When,
)
from django.db.models.functions import (
    Coalesce,
    TruncMonth,
    Floor,
)
from django.http import HttpResponse, HttpResponseForbidden, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import render_to_string
from django.templatetags.static import static
from django.urls import reverse, reverse_lazy
from django.utils import timezone
from django.utils.decorators import method_decorator
from django.utils.text import slugify
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.contrib.messages import get_messages
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from .forms import (
    CommentForm,
    FranchiseCommentForm,
    AgencyCommentForm,
    SpecialistCommentForm,
    ContactForm,
    LoginForm,
    MessageForm,
    ModeratorTicketForm,
    ProfileEditForm,
    RegisterForm,
    StartupForm,
    FranchiseForm,
    AgencyForm,
    SpecialistForm,
    SupportTicketForm,
    UserSearchForm,
)
from .models import (
    ChatConversations,
    ChatParticipants,
    Comments,
    Directions,
    FranchiseDirections,
    EntityTypes,
    FileStorage,
    FileTypes,
    Franchises,
    InvestmentTransactions,
    Messages,
    MessageStatuses,
    NewsArticles,
    NewsLikes,
    NewsViews,
    PaymentMethods,
    ReviewStatuses,
    Startups,
    StartupTimeline,
    SupportTicket,
    TransactionTypes,
    Users,
    UserVotes,
    FranchiseVotes,
    Agencies,
    AgencyComments,
    AgencyVotes,
    Specialists,
    SpecialistComments,
    SpecialistVotes,
)
from .utils import send_telegram_support_message, send_telegram_contact_form_message
logger = logging.getLogger(__name__)

RATE_WINDOW_SECONDS = 60
RATE_MAX_ATTEMPTS = 15
BLOCK_SECONDS = 30
CAPTCHA_FAILS_THRESHOLD = 3
FREQUENT_ATTEMPTS_THRESHOLD = 3
CAPTCHA_INVALID_MESSAGE = "Неверный ответ на капчу."

def _session_key(prefix: str, suffix: str) -> str:
    return f"{prefix}_{suffix}"

def _now_ts() -> int:
    return int(time.time())

def _get_attempts_in_window(session, prefix: str) -> int:
    key = _session_key(prefix, "attempt_times")
    times = session.get(key, [])
    now_ts = _now_ts()
    recent = [t for t in times if now_ts - t <= RATE_WINDOW_SECONDS]
    session[key] = recent
    return len(recent)

def _record_attempt(session, prefix: str) -> None:
    key = _session_key(prefix, "attempt_times")
    times = session.get(key, [])
    times.append(_now_ts())
    session[key] = times
    if _get_attempts_in_window(session, prefix) > RATE_MAX_ATTEMPTS:
        session[_session_key(prefix, "block_until")] = _now_ts() + BLOCK_SECONDS

def _is_blocked(session, prefix: str) -> int:
    block_until = session.get(_session_key(prefix, "block_until"))
    if not block_until:
        return 0
    remaining = block_until - _now_ts()
    if remaining <= 0:
        session.pop(_session_key(prefix, "block_until"), None)
        return 0
    return remaining

def _should_require_captcha(session, prefix: str) -> bool:
    if session.get(_session_key(prefix, "captcha_required")):
        logger.debug(f"[{prefix}] Captcha required flag set in session")
        return True
    fail_count = _get_fail_count(session, prefix)
    attempts = _get_attempts_in_window(session, prefix)
    require = fail_count >= CAPTCHA_FAILS_THRESHOLD or attempts >= FREQUENT_ATTEMPTS_THRESHOLD
    logger.debug(f"[{prefix}] should_require_captcha? fail_count={fail_count}, attempts_in_window={attempts} => {require}")
    return require

def _generate_captcha(session, prefix: str) -> None:
    import random
    a = random.randint(1, 9)
    b = random.randint(1, 9)
    session[_session_key(prefix, "captcha_question")] = f"Сколько будет {a} + {b}?"
    session[_session_key(prefix, "captcha_expected")] = str(a + b)
    session[_session_key(prefix, "captcha_required")] = True
    session[_session_key(prefix, "captcha_set_at")] = _now_ts()
    logger.debug(f"[{prefix}] Generated captcha: question={session.get(_session_key(prefix, 'captcha_question'))}, expected={session.get(_session_key(prefix, 'captcha_expected'))}")

def _clear_captcha(session, prefix: str) -> None:
    session.pop(_session_key(prefix, "captcha_question"), None)
    session.pop(_session_key(prefix, "captcha_expected"), None)
    session.pop(_session_key(prefix, "captcha_required"), None)
    session.pop(_session_key(prefix, "captcha_set_at"), None)
    logger.debug(f"[{prefix}] Cleared captcha requirement")

def _reset_limits(session, prefix: str) -> None:
    session.pop(_session_key(prefix, "attempt_times"), None)
    session.pop(_session_key(prefix, "fail_count"), None)
    session.pop(_session_key(prefix, "fail_last_ts"), None)
    session.pop(_session_key(prefix, "block_until"), None)
    _clear_captcha(session, prefix)
    logger.debug(f"[{prefix}] Reset limits")

def _get_fail_count(session, prefix: str) -> int:
    last_ts = session.get(_session_key(prefix, "fail_last_ts"))
    if last_ts and (_now_ts() - last_ts) > RATE_WINDOW_SECONDS:
        session[_session_key(prefix, "fail_count")] = 0
        session.pop(_session_key(prefix, "fail_last_ts"), None)
        logger.debug(f"[{prefix}] Fail count expired window -> reset to 0")
    return session.get(_session_key(prefix, "fail_count"), 0)

def _inc_fail_count(session, prefix: str) -> int:
    count = _get_fail_count(session, prefix) + 1
    session[_session_key(prefix, "fail_count")] = count
    session[_session_key(prefix, "fail_last_ts")] = _now_ts()
    logger.debug(f"[{prefix}] Increased fail_count -> {count}")
    return count

def _expire_captcha_if_old(session, prefix: str) -> None:
    set_at = session.get(_session_key(prefix, "captcha_set_at"))
    if set_at and (_now_ts() - set_at) > RATE_WINDOW_SECONDS:
        logger.debug(f"[{prefix}] Captcha expired by time window")
        _clear_captcha(session, prefix)

def _clear_captcha_messages(request):
    storage = get_messages(request)
    kept = []
    for m in storage:
        if str(m) != CAPTCHA_INVALID_MESSAGE:
            kept.append((m.level, str(m)))
    for level, msg in kept:
        messages.add_message(request, level, msg)
def safe_create_file_storage(entity_type, entity_id, file_type, file_url, uploaded_at, startup, original_file_name):
    """
    Безопасно создает объект FileStorage, учитывая наличие/отсутствие поля original_file_name
    """
    if hasattr(FileStorage, 'original_file_name'):
        try:
            return FileStorage.objects.create(
                entity_type=entity_type,
                entity_id=entity_id,
                file_type=file_type,
                file_url=file_url,
                uploaded_at=uploaded_at,
                startup=startup,
                original_file_name=original_file_name,
            )
        except Exception:
            return FileStorage.objects.create(
                entity_type=entity_type,
                entity_id=entity_id,
                file_type=file_type,
                file_url=file_url,
                uploaded_at=uploaded_at,
                startup=startup,
            )
    else:
        return FileStorage.objects.create(
            entity_type=entity_type,
            entity_id=entity_id,
            file_type=file_type,
            file_url=file_url,
            uploaded_at=uploaded_at,
            startup=startup,
        )
def safe_create_file_storage_instance(entity_type, entity_id, file_type, file_url, uploaded_at, startup, original_file_name):
    """
    Безопасно создает и сохраняет экземпляр FileStorage, учитывая наличие/отсутствие поля original_file_name
    """
    if hasattr(FileStorage, 'original_file_name'):
        try:
            file_storage = FileStorage(
                entity_type=entity_type,
                entity_id=entity_id,
                file_type=file_type,
                file_url=file_url,
                uploaded_at=uploaded_at,
                startup=startup,
                original_file_name=original_file_name,
            )
            file_storage.save()
            return file_storage
        except Exception:
            file_storage = FileStorage(
                entity_type=entity_type,
                entity_id=entity_id,
                file_type=file_type,
                file_url=file_url,
                uploaded_at=uploaded_at,
                startup=startup,
            )
            file_storage.save()
            return file_storage
    else:
        file_storage = FileStorage(
            entity_type=entity_type,
            entity_id=entity_id,
            file_type=file_type,
            file_url=file_url,
            uploaded_at=uploaded_at,
            startup=startup,
        )
        file_storage.save()
        return file_storage
def get_unique_filename(original_name, startup_id, file_type_name):
    """
    Генерирует уникальное имя файла, добавляя (2), (3) и т.д. если файл с таким именем уже существует
    """
    name, ext = os.path.splitext(original_name)
    try:
        file_type = FileTypes.objects.get(type_name=file_type_name)
        if not hasattr(FileStorage, 'original_file_name'):
            return original_name
        try:
            existing_files = FileStorage.objects.filter(
                startup_id=startup_id,
                file_type=file_type,
                original_file_name=original_name
            )
            if not existing_files.exists():
                return original_name
            counter = 2
            while True:
                new_name = f"{name} ({counter}){ext}"
                existing_duplicate = FileStorage.objects.filter(
                    startup_id=startup_id,
                    file_type=file_type,
                    original_file_name=new_name
                )
                if not existing_duplicate.exists():
                    return new_name
                counter += 1
        except Exception:
            return original_name
    except FileTypes.DoesNotExist:
        logger.error(f"FileType '{file_type_name}' не найден")
        return original_name
DIRECTION_TRANSLATIONS = {
    'Beauty': 'Красота', 'Technology': 'Технологии', 'Healthcare': 'Здравоохранение', 'Health': 'Здоровье',
    'Finance': 'Финансы', 'Cafe': 'Кафе/рестораны', 'Restaurant': 'Кафе/рестораны', 'Delivery': 'Доставка',
    'Fastfood': 'Фастфуд', 'Sport': 'Спорт', 'Transport': 'Транспорт', 'Psychology': 'Психология',
    'AI': 'ИИ', 'Auto': 'Авто',
    'Education': 'Образование', 'Entertainment': 'Развлечения',
    'Fashion': 'Мода', 'Food': 'Еда', 'Gaming': 'Игры', 'Real Estate': 'Недвижимость', 'Travel': 'Путешествия',
    'Agriculture': 'Сельское хозяйство', 'Energy': 'Энергетика', 'Environment': 'Экология', 'Social': 'Социальные проекты', 'Media': 'Медиа',
    'E-commerce': 'Электронная коммерция', 'Biotech': 'Биотехнологии', 'Construction': 'Строительство',
    'Logistics': 'Логистика', 'Manufacturing': 'Производство', 'Retail': 'Розничная торговля', 'Security': 'Безопасность', 'Insurance': 'Страхование',
    'Legal': 'Юридические услуги', 'Consulting': 'Консалтинг', 'Marketing': 'Маркетинг', 'IT': 'ИТ', 'Software': 'Программное обеспечение',
    'Hardware': 'Аппаратное обеспечение', 'Mobile': 'Мобильные приложения', 'Web': 'Веб-разработка', 'Blockchain': 'Блокчейн',
    'Cryptocurrency': 'Криптовалюты', 'VR': 'Виртуальная реальность', 'AR': 'Дополненная реальность', 'IoT': 'Интернет вещей',
    'Robotics': 'Робототехника', 'Space': 'Космические технологии', 'Science': 'Наука', 'Research': 'Исследования',     'Other': 'Другое',
}
FIXED_CATEGORIES = [
    {'original_name': 'Technology', 'direction_name': 'Технологии'},
    {'original_name': 'Healthcare', 'direction_name': 'Здравоохранение'},
    {'original_name': 'Finance', 'direction_name': 'Финансы'},
    {'original_name': 'Education', 'direction_name': 'Образование'},
    {'original_name': 'Entertainment', 'direction_name': 'Развлечения'},
    {'original_name': 'Fashion', 'direction_name': 'Мода'},
    {'original_name': 'Food', 'direction_name': 'Еда'},
    {'original_name': 'Gaming', 'direction_name': 'Игры'},
    {'original_name': 'Real Estate', 'direction_name': 'Недвижимость'},
    {'original_name': 'Travel', 'direction_name': 'Путешествия'},
    {'original_name': 'Agriculture', 'direction_name': 'Сельское хозяйство'},
    {'original_name': 'Energy', 'direction_name': 'Энергетика'},
    {'original_name': 'Environment', 'direction_name': 'Экология'},
    {'original_name': 'Social', 'direction_name': 'Социальные проекты'},
]
def home(request):
    if not request.user.is_authenticated:
        import random
        from django.db.models import Avg, Count, Sum, F, Case, When, Value, FloatField, DecimalField
        from django.db.models.functions import Coalesce
        from django.templatetags.static import static
        startups_query = Startups.objects.filter(status="approved").annotate(
            rating_avg=Coalesce(Avg("uservotes__rating"), 0.0, output_field=FloatField()),
            voters_count=Count("uservotes", distinct=True),
            total_investors=Count("investmenttransactions__investor", distinct=True),
            current_funding=Coalesce(
                Sum("investmenttransactions__amount"), 0, output_field=DecimalField()
            ),
            comment_count=Count("comments", distinct=True),
            progress=Case(
                When(funding_goal__gt=0, then=(F("amount_raised") * 100.0 / F("funding_goal"))),
                default=Value(0),
                output_field=FloatField(),
            )
        )
        all_startups = list(startups_query)
        demo_startups = []
        if all_startups:
            num_startups = min(6, len(all_startups))
            demo_startups = random.sample(all_startups, num_startups)
        startups_data = []
        for startup in demo_startups:
            folder_choice = random.choice(['planets_round', 'planets_ring'])
            if folder_choice == 'planets_round':
                planet_num = random.randint(1, 15)
                planet_image_url = static(f"accounts/images/planetary_system/planets_round/{planet_num}.png")
            else:
                planet_num = random.randint(1, 6)
                planet_image_url = static(f"accounts/images/planetary_system/planets_ring/{planet_num}.png")
            startups_data.append({
                "id": startup.startup_id,
                "name": startup.title,
                "description": startup.short_description or startup.description[:200] if startup.description else "",
                "image": planet_image_url,
                "rating": round(startup.rating_avg, 2),
                "voters_count": startup.voters_count,
                "comment_count": startup.comment_count,
                "direction": startup.direction.direction_name if startup.direction else "Не указано",
                "funding_goal": f"{startup.funding_goal:,.0f} ₽".replace(",", " ") if startup.funding_goal else "Не указано",
                "valuation": f"{startup.valuation:,.0f} ₽".replace(",", " ") if startup.valuation else "Не указано",
                "investors": startup.total_investors,
                "progress": round(startup.progress, 2) if startup.progress is not None else 0,
                "investment_type": "Выкуп+инвестирование" if startup.both_mode else ("Только выкуп" if startup.only_buy else "Только инвестирование")
            })
        
        directions_data = FIXED_CATEGORIES.copy()
        selected_startups = []
        if len(all_startups) > 0:
            selected_startups = all_startups[:6]
        else:
            selected_startups = []
        planets_data = []
        for i, startup in enumerate(selected_startups):
            planet_image_url = None
            
            if startup.planet_image:
                planet_image_url = f"{settings.S3_PUBLIC_BASE_URL}/choosable_planets/{startup.planet_image}"
            
            if not planet_image_url:
                import random
                folder_choice = random.choice(['planets_round', 'planets_ring'])
                if folder_choice == 'planets_round':
                    planet_image_num = (i % 15) + 1
                    planet_image_url = f"/static/accounts/images/planetary_system/planets_round/{planet_image_num}.png"
                else:
                    planet_image_num = (i % 6) + 1
                    planet_image_url = f"/static/accounts/images/planetary_system/planets_ring/{planet_image_num}.png"
            
            direction_original = 'Не указано'
            if startup.direction:
                for cat in directions_data:
                    if cat['direction_name'] == startup.direction.direction_name or cat['original_name'] == getattr(startup.direction, 'original_name', None):
                        direction_original = cat['original_name']
                        break
            planets_data.append({
                "id": startup.startup_id,
                "startup_id": startup.startup_id,
                "name": startup.title,
                "description": startup.short_description or startup.description[:200] if startup.description else "",
                "image": planet_image_url,
                "rating": startup.get_average_rating(),
                "voters_count": startup.total_voters,
                "comment_count": startup.comments.count(),
                "direction": direction_original,
                "funding_goal": f"{startup.funding_goal:,.0f} ₽".replace(",", " ") if startup.funding_goal else "Не указано",
                "valuation": f"{startup.valuation:,.0f} ₽".replace(",", " ") if startup.valuation else "Не указано",
                "investors": startup.get_investors_count(),
                "progress": startup.get_progress_percentage(),
                "investment_type": "Выкуп+инвестирование" if startup.both_mode else ("Только выкуп" if startup.only_buy else "Только инвестирование")
            })
        
        # Получаем случайных стартаперов для блока "Последние обновления от стартаперов"
        random_startupers = []
        try:
            # Получаем пользователей с ролью стартапера, у которых есть рейтинг
            startuper_users = Users.objects.filter(
                role__role_name__iexact='startuper',
                rating__isnull=False
            ).exclude(rating=0).order_by('?')[:3]
            
            for user in startuper_users:
                # Формируем звездный рейтинг
                rating = float(user.rating or 0)
                full_stars = int(rating)
                has_half_star = rating % 1 >= 0.5
                
                stars_html = "★" * full_stars
                if has_half_star:
                    stars_html += "☆"
                else:
                    stars_html += "☆" * (5 - full_stars)
                
                random_startupers.append({
                    'name': user.get_full_name() or user.username or f"Стартапер {user.user_id}",
                    'rating': rating,
                    'stars_html': stars_html,
                    'avatar_url': user.get_profile_picture_url() or static('accounts/images/avatars/default_avatar_ufo.png')
                })
            
            # Если стартаперов с рейтингом меньше 3, дополняем список
            if len(random_startupers) < 3:
                # Получаем дополнительных стартаперов без рейтинга
                additional_startupers = Users.objects.filter(
                    role__role_name__iexact='startuper'
                ).exclude(user_id__in=[s.get('user_id', 0) for s in random_startupers]).order_by('?')[:3-len(random_startupers)]
                
                for user in additional_startupers:
                    # Генерируем случайный рейтинг от 3.5 до 5.0
                    import random
                    rating = round(random.uniform(3.5, 5.0), 1)
                    full_stars = int(rating)
                    has_half_star = rating % 1 >= 0.5
                    
                    stars_html = "★" * full_stars
                    if has_half_star:
                        stars_html += "☆"
                    else:
                        stars_html += "☆" * (5 - full_stars)
                    
                    random_startupers.append({
                        'name': user.get_full_name() or user.username or f"Стартапер {user.user_id}",
                        'rating': rating,
                        'stars_html': stars_html,
                        'avatar_url': user.get_profile_picture_url() or static('accounts/images/avatars/default_avatar_ufo.png')
                    })
            
            # Если все еще нет стартаперов, используем fallback
            if len(random_startupers) == 0:
                random_startupers = [
                    {
                        'name': 'Виктор Смирнов',
                        'rating': 4.5,
                        'stars_html': '★★★★☆',
                        'avatar_url': static('accounts/images/avatars/default_avatar_ufo.png')
                    },
                    {
                        'name': 'Анна Кузнецова',
                        'rating': 4.9,
                        'stars_html': '★★★★★',
                        'avatar_url': static('accounts/images/avatars/default_avatar_ufo.png')
                    },
                    {
                        'name': 'Дмитрий Иванов',
                        'rating': 4.3,
                        'stars_html': '★★★★☆',
                        'avatar_url': static('accounts/images/avatars/default_avatar_ufo.png')
                    }
                ]
                
        except Exception as e:
            logger.error(f"Error getting random startupers: {e}")
            # Fallback данные если что-то пошло не так
            random_startupers = [
                {
                    'name': 'Виктор Смирнов',
                    'rating': 4.5,
                    'stars_html': '★★★★☆',
                    'avatar_url': static('accounts/images/avatars/default_avatar_ufo.png')
                },
                {
                    'name': 'Анна Кузнецова',
                    'rating': 4.9,
                    'stars_html': '★★★★★',
                    'avatar_url': static('accounts/images/avatars/default_avatar_ufo.png')
                },
                {
                    'name': 'Дмитрий Иванов',
                    'rating': 4.3,
                    'stars_html': '★★★★☆',
                    'avatar_url': static('accounts/images/avatars/default_avatar_ufo.png')
                }
            ]
            logger.info("Using fallback startuper data")
        
        # Получаем случайные стартапы для блока "Исследуйте новые миры"
        random_startups = []
        try:
            # Получаем случайные одобренные стартапы
            featured_startups = Startups.objects.filter(
                status="approved"
            ).order_by('?')[:3]
            
            # Если нет одобренных стартапов, пробуем получить любые
            if len(featured_startups) == 0:
                featured_startups = Startups.objects.all().order_by('?')[:3]
            
            for startup in featured_startups:
                # Получаем рейтинг стартапа
                try:
                    rating = startup.get_average_rating() or 0
                except Exception as e:
                    logger.warning(f"Could not get rating for startup {getattr(startup, 'title', 'Unknown')}: {e}")
                    rating = 4.5  # Fallback рейтинг
                
                rating_formatted = f"{rating:.1f}".replace('.', ',')
                
                # Получаем изображение стартапа
                startup_image = None
                if hasattr(startup, 'planet_image') and startup.planet_image:
                    startup_image = f"{settings.S3_PUBLIC_BASE_URL}/choosable_planets/{startup.planet_image}"
                else:
                    # Fallback на случайные изображения планет
                    import random
                    folder_choice = random.choice(['planets_round', 'planets_ring'])
                    if folder_choice == 'planets_round':
                        planet_num = random.randint(1, 15)
                        startup_image = static(f"accounts/images/planetary_system/planets_round/{planet_num}.png")
                    else:
                        planet_num = random.randint(1, 6)
                        startup_image = static(f"accounts/images/planetary_system/planets_ring/{planet_num}.png")
                
                # Получаем аватар владельца стартапа
                owner_avatar = static('accounts/images/avatars/default_avatar_ufo.png')
                try:
                    if hasattr(startup, 'owner') and startup.owner and hasattr(startup.owner, 'get_profile_picture_url'):
                        owner_avatar = startup.owner.get_profile_picture_url() or owner_avatar
                except Exception as e:
                    logger.warning(f"Could not get owner avatar for startup {getattr(startup, 'startup_id', 'Unknown')}: {e}")
                    owner_avatar = static('accounts/images/avatars/default_avatar_ufo.png')
                
                # Формируем описание
                description = getattr(startup, 'short_description', None) or getattr(startup, 'description', None) or "Описание стартапа"
                if len(description) > 100:
                    description = description[:97] + "..."
                
                # Проверяем, что startup_id является валидным числом
                startup_id = getattr(startup, 'startup_id', None)
                if startup_id and str(startup_id).isdigit():
                    startup_url = f"/startups/{startup_id}/"
                else:
                    startup_url = "/startups_list/"
                
                startup_data = {
                    'id': startup_id or 'Unknown',
                    'name': getattr(startup, 'title', 'Unknown'),
                    'rating': rating_formatted,
                    'description': description,
                    'image': startup_image,
                    'owner_avatar': owner_avatar,
                    'url': startup_url
                }
                
                random_startups.append(startup_data)
            
            # Если все еще нет стартапов, используем fallback
            if len(random_startups) == 0:
                random_startups = [
                    {
                        'id': 1,
                        'name': 'VoltForge Dynamics',
                        'rating': '4,4',
                        'description': 'VoltForge разрабатывает твердотельные батареи с графеновыми наноструктурами, которые заряжаются...',
                        'image': static('accounts/images/main_page/volt_forge.webp'),
                        'owner_avatar': static('accounts/images/avatars/default_avatar_ufo.png'),
                        'url': '/startups_list/'
                    },
                    {
                        'id': 2,
                        'name': 'NeuroBloom',
                        'rating': '4,7',
                        'description': 'NeuroBloom предлагает носимый гаджет с ИИ, который анализирует нейронные паттерны...',
                        'image': static('accounts/images/main_page/neuro_bloom.webp'),
                        'owner_avatar': static('accounts/images/avatars/default_avatar_ufo.png'),
                        'url': '/startups_list/'
                    },
                    {
                        'id': 3,
                        'name': 'BioCrop Nexus',
                        'rating': '4,2',
                        'description': 'BioCrop Nexus создает генетически оптимизированные семена, устойчивые к климату...',
                        'image': static('accounts/images/main_page/biocrop_nexus.webp'),
                        'owner_avatar': static('accounts/images/avatars/default_avatar_ufo.png'),
                        'url': '/startups_list/'
                    }
                ]
                
        except Exception as e:
            logger.error(f"Error getting random startups: {e}")
            # Fallback данные если что-то пошло не так
            random_startups = [
                {
                    'id': 1,
                    'name': 'VoltForge Dynamics',
                    'rating': '4,4',
                    'description': 'VoltForge разрабатывает твердотельные батареи с графеновыми наноструктурами, которые заряжаются...',
                    'image': static('accounts/images/main_page/volt_forge.webp'),
                    'owner_avatar': static('accounts/images/avatars/default_avatar_ufo.png'),
                    'url': '/startups_list/'
                },
                {
                    'id': 2,
                    'name': 'NeuroBloom',
                    'rating': '4,7',
                    'description': 'NeuroBloom предлагает носимый гаджет с ИИ, который анализирует нейронные паттерны...',
                    'image': static('accounts/images/main_page/neuro_bloom.webp'),
                    'owner_avatar': static('accounts/images/avatars/default_avatar_ufo.png'),
                    'url': '/startups_list/'
                },
                {
                    'id': 3,
                    'name': 'BioCrop Nexus',
                    'rating': '4,2',
                    'description': 'BioCrop Nexus создает генетически оптимизированные семена, устойчивые к климату...',
                    'image': static('accounts/images/main_page/biocrop_nexus.webp'),
                    'owner_avatar': static('accounts/images/avatars/default_avatar_ufo.png'),
                    'url': '/startups_list/'
                }
            ]
            logger.info("Using fallback startup data")
        
        context = {
            "demo_startups_data": json.dumps(startups_data, cls=DjangoJSONEncoder),
            "planets_data_json": json.dumps(planets_data, ensure_ascii=False),
            "directions_data_json": json.dumps(directions_data, ensure_ascii=False),
            "directions": directions_data,
            "random_startupers": random_startupers,
            "random_startups": random_startups,
        }
        
        return render(request, "accounts/main.html", context)
        return render(request, "accounts/main.html", context)
    if hasattr(request.user, "role") and request.user.role:
        role_name = request.user.role.role_name.lower()
        if role_name == "investor":
            return redirect("investor_main")
        elif role_name == "startuper":
            return redirect("startuper_main")
        elif role_name == "moderator":
            return redirect("main_page_moderator")
    return redirect("profile")
def faq_page_view(request):
    return render(request, "accounts/faq.html")
@login_required
def contacts_page_view(request):
    prefix = "contacts"
    _expire_captcha_if_old(request.session, prefix)
    captcha_q = None
    
    if request.method == "POST":
        form = ContactForm(request.POST)
        
        # Для формы контактов капча всегда обязательна
        if True:  # Всегда показываем капчу для формы контактов
            _expire_captcha_if_old(request.session, prefix)
            expected = request.session.get(_session_key(prefix, "captcha_expected"))
            answer_raw = (form.data.get("captcha_answer") or "").strip()
            try:
                answer_normalized = str(int(answer_raw))
            except ValueError:
                answer_normalized = ""
            if not expected or answer_normalized != expected:
                _clear_captcha_messages(request)
                messages.error(request, CAPTCHA_INVALID_MESSAGE)
                _record_attempt(request.session, prefix)
                _inc_fail_count(request.session, prefix)
                _generate_captcha(request.session, prefix)
                captcha_q = request.session.get(_session_key(prefix, "captcha_question"))
                return render(request, "accounts/contacts.html", {"form": form, "captcha_question": captcha_q})
            else:
                logger.debug("[contacts] captcha ok, clearing requirement")
                _clear_captcha(request.session, prefix)
        
        if form.is_valid():
            name = form.cleaned_data['name']
            email = form.cleaned_data['email']
            subject = form.cleaned_data['subject']
            message = form.cleaned_data['message']
            
            try:
                # Отправляем сообщение в Telegram
                logger.info(f"Sending contact form message from {email} to Telegram")
                sent_ok = send_telegram_contact_form_message(name, email, subject, message)
                logger.info(f"Telegram dispatch result for contact form from {email}: {sent_ok}")
                
                if sent_ok:
                    messages.success(request, "Спасибо за ваше сообщение! Мы свяжемся с вами в ближайшее время.")
                else:
                    messages.warning(request, "Сообщение отправлено, но возникли проблемы с уведомлением. Мы все равно получим ваше обращение.")
                
                # Перенаправляем на ту же страницу с GET запросом, чтобы избежать повторной отправки
                return redirect('contacts')
            except Exception as e:
                logger.error(f"Unexpected error during Telegram dispatch for contact form from {email}: {e}", exc_info=True)
                messages.success(request, "Спасибо за ваше сообщение! Мы свяжемся с вами в ближайшее время.")
                # Перенаправляем на ту же страницу с GET запросом, чтобы избежать повторной отправки
                return redirect('contacts')
        else:
            # Для формы контактов капча всегда обязательна
            _generate_captcha(request.session, prefix)
            captcha_q = request.session.get(_session_key(prefix, "captcha_question"))
    else:
        form = ContactForm()
        # Для формы контактов капча всегда обязательна
        _generate_captcha(request.session, prefix)
        captcha_q = request.session.get(_session_key(prefix, "captcha_question"))
    
    return render(request, "accounts/contacts.html", {"form": form, "captcha_question": captcha_q})
def register(request):
    next_url = request.GET.get("next") or request.POST.get("next")
    prefix = "register"
    block_left = _is_blocked(request.session, prefix)
    _expire_captcha_if_old(request.session, prefix)
    captcha_q = None

    if request.method == "POST":
        if block_left:
            _clear_captcha_messages(request)
            messages.error(request, f"Слишком много попыток. Попробуйте через {block_left} сек.")
            form = RegisterForm(request.POST)
            return render(request, "accounts/register.html", {"form": form, "next": next_url})

        form = RegisterForm(request.POST)
        if _should_require_captcha(request.session, prefix):
            _expire_captcha_if_old(request.session, prefix)
            expected = request.session.get(_session_key(prefix, "captcha_expected"))
            answer_raw = (form.data.get("captcha_answer") or "").strip()
            try:
                answer_normalized = str(int(answer_raw))
            except ValueError:
                answer_normalized = ""
            if not expected or answer_normalized != expected:
                _clear_captcha_messages(request)
                messages.error(request, CAPTCHA_INVALID_MESSAGE)
                _record_attempt(request.session, prefix)
                _inc_fail_count(request.session, prefix)
                _generate_captcha(request.session, prefix)
                captcha_q = request.session.get(_session_key(prefix, "captcha_question"))
                logger.debug(f"[register] captcha invalid: provided={answer_normalized!r}, expected={expected!r}")
                return render(request, "accounts/register.html", {"form": form, "next": next_url, "captcha_question": captcha_q})
            else:
                logger.debug("[register] captcha ok, clearing requirement")
                _clear_captcha(request.session, prefix)

        if form.is_valid():
            user = form.save(commit=False)
            user.set_password(form.cleaned_data["password"])
            user.save()
            _reset_limits(request.session, prefix)
            messages.success(
                request, "Регистрация прошла успешно! Теперь вы можете войти."
            )
            if next_url:
                login_url = reverse("login") + f"?next={next_url}"
                return redirect(login_url)
            return redirect("login")
        else:
            _record_attempt(request.session, prefix)
            _inc_fail_count(request.session, prefix)
            if _should_require_captcha(request.session, prefix):
                _generate_captcha(request.session, prefix)
                captcha_q = request.session.get(_session_key(prefix, "captcha_question"))
                return render(request, "accounts/register.html", {"form": form, "next": next_url, "captcha_question": captcha_q})
            return render(request, "accounts/register.html", {"form": form, "next": next_url})
    else:
        form = RegisterForm()
    _clear_captcha_messages(request)
    return render(request, "accounts/register.html", {"form": form, "next": next_url})
def user_login(request):
    logger.debug("Entering user_login view")
    next_url = request.GET.get("next") or request.POST.get("next")
    prefix = "login"
    block_left = _is_blocked(request.session, prefix)
    _expire_captcha_if_old(request.session, prefix)
    captcha_q = None

    if request.method == "POST":
        logger.debug("Processing POST request in user_login")
        if block_left:
            _clear_captcha_messages(request)
            messages.error(request, f"Слишком много попыток. Попробуйте через {block_left} сек.")
            form = LoginForm(request.POST)
            return render(request, "accounts/login.html", {"form": form, "next": next_url})

        form = LoginForm(request.POST)
        if _should_require_captcha(request.session, prefix):
            _expire_captcha_if_old(request.session, prefix)
            expected = request.session.get(_session_key(prefix, "captcha_expected"))
            answer_raw = (form.data.get("captcha_answer") or "").strip()
            try:
                answer_normalized = str(int(answer_raw))
            except ValueError:
                answer_normalized = ""
            if not expected or answer_normalized != expected:
                _clear_captcha_messages(request)
                messages.error(request, CAPTCHA_INVALID_MESSAGE)
                _record_attempt(request.session, prefix)
                _inc_fail_count(request.session, prefix)
                _generate_captcha(request.session, prefix)
                captcha_q = request.session.get(_session_key(prefix, "captcha_question"))
                logger.debug(f"[login] captcha invalid: provided={answer_normalized!r}, expected={expected!r}")
                return render(request, "accounts/login.html", {"form": form, "next": next_url, "captcha_question": captcha_q})
            else:
                logger.debug("[login] captcha ok, clearing requirement")
                _clear_captcha(request.session, prefix)

        if form.is_valid():
            logger.debug(f"Form is valid. Email: {form.cleaned_data['email']}")
            user = authenticate(
                request,
                username=form.cleaned_data["email"],
                password=form.cleaned_data["password"],
            )
            if user is not None:
                logger.info(f"User authenticated: {user.email}")
                _reset_limits(request.session, prefix)
                login(request, user)
                messages.success(
                    request, f"Добро пожаловать, {user.first_name or user.email}!"
                )
                if next_url == reverse("create_startup"):
                    role_name = user.role.role_name.lower() if hasattr(user, "role") and user.role else None
                    if role_name in ["startuper", "moderator"]:
                        return redirect(next_url)
                    else:
                        if role_name == "investor":
                            return redirect("investor_main")
                        elif role_name == "moderator":
                            return redirect("main_page_moderator")
                        else:
                            return redirect("home")
                if hasattr(user, "role") and user.role:
                    role_name = user.role.role_name.lower()
                    if role_name == "investor":
                        return redirect("investor_main")
                    elif role_name == "startuper":
                        return redirect("startuper_main")
                    elif role_name == "moderator":
                        return redirect("main_page_moderator")
                return redirect("home")
            else:
                logger.warning("Authentication failed for email")
                messages.error(request, "Неверный email или пароль.")
                _record_attempt(request.session, prefix)
                _inc_fail_count(request.session, prefix)
                if _should_require_captcha(request.session, prefix):
                    _generate_captcha(request.session, prefix)
                    captcha_q = request.session.get(_session_key(prefix, "captcha_question"))
        else:
            logger.warning(f"Form invalid: {form.errors}")
            _record_attempt(request.session, prefix)
            _inc_fail_count(request.session, prefix)
            if _should_require_captcha(request.session, prefix):
                _generate_captcha(request.session, prefix)
                captcha_q = request.session.get(_session_key(prefix, "captcha_question"))
        return render(request, "accounts/login.html", {"form": form, "next": next_url, "captcha_question": captcha_q})
    else:
        logger.debug("Rendering login form")
        form = LoginForm()
    _clear_captcha_messages(request)
    return render(request, "accounts/login.html", {"form": form, "next": next_url})
def user_logout(request):
    logout(request)
    messages.success(request, "Вы успешно вышли из системы.")
    return redirect("home")

def startups_list(request):
    startup_directions = Directions.objects.filter(
        direction_name__in=[
            'Technology', 'Healthcare', 'Finance', 'Education', 'Entertainment', 
            'Fashion', 'Food', 'Gaming', 'Real Estate', 'Travel', 'Agriculture', 
            'Energy', 'Environment', 'Social', 'Medicine', 'Auto', 'Delivery', 
            'Cafe', 'Fastfood', 'Health', 'Beauty', 'Transport', 'Sport', 
            'Psychology', 'AI', 'IT', 'Retail'
        ]
    ).order_by('direction_name')
    
    startups_qs = Startups.objects.filter(status="approved")
    selected_categories = request.GET.getlist("category")
    micro_investment_str = request.GET.get("micro_investment", "0")
    min_goal_str = request.GET.get("min_goal", "0")
    max_goal_str = request.GET.get("max_goal", "10000000")
    min_micro_str = request.GET.get("min_micro", "0")
    max_micro_str = request.GET.get("max_micro", "1000000")
    search_query = request.GET.get("search", "").strip()
    min_rating_str = request.GET.get("min_rating", "0")
    max_rating_str = request.GET.get("max_rating", "5")
    sort_order = request.GET.get("sort_order", "newest")
    page_number = request.GET.get("page", 1)
    
    startups_qs = startups_qs.annotate(
        total_voters_agg=Count("uservotes", distinct=True),
        rating_agg=ExpressionWrapper(
            Coalesce(Avg("uservotes__rating"), 0.0), output_field=FloatField()
        ),
        total_investors_agg=Count("investmenttransactions__investor", distinct=True),
        rating_bucket=Floor(Coalesce(Avg("uservotes__rating"), 0.0)),
    )
    
    categories = list(
        Directions.objects.annotate(id=F("direction_id"), name=F("direction_name"))
        .values("id", "name")
        .order_by("name")
    )
    
    if selected_categories:
        startups_qs = startups_qs.filter(
            direction__direction_name__in=selected_categories
        )
    
    micro_investment = micro_investment_str == "1"
    if micro_investment:
        startups_qs = startups_qs.filter(micro_investment_available=True)

    if search_query:
        startups_qs = startups_qs.filter(title__icontains=search_query)
    
    try:
        min_goal = int(min_goal_str)
        max_goal = int(max_goal_str)
        if min_goal > 0:
            startups_qs = startups_qs.filter(funding_goal__gte=min_goal)
        if max_goal < 10000000:
            startups_qs = startups_qs.filter(funding_goal__lte=max_goal)
    except ValueError:
        min_goal = 0
        max_goal = 10000000
    
    try:
        min_micro = int(min_micro_str)
        max_micro = int(max_micro_str)
        if min_micro > 0:
            startups_qs = startups_qs.filter(percent_amount__gte=min_micro)
        if max_micro < 1000000:
            startups_qs = startups_qs.filter(percent_amount__lte=max_micro)
    except ValueError:
        min_micro = 0
        max_micro = 1000000
    
    try:
        min_rating = float(min_rating_str)
        max_rating = float(max_rating_str)
        if min_rating > 0:
            startups_qs = startups_qs.filter(rating_agg__gte=min_rating)
        if max_rating < 5:
            startups_qs = startups_qs.filter(rating_agg__lte=max_rating)
    except ValueError:
        min_rating = 0
        max_rating = 5
    
    filters_active = (
        bool(selected_categories) or
        (search_query != "") or
        (min_goal > 0) or (max_goal < 10000000) or
        (min_micro > 0) or (max_micro < 1000000) or
        (min_rating > 0) or (max_rating < 5) or
        micro_investment
    )
    rating_active = (min_rating > 0 or max_rating < 5)
    goal_active = (min_goal > 0 or max_goal < 10000000)
    micro_active = (min_micro > 0 or max_micro < 1000000)
    if goal_active:
        startups_qs = startups_qs.order_by("funding_goal", "rating_bucket", "rating_agg", "-created_at")
    elif micro_active:
        startups_qs = startups_qs.order_by("percent_amount", "rating_bucket", "rating_agg", "-created_at")
    elif rating_active:
        startups_qs = startups_qs.order_by("rating_bucket", "rating_agg", "-created_at")
    elif filters_active:
        startups_qs = startups_qs.order_by("-created_at")
    else:
        if sort_order == "newest":
            startups_qs = startups_qs.order_by("-created_at")
        elif sort_order == "oldest":
            startups_qs = startups_qs.order_by("created_at")
    
    paginator = Paginator(startups_qs, 6)
    page_obj = paginator.get_page(page_number)
    
    is_ajax = request.headers.get("x-requested-with") == "XMLHttpRequest"
    if is_ajax:
        html = render_to_string(
            "accounts/partials/_startup_cards.html", {"page_obj": page_obj}
        )
        return JsonResponse(
            {
                "html": html,
                "has_next": page_obj.has_next(),
                "page_number": page_obj.number,
                "num_pages": paginator.num_pages,
                "count": paginator.count,
            }
        )
    else:
        context = {
            "page_obj": page_obj,
            "paginator": paginator,
            "initial_has_next": page_obj.has_next(),
            "selected_categories": selected_categories,
            "search_query": search_query,
            "min_rating": min_rating,
            "max_rating": max_rating,
            "min_goal": min_goal,
            "max_goal": max_goal,
            "min_micro": min_micro,
            "max_micro": max_micro,
            "micro_investment": micro_investment,
            "sort_order": sort_order,
            "directions": startup_directions,
        }
        return render(request, "accounts/startups_list.html", context)

def franchises_list(request):
    # Динамически формируем список доступных категорий по реально существующим направлениям у франшиз
    existing_dir_ids = (
        Franchises.objects.filter(status="approved", direction__isnull=False)
        .values_list("direction_id", flat=True)
        .distinct()
    )
    franchise_directions = Directions.objects.filter(direction_id__in=existing_dir_ids).order_by("direction_name")
    
    if not Franchises.objects.exists():
        franchise_names = [
            "БургерХаус", "КофеМир", "ПиццаПлюс", "СушиБар", "ШаурмаСтар", "КебабКинг",
            "БлинХаус", "ПельменХаус", "СупБар", "СалатСтар", "ДесертХаус", "МороженоМир",
            "ЧайХаус", "СокБар", "КоктейльХаус", "СмузиСтар", "МилкшейкБар", "ЛимонадХаус",
            "СпортХаус", "ФитнесСтар", "ЙогаБар", "ТанцыХаус", "БоксСтар", "ПлаваниеБар",
            "БегХаус", "ВелосипедСтар", "ТренажерБар", "БассейнХаус", "СаунаСтар", "МассажБар",
            "КрасотаХаус", "СалонСтар", "ПарикмахерБар", "МаникюрХаус", "ПедикюрСтар", "СПАБар",
            "КосметикаХаус", "ПарфюмСтар", "УходБар", "МакияжХаус", "ПрическаСтар", "СтильБар",
            "ОдеждаХаус", "ОбувьСтар", "АксессуарБар", "СумкаХаус", "РеменьСтар", "ЧасыБар",
            "УкрашениеХаус", "БижутерияСтар", "ЗолотоБар", "СереброХаус", "ПлатинаСтар", "БриллиантБар"
        ]
        
        startups = Startups.objects.filter(status="approved")
        for i, startup in enumerate(startups):
            franchise_name = franchise_names[i % len(franchise_names)]
            franchise_direction = franchise_directions[i % len(franchise_directions)] if franchise_directions.exists() else None
            
            franchise = Franchises.objects.create(
                title=franchise_name,
                short_description=startup.short_description,
                description=startup.description,
                terms=startup.terms,
                direction=franchise_direction,
                stage=startup.stage,
                investment_size=startup.funding_goal,
                payback_period=12,
                own_businesses=5,
                franchise_businesses=15,
                valuation=startup.valuation,
                pitch_deck_url=startup.pitch_deck_url,
                created_at=startup.created_at,
                updated_at=startup.updated_at,
                status="approved",
                total_invested=startup.total_invested,
                info_url=startup.info_url,
                percent_amount=startup.percent_amount,
                customization_data=startup.customization_data,
                total_voters=startup.total_voters,
                sum_votes=startup.sum_votes,
                is_edited=startup.is_edited,
                moderator_comment=startup.moderator_comment,
                step_number=startup.step_number,
                logo_urls=startup.logo_urls,
                creatives_urls=startup.creatives_urls,
                proofs_urls=startup.proofs_urls,
                video_urls=startup.video_urls,
                planet_image=startup.planet_image,
                owner=startup.owner
            )
    
    franchises_qs = Franchises.objects.filter(status="approved")
    selected_categories = request.GET.getlist("category")
    min_payback_str = request.GET.get("min_payback", "0")
    max_payback_str = request.GET.get("max_payback", "60")
    min_investment_str = request.GET.get("min_investment", "0")
    max_investment_str = request.GET.get("max_investment", "10000000")
    search_query = request.GET.get("search", "").strip()
    min_rating_str = request.GET.get("min_rating", "0")
    max_rating_str = request.GET.get("max_rating", "5")
    sort_order = request.GET.get("sort_order", "newest")
    page_number = request.GET.get("page", 1)
    
    franchises_qs = franchises_qs.annotate(
        rating_agg=ExpressionWrapper(
            Case(
                When(total_voters__gt=0, then=F('sum_votes') * 1.0 / F('total_voters')),
                default=Value(0.0),
                output_field=FloatField(),
            ),
            output_field=FloatField()
        ),
        rating_bucket=Floor(
            Case(
                When(total_voters__gt=0, then=F('sum_votes') * 1.0 / F('total_voters')),
                default=Value(0.0),
                output_field=FloatField(),
            )
        ),
    )
    
    
    if selected_categories:
        try:
            selected_ids = [int(cid) for cid in selected_categories]
        except Exception:
            selected_ids = []
        if selected_ids:
            franchises_qs = franchises_qs.filter(direction_id__in=selected_ids)
    
    if search_query:
        franchises_qs = franchises_qs.filter(title__icontains=search_query)
    
    try:
        min_payback = int(min_payback_str)
        max_payback = int(max_payback_str)
        if min_payback > 0:
            franchises_qs = franchises_qs.filter(payback_period__gte=min_payback)
        if max_payback < 60:
            franchises_qs = franchises_qs.filter(payback_period__lte=max_payback)
    except ValueError:
        min_payback = 0
        max_payback = 60
    
    try:
        min_investment = int(min_investment_str)
        max_investment = int(max_investment_str)
        if max_investment <= 0 or max_investment < min_investment:
            max_investment = 10000000
        if min_investment < 0:
            min_investment = 0
        if min_investment > 0:
            franchises_qs = franchises_qs.filter(investment_size__gte=min_investment)
        if max_investment < 10000000:
            franchises_qs = franchises_qs.filter(investment_size__lte=max_investment)
    except ValueError:
        min_investment = 0
        max_investment = 10000000
    
    try:
        min_rating = float(min_rating_str)
        max_rating = float(max_rating_str)
        if min_rating > 0:
            franchises_qs = franchises_qs.filter(rating_agg__gte=min_rating)
        if max_rating < 5:
            franchises_qs = franchises_qs.filter(rating_agg__lte=max_rating)
    except ValueError:
        min_rating = 0
        max_rating = 5
    
    filters_active = (
        bool(selected_categories) or
        (search_query != "") or
        (min_payback > 0) or (max_payback < 60) or
        (min_investment > 0) or (max_investment < 10000000) or
        (min_rating > 0) or (max_rating < 5)
    )
    rating_active = (min_rating > 0 or max_rating < 5)
    payback_active = (min_payback > 0 or max_payback < 60)
    investment_active = (min_investment > 0 or max_investment < 10000000)
    if payback_active:
        franchises_qs = franchises_qs.order_by("payback_period", "investment_size", "rating_bucket", "rating_agg", "-created_at")
    elif investment_active:
        franchises_qs = franchises_qs.order_by("investment_size", "rating_bucket", "rating_agg", "-created_at")
    elif rating_active:
        franchises_qs = franchises_qs.order_by("rating_bucket", "rating_agg", "-created_at")
    elif filters_active:
        franchises_qs = franchises_qs.order_by("-created_at")
    else:
        if sort_order == "newest":
            franchises_qs = franchises_qs.order_by("-created_at")
        elif sort_order == "oldest":
            franchises_qs = franchises_qs.order_by("created_at")
    
    try:
        approved_count = Franchises.objects.filter(status="approved").count()
        if approved_count < 18:
            from .models import Startups
            available_startups = list(Startups.objects.filter(status="approved")[:100])
            seed_names = [
                "Orbit Digital", "NovaLab Studio", "Cometix", "PixelFoundry",
                "NeuroCraft", "Skyline Media", "Quantum Works", "AstroBrand",
                "DeepWave", "Hyperlink", "BlueOrbit", "MetaForge",
                "Brandverse", "CodeSmiths", "UXia Lab", "Visionary",
                "IdeaGarden", "EchoPixel", "CraftLabs", "MotionQuark",
            ]
            import random
            missing = 18 - approved_count
            for i in range(max(0, missing)):
                base_title = seed_names[i % len(seed_names)]
                title = f"{base_title} {random.randint(100, 999)}"
                st = available_startups[i % len(available_startups)] if available_startups else None
                try:
                    Franchises.objects.create(
                        title=title,
                        short_description=(st.short_description if st else "Агентство полного цикла"),
                        description=(st.description if st else "Услуги: разработка, дизайн, маркетинг."),
                        terms=(st.terms if st else "Условия обсуждаются индивидуально."),
                        direction=(st.direction if st else None),
                        stage=(st.stage if st else None),
                        investment_size=(st.funding_goal if st else 0),
                        payback_period=12,
                        own_businesses=0,
                        franchise_businesses=0,
                        valuation=(st.valuation if st else 0),
                        pitch_deck_url=(st.pitch_deck_url if st else None),
                        created_at=timezone.now(),
                        updated_at=timezone.now(),
                        status="approved",
                        total_invested=0,
                        info_url=(st.info_url if st else None),
                        percent_amount=(st.percent_amount if st else 0),
                        customization_data={"agency_category": random.choice([
                            "Веб-разработка", "Мобильная разработка", "Дизайн",
                            "Маркетинг", "ИИ", "Брендинг", "Видео и мультимедиа"
                        ])},
                        total_voters=0,
                        sum_votes=0,
                        is_edited=False,
                        moderator_comment=None,
                        step_number=1,
                        logo_urls=(st.logo_urls if st else []),
                        creatives_urls=(st.creatives_urls if st else []),
                        proofs_urls=(st.proofs_urls if st else []),
                        video_urls=(st.video_urls if st else []),
                        planet_image=(st.planet_image if st else None),
                        owner=(st.owner if st else None),
                        franchise_cost=0,
                        profit_calculation=None,
                    )
                except Exception:
                    pass
    except Exception:
        pass
    
    paginator = Paginator(franchises_qs, 6)
    page_obj = paginator.get_page(page_number)
    
    is_ajax = request.headers.get("x-requested-with") == "XMLHttpRequest"
    if is_ajax:
        html = render_to_string(
            "accounts/partials/_franchise_cards.html", {"page_obj": page_obj}
        )
        return JsonResponse(
            {
                "html": html,
                "has_next": page_obj.has_next(),
                "page_number": page_obj.number,
                "num_pages": paginator.num_pages,
                "count": paginator.count,
            }
        )
    else:
        context = {
            "page_obj": page_obj,
            "paginator": paginator,
            "initial_has_next": page_obj.has_next(),
            "selected_categories": selected_categories,
            "search_query": search_query,
            "min_rating": min_rating,
            "max_rating": max_rating,
            "min_payback": min_payback,
            "max_payback": max_payback,
            "min_investment": min_investment,
            "max_investment": max_investment,
            "sort_order": sort_order,
            "franchise_directions": franchise_directions,
        }
        return render(request, "accounts/franchises_list.html", context)
def agencies_list(request):

    agencies_qs = Agencies.objects.filter(status="approved")
    agency_categories = [
        "Веб-разработка",
        "Мобильная разработка",
        "Дизайн",
        "Маркетинг",
        "ИИ",
        "Брендинг",
        "Видео и мультимедиа",
    ]

    selected_categories = request.GET.getlist("category")
    search_query = request.GET.get("search", "").strip()
    min_rating_str = request.GET.get("min_rating", "0")
    max_rating_str = request.GET.get("max_rating", "5")
    sort_order = request.GET.get("sort_order", "newest")
    page_number = request.GET.get("page", 1)

    agencies_qs = agencies_qs.annotate(
        rating_agg=ExpressionWrapper(
            Case(
                When(total_voters__gt=0, then=F('sum_votes') * 1.0 / F('total_voters')),
                default=Value(0.0),
                output_field=FloatField(),
            ),
            output_field=FloatField()
        ),
        rating_bucket=Floor(
            Case(
                When(total_voters__gt=0, then=F('sum_votes') * 1.0 / F('total_voters')),
                default=Value(0.0),
                output_field=FloatField(),
            )
        ),
    )

    if selected_categories:
        agencies_qs = agencies_qs.filter(
            Q(customization_data__agency_category__in=selected_categories)
        )

    if search_query:
        agencies_qs = agencies_qs.filter(title__icontains=search_query)

    try:
        min_rating = float(min_rating_str)
        max_rating = float(max_rating_str)
        if min_rating > 0:
            agencies_qs = agencies_qs.filter(rating_agg__gte=min_rating)
        if max_rating < 5:
            agencies_qs = agencies_qs.filter(rating_agg__lte=max_rating)
    except ValueError:
        min_rating = 0
        max_rating = 5

    filters_active = (
        bool(selected_categories) or
        (search_query != "") or
        (min_rating > 0) or (max_rating < 5)
    )
    rating_active = (min_rating > 0 or max_rating < 5)
    if rating_active:
        agencies_qs = agencies_qs.order_by("rating_bucket", "rating_agg", "-created_at")
    elif filters_active:
        agencies_qs = agencies_qs.order_by("-created_at")
    else:
        if sort_order == "newest":
            agencies_qs = agencies_qs.order_by("-created_at")
        elif sort_order == "oldest":
            agencies_qs = agencies_qs.order_by("created_at")

    paginator = Paginator(agencies_qs, 6)
    page_obj = paginator.get_page(page_number)

    is_ajax = request.headers.get("x-requested-with") == "XMLHttpRequest"
    if is_ajax:
        html = render_to_string(
            "accounts/partials/_agency_cards.html", {"page_obj": page_obj}
        )
        return JsonResponse(
            {
                "html": html,
                "has_next": page_obj.has_next(),
                "page_number": page_obj.number,
                "num_pages": paginator.num_pages,
                "count": paginator.count,
            }
        )
    else:
        context = {
            "page_obj": page_obj,
            "paginator": paginator,
            "initial_has_next": page_obj.has_next(),
            "selected_categories": selected_categories,
            "search_query": search_query,
            "min_rating": min_rating,
            "max_rating": max_rating,
            "sort_order": sort_order,
            "agency_categories": agency_categories,
        }
        return render(request, "accounts/agencies_list.html", context)

def specialists_list(request):
    specialists_qs = Specialists.objects.filter(status="approved")
    specialist_categories = [
        "Веб-разработка",
        "Мобильная разработка",
        "Дизайн",
        "Маркетинг",
        "ИИ",
        "Брендинг",
        "Видео и мультимедиа",
    ]

    selected_categories = request.GET.getlist("category")
    search_query = request.GET.get("search", "").strip()
    min_rating_str = request.GET.get("min_rating", "0")
    max_rating_str = request.GET.get("max_rating", "5")
    sort_order = request.GET.get("sort_order", "newest")
    page_number = request.GET.get("page", 1)

    specialists_qs = specialists_qs.annotate(
        rating_agg=ExpressionWrapper(
            Case(
                When(total_voters__gt=0, then=F('sum_votes') * 1.0 / F('total_voters')),
                default=Value(0.0),
                output_field=FloatField(),
            ),
            output_field=FloatField()
        ),
        rating_bucket=Floor(
            Case(
                When(total_voters__gt=0, then=F('sum_votes') * 1.0 / F('total_voters')),
                default=Value(0.0),
                output_field=FloatField(),
            )
        ),
    )

    if selected_categories:
        specialists_qs = specialists_qs.filter(
            Q(customization_data__specialist_category__in=selected_categories)
        )

    if search_query:
        specialists_qs = specialists_qs.filter(title__icontains=search_query)

    try:
        min_rating = float(min_rating_str)
        max_rating = float(max_rating_str)
        if min_rating > 0:
            specialists_qs = specialists_qs.filter(rating_agg__gte=min_rating)
        if max_rating < 5:
            specialists_qs = specialists_qs.filter(rating_agg__lte=max_rating)
    except ValueError:
        min_rating = 0
        max_rating = 5

    filters_active = (
        bool(selected_categories) or
        (search_query != "") or
        (min_rating > 0) or (max_rating < 5)
    )
    rating_active = (min_rating > 0 or max_rating < 5)
    if rating_active:
        specialists_qs = specialists_qs.order_by("rating_bucket", "rating_agg", "-created_at")
    elif filters_active:
        specialists_qs = specialists_qs.order_by("-created_at")
    else:
        if sort_order == "newest":
            specialists_qs = specialists_qs.order_by("-created_at")
        elif sort_order == "oldest":
            specialists_qs = specialists_qs.order_by("created_at")

    paginator = Paginator(specialists_qs, 6)
    page_obj = paginator.get_page(page_number)

    is_ajax = request.headers.get("x-requested-with") == "XMLHttpRequest"
    if is_ajax:
        html = render_to_string(
            "accounts/partials/_specialist_cards.html", {"page_obj": page_obj}
        )
        return JsonResponse(
            {
                "html": html,
                "has_next": page_obj.has_next(),
                "page_number": page_obj.number,
                "num_pages": paginator.num_pages,
                "count": paginator.count,
            }
        )
    else:
        context = {
            "page_obj": page_obj,
            "paginator": paginator,
            "initial_has_next": page_obj.has_next(),
            "selected_categories": selected_categories,
            "search_query": search_query,
            "min_rating": min_rating,
            "max_rating": max_rating,
            "sort_order": sort_order,
            "specialist_categories": specialist_categories,
        }
        return render(request, "accounts/specialists_list.html", context)
def agency_detail(request, franchise_id):
    try:
        franchise = Agencies.objects.get(agency_id=franchise_id)
    except Agencies.DoesNotExist:
        return render(request, "accounts/404.html", status=404)

    if request.method == "POST":
        if not request.user.is_authenticated:
            return redirect("login")
        form = AgencyCommentForm(request.POST)
        if form.is_valid():
            comment = form.save(commit=False)
            comment.agency = franchise
            comment.user = request.user

            try:
                new_rating = int(form.cleaned_data.get("user_rating") or 0)
            except (TypeError, ValueError):
                new_rating = 0
            user_vote = AgencyVotes.objects.filter(user=request.user, agency=franchise).first()
            if 1 <= new_rating <= 5:
                comment.user_rating = new_rating
                if user_vote:
                    if user_vote.rating != new_rating:
                        franchise.sum_votes = (franchise.sum_votes or 0) + (new_rating - int(user_vote.rating or 0))
                        user_vote.rating = new_rating
                        user_vote.save(update_fields=["rating"])
                        franchise.save(update_fields=["sum_votes"])
                else:
                    AgencyVotes.objects.create(user=request.user, agency=franchise, rating=new_rating)
                    franchise.total_voters = (franchise.total_voters or 0) + 1
                    franchise.sum_votes = (franchise.sum_votes or 0) + new_rating
                    franchise.save(update_fields=["total_voters", "sum_votes"])
            else:
                if user_vote:
                    comment.user_rating = user_vote.rating

            comment.save()
            messages.success(request, "Ваш комментарий был добавлен.")
            return redirect("agency_detail", franchise_id=franchise.franchise_id)
        else:
            messages.error(request, "Ошибка при добавлении комментария.")
    else:
        form = AgencyCommentForm()

    agency_category = None
    try:
        agency_category = (franchise.customization_data or {}).get("agency_category")
    except Exception:
        agency_category = None

    if agency_category:
        candidates_qs = Agencies.objects.filter(
            customization_data__agency_category=agency_category,
            status="approved",
        ).exclude(agency_id=franchise_id)
    else:
        candidates_qs = Agencies.objects.filter(
            status="approved",
        ).exclude(agency_id=franchise_id)
    similar_franchises = candidates_qs.order_by("-created_at")[:4]

    comments_with_rating = (
        AgencyComments.objects.filter(agency=franchise, parent_comment__isnull=True)
        .annotate(
            user_vote_rating=models.Subquery(
                AgencyVotes.objects.filter(
                    agency=franchise, user=models.OuterRef("user_id")
                ).values("rating")[:1]
            )
        )
        .order_by("-created_at")
    )
    average_rating = franchise.get_average_rating()
    total_votes = franchise.total_voters
    user_has_voted = False
    if request.user.is_authenticated:
        user_has_voted = AgencyVotes.objects.filter(
            user=request.user, agency=franchise
        ).exists()
    rating_distribution_query = (
        AgencyVotes.objects.filter(agency=franchise)
        .values("rating")
        .annotate(count=Count("rating"))
        .order_by("-rating")
    )
    rating_distribution = {item["rating"]: item["count"] for item in rating_distribution_query}
    for i in range(1, 6):
        rating_distribution.setdefault(i, 0)

    context = {
        "franchise": franchise,
        "similar_franchises": similar_franchises,
        "has_similar": candidates_qs.exists(),
        "comments": comments_with_rating,
        "form": form,
        "average_rating": average_rating,
        "total_votes_count": total_votes,
        "user_has_voted": user_has_voted,
        "rating_distribution": rating_distribution,
    }
    return render(request, "accounts/agency_detail.html", context)

def specialist_detail(request, specialist_id):
    try:
        specialist = Specialists.objects.get(specialist_id=specialist_id)
    except Specialists.DoesNotExist:
        return render(request, "accounts/404.html", status=404)

    if request.method == "POST":
        if not request.user.is_authenticated:
            return redirect("login")
        form = SpecialistCommentForm(request.POST)
        if form.is_valid():
            comment = form.save(commit=False)
            comment.specialist = specialist
            comment.user = request.user

            try:
                new_rating = int(form.cleaned_data.get("user_rating") or 0)
            except (TypeError, ValueError):
                new_rating = 0
            user_vote = SpecialistVotes.objects.filter(user=request.user, specialist=specialist).first()
            if 1 <= new_rating <= 5:
                comment.user_rating = new_rating
                if user_vote:
                    if user_vote.rating != new_rating:
                        specialist.sum_votes = (specialist.sum_votes or 0) + (new_rating - int(user_vote.rating or 0))
                        user_vote.rating = new_rating
                        user_vote.save(update_fields=["rating"])
                        specialist.save(update_fields=["sum_votes"])
                else:
                    SpecialistVotes.objects.create(user=request.user, specialist=specialist, rating=new_rating)
                    specialist.total_voters = (specialist.total_voters or 0) + 1
                    specialist.sum_votes = (specialist.sum_votes or 0) + new_rating
                    specialist.save(update_fields=["total_voters", "sum_votes"])
            else:
                if user_vote:
                    comment.user_rating = user_vote.rating

            comment.save()
            messages.success(request, "Ваш комментарий был добавлен.")
            return redirect("specialist_detail", specialist_id=specialist.specialist_id)
        else:
            messages.error(request, "Ошибка при добавлении комментария.")
    else:
        form = SpecialistCommentForm()

    specialist_category = None
    try:
        specialist_category = (specialist.customization_data or {}).get("specialist_category")
    except Exception:
        specialist_category = None

    if specialist_category:
        candidates_qs = Specialists.objects.filter(
            customization_data__specialist_category=specialist_category,
            status="approved",
        ).exclude(specialist_id=specialist_id)
    else:
        candidates_qs = Specialists.objects.filter(
            status="approved",
        ).exclude(specialist_id=specialist_id)
    similar_specialists = candidates_qs.order_by("-created_at")[:4]

    comments_with_rating = (
        SpecialistComments.objects.filter(specialist=specialist, parent_comment__isnull=True)
        .annotate(
            user_vote_rating=models.Subquery(
                SpecialistVotes.objects.filter(
                    specialist=specialist, user=models.OuterRef("user_id")
                ).values("rating")[:1]
            )
        )
        .order_by("-created_at")
    )
    average_rating = specialist.get_average_rating()
    total_votes = specialist.total_voters
    user_has_voted = False
    if request.user.is_authenticated:
        user_has_voted = SpecialistVotes.objects.filter(
            user=request.user, specialist=specialist
        ).exists()
    rating_distribution_query = (
        SpecialistVotes.objects.filter(specialist=specialist)
        .values("rating")
        .annotate(count=Count("rating"))
        .order_by("-rating")
    )
    rating_distribution = {item["rating"]: item["count"] for item in rating_distribution_query}
    for i in range(1, 6):
        rating_distribution.setdefault(i, 0)

    context = {
        "specialist": specialist,
        "similar_specialists": similar_specialists,
        "has_similar": candidates_qs.exists(),
        "comments": comments_with_rating,
        "form": form,
        "average_rating": average_rating,
        "total_votes_count": total_votes,
        "user_has_voted": user_has_voted,
        "rating_distribution": rating_distribution,
    }
    return render(request, "accounts/specialist_detail.html", context)
def franchise_detail(request, franchise_id):
    try:
        franchise = Franchises.objects.get(franchise_id=franchise_id)
    except Franchises.DoesNotExist:
        return render(request, "accounts/404.html", status=404)

    if request.method == "POST":
        if not request.user.is_authenticated:
            return redirect("login")
        form = FranchiseCommentForm(request.POST)
        if form.is_valid():
            from .models import FranchiseComments
            comment = form.save(commit=False)
            comment.franchise = franchise
            comment.user = request.user
            try:
                new_rating = int(form.cleaned_data.get("user_rating") or 0)
            except (TypeError, ValueError):
                new_rating = 0
            user_vote = FranchiseVotes.objects.filter(user=request.user, franchise=franchise).first()
            if 1 <= new_rating <= 5:
                comment.user_rating = new_rating
                if user_vote:
                    if user_vote.rating != new_rating:
                        franchise.sum_votes = (franchise.sum_votes or 0) + (new_rating - int(user_vote.rating or 0))
                        user_vote.rating = new_rating
                        user_vote.save(update_fields=["rating"])
                        franchise.save(update_fields=["sum_votes"])
                else:
                    FranchiseVotes.objects.create(user=request.user, franchise=franchise, rating=new_rating)
                    franchise.total_voters = (franchise.total_voters or 0) + 1
                    franchise.sum_votes = (franchise.sum_votes or 0) + new_rating
                    franchise.save(update_fields=["total_voters", "sum_votes"])
            else:
                if user_vote:
                    comment.user_rating = user_vote.rating
            comment.save()
            messages.success(request, "Ваш комментарий был добавлен.")
            return redirect("franchise_detail", franchise_id=franchise.franchise_id)
        else:
            messages.error(request, "Ошибка при добавлении комментария.")
    else:
        form = FranchiseCommentForm()

    candidates_qs = Franchises.objects.filter(
        direction=franchise.direction,
        status="approved",
    ).exclude(franchise_id=franchise_id)
    similar_franchises = candidates_qs.order_by("-created_at")[:4]

    from .models import FranchiseComments
    comments_with_rating = (
        FranchiseComments.objects.filter(franchise=franchise, parent_comment__isnull=True)
        .annotate(
            user_vote_rating=models.Subquery(
                FranchiseVotes.objects.filter(
                    franchise=franchise, user=models.OuterRef("user_id")
                ).values("rating")[:1]
            )
        )
        .order_by("-created_at")
    )
    average_rating = franchise.get_average_rating()
    total_votes = franchise.total_voters
    user_has_voted = False
    if request.user.is_authenticated:
        user_has_voted = FranchiseVotes.objects.filter(
            user=request.user, franchise=franchise
        ).exists()
    rating_distribution_query = (
        FranchiseVotes.objects.filter(franchise=franchise)
        .values("rating")
        .annotate(count=Count("rating"))
        .order_by("-rating")
    )
    rating_distribution = {item["rating"]: item["count"] for item in rating_distribution_query}
    for i in range(1, 6):
        rating_distribution.setdefault(i, 0)

    context = {
        "franchise": franchise,
        "similar_franchises": similar_franchises,
        "has_similar": candidates_qs.exists(),
        "comments": comments_with_rating,
        "form": form,
        "average_rating": average_rating,
        "total_votes_count": total_votes,
        "user_has_voted": user_has_voted,
        "rating_distribution": rating_distribution,
    }
    return render(request, "accounts/franchise_detail.html", context)

def search_suggestions(request):
    query = request.GET.get("q", "").strip()
    users = []
    if len(query) >= 2:
        search_results = Users.objects.filter(
            Q(first_name__icontains=query)
            | Q(last_name__icontains=query)
            | Q(email__icontains=query)
        ).distinct()[:10]
        users = [
            {
                "id": user.user_id,
                "name": f"{user.first_name or ''} {user.last_name or ''} ({user.email})".strip(),
            }
            for user in search_results
        ]
    return JsonResponse({"suggestions": users})

def global_search(request):
    """Глобальный поиск по всем типам карточек"""
    try:
        query = request.GET.get("q", "").strip()
        
        if len(query) < 2:
            return JsonResponse({
                "users": [],
                "startups": [],
                "franchises": [],
                "agencies": [],
                "specialists": []
            })
        
        results = {
            "users": [],
            "startups": [],
            "franchises": [],
            "agencies": [],
            "specialists": []
        }
        
        # Поиск пользователей
        try:
            users = Users.objects.filter(
                Q(first_name__icontains=query) |
                Q(last_name__icontains=query) |
                Q(email__icontains=query)
            ).distinct()[:5]
            
            for user in users:
                try:
                    results["users"].append({
                        "id": user.user_id,
                        "name": f"{user.first_name or ''} {user.last_name or ''}".strip() or user.email,
                        "type": "user",
                        "url": reverse('profile', kwargs={'user_id': user.user_id})
                    })
                except Exception as e:
                    logger.error(f"Ошибка при обработке пользователя {user.user_id}: {e}")
                    continue
        except Exception as e:
            logger.error(f"Ошибка при поиске пользователей: {e}")
        
        # Поиск стартапов
        try:
            # У Startups есть и status (CharField) и status_id (ForeignKey)
            try:
                # Сначала пробуем найти по status_id
                approved_status = ReviewStatuses.objects.get(status_name="Approved")
                startups = Startups.objects.filter(
                    Q(title__icontains=query) |
                    Q(short_description__icontains=query)
                ).filter(status_id=approved_status).distinct()[:5]
            except ReviewStatuses.DoesNotExist:
                # Если статус не найден, ищем по полю status
                startups = Startups.objects.filter(
                    Q(title__icontains=query) |
                    Q(short_description__icontains=query)
                ).filter(status="approved").distinct()[:5]
            
            for startup in startups:
                try:
                    # Проверяем, что обязательные поля не пустые
                    if not startup.title:
                        continue
                        
                    results["startups"].append({
                        "id": startup.startup_id,
                        "name": startup.title,
                        "type": "startup",
                        "url": reverse('startup_detail', kwargs={'startup_id': startup.startup_id})
                    })
                except Exception as e:
                    logger.error(f"Ошибка при обработке стартапа {startup.startup_id}: {e}")
                    continue
        except Exception as e:
            logger.error(f"Ошибка при поиске стартапов: {e}")
        
        # Поиск франшиз
        try:
            # У Franchises есть и status (CharField) и status_id (ForeignKey)
            try:
                # Сначала пробуем найти по status_id
                approved_status = ReviewStatuses.objects.get(status_name="Approved")
                franchises = Franchises.objects.filter(
                    Q(title__icontains=query) |
                    Q(short_description__icontains=query)
                ).filter(status_id=approved_status).distinct()[:5]
            except ReviewStatuses.DoesNotExist:
                # Если статус не найден, ищем по полю status
                franchises = Franchises.objects.filter(
                    Q(title__icontains=query) |
                    Q(short_description__icontains=query)
                ).filter(status="approved").distinct()[:5]
            
            for franchise in franchises:
                try:
                    # Проверяем, что обязательные поля не пустые
                    if not franchise.title:
                        continue
                        
                    results["franchises"].append({
                        "id": franchise.franchise_id,
                        "name": franchise.title,
                        "type": "franchise",
                        "url": reverse('franchise_detail', kwargs={'franchise_id': franchise.franchise_id})
                    })
                except Exception as e:
                    logger.error(f"Ошибка при обработке франшизы {franchise.franchise_id}: {e}")
                    continue
        except Exception as e:
            logger.error(f"Ошибка при поиске франшиз: {e}")
        
        # Поиск агентств
        try:
            # У Agencies только поле status (CharField)
            agencies = Agencies.objects.filter(
                Q(title__icontains=query) |
                Q(short_description__icontains=query)
            ).filter(status="approved").distinct()[:5]
            
            for agency in agencies:
                try:
                    # Проверяем, что обязательные поля не пустые
                    if not agency.title:
                        continue
                        
                    results["agencies"].append({
                        "id": agency.agency_id,
                        "name": agency.title,
                        "type": "agency",
                        "url": reverse('agency_detail', kwargs={'agency_id': agency.agency_id})
                    })
                except Exception as e:
                    logger.error(f"Ошибка при обработке агентства {agency.agency_id}: {e}")
                    continue
        except Exception as e:
            logger.error(f"Ошибка при поиске агентств: {e}")
        
        # Поиск специалистов
        try:
            # У Specialists только поле status (CharField)
            specialists = Specialists.objects.filter(
                Q(title__icontains=query) |
                Q(short_description__icontains=query)
            ).filter(status="approved").distinct()[:5]
            
            for specialist in specialists:
                try:
                    # Проверяем, что обязательные поля не пустые
                    if not specialist.title:
                        continue
                        
                    results["specialists"].append({
                        "id": specialist.specialist_id,
                        "name": specialist.title,
                        "type": "specialist",
                        "url": reverse('specialist_detail', kwargs={'specialist_id': specialist.specialist_id})
                    })
                except Exception as e:
                    logger.error(f"Ошибка при обработке специалиста {specialist.specialist_id}: {e}")
                    continue
        except Exception as e:
            logger.error(f"Ошибка при поиске специалистов: {e}")
        
        return JsonResponse(results)
        
    except Exception as e:
        logger.error(f"Критическая ошибка в global_search: {e}")
        return JsonResponse({
            "error": "Ошибка при выполнении поиска",
            "details": str(e) if settings.DEBUG else "Внутренняя ошибка сервера"
        }, status=500)

def startup_detail(request, startup_id):
    try:
        startup = Startups.objects.select_related("owner", "direction", "stage").get(
            startup_id=startup_id
        )
    except Startups.DoesNotExist:
        return get_object_or_404(Startups, startup_id=startup_id)
    if request.method == "POST":
        if "status" in request.POST:
            if not request.user.is_authenticated or not hasattr(request.user, "role") or (request.user.role.role_name or "") != "moderator":
                messages.error(request, "У вас нет прав для этого действия.")
                return redirect("startup_detail", startup_id=startup.startup_id)
            new_status = (request.POST.get("status", "") or "").strip().lower()
            allowed_statuses = {"approved", "blocked", "closed", "pending", "rejected"}
            if new_status in allowed_statuses:
                startup.status = new_status
                try:
                    status_map = {
                        "approved": "Approved",
                        "blocked": "Blocked",
                        "closed": "Closed",
                        "pending": "Pending",
                        "rejected": "Rejected",
                    }
                    startup.status_id = ReviewStatuses.objects.get(status_name=status_map[new_status])
                except ReviewStatuses.DoesNotExist:
                    pass
                startup.save(update_fields=["status", "status_id"])
                messages.success(request, "Статус стартапа обновлён.")
            else:
                messages.error(request, "Недопустимый статус.")
            return redirect("startup_detail", startup_id=startup.startup_id)
        if not request.user.is_authenticated:
            return redirect("login")
        form = CommentForm(request.POST)
        if form.is_valid():
            comment = form.save(commit=False)
            comment.startup_id = startup
            comment.user_id = request.user

            try:
                new_rating = int(form.cleaned_data.get("user_rating") or 0)
            except (TypeError, ValueError):
                new_rating = 0
            user_vote = UserVotes.objects.filter(user=request.user, startup=startup).first()
            if 1 <= new_rating <= 5:
                comment.user_rating = new_rating
                if user_vote:
                    if user_vote.rating != new_rating:
                        startup.sum_votes = (startup.sum_votes or 0) + (new_rating - int(user_vote.rating or 0))
                        user_vote.rating = new_rating
                        user_vote.save(update_fields=["rating"])
                        startup.save(update_fields=["sum_votes"])
                else:
                    UserVotes.objects.create(user=request.user, startup=startup, rating=new_rating)
                    startup.total_voters = (startup.total_voters or 0) + 1
                    startup.sum_votes = (startup.sum_votes or 0) + new_rating
                    startup.save(update_fields=["total_voters", "sum_votes"])
            else:
                if user_vote:
                    comment.user_rating = user_vote.rating

            comment.save()
            messages.success(request, "Ваш комментарий был добавлен.")
            return redirect("startup_detail", startup_id=startup.startup_id)
        else:
            messages.error(request, "Ошибка при добавлении комментария.")
    else:
        form = CommentForm()
    comments_with_rating = (
        Comments.objects.filter(startup_id=startup, parent_comment_id__isnull=True)
        .annotate(
            user_vote_rating=models.Subquery(
                UserVotes.objects.filter(
                    startup=startup, user=models.OuterRef("user_id_id")
                ).values("rating")[:1]
            )
        )
        .order_by("-created_at")
    )
    average_rating = (
        startup.sum_votes / startup.total_voters if startup.total_voters > 0 else 0
    )
    comments = Comments.objects.filter(
        startup_id=startup, parent_comment_id__isnull=True
    ).order_by("-created_at")
    form = CommentForm()
    average_rating = startup.get_average_rating()
    total_votes = startup.total_voters
    user_has_voted = False
    if request.user.is_authenticated:
        user_has_voted = UserVotes.objects.filter(
            user=request.user, startup=startup
        ).exists()
    rating_distribution_query = (
        UserVotes.objects.filter(startup=startup)
        .values("rating")
        .annotate(count=Count("rating"))
        .order_by("-rating")
    )
    rating_distribution = {
        item["rating"]: item["count"] for item in rating_distribution_query
    }
    for i in range(1, 6):
        rating_distribution.setdefault(i, 0)
    similar_startups = (
        Startups.objects.filter(status="approved")
        .exclude(startup_id=startup.startup_id)
        .order_by("?")[:4]
    )
    similar_startups = similar_startups.annotate(
        average_rating_calc=Avg(
            models.ExpressionWrapper(
                models.F("sum_votes") * 1.0 / models.F("total_voters"),
                output_field=FloatField(),
            ),
            filter=models.Q(total_voters__gt=0),
        )
    ).annotate(average_rating=Coalesce("average_rating_calc", 0.0))
    logo_urls = startup.logo_urls if isinstance(startup.logo_urls, list) else []
    creatives_urls = (
        startup.creatives_urls if isinstance(startup.creatives_urls, list) else []
    )
    video_urls = startup.video_urls if isinstance(startup.video_urls, list) else []
    show_moderator_comment = False
    if startup.moderator_comment and (
        request.user == startup.owner
        or (
            request.user.is_authenticated
            and hasattr(request.user, "role")
            and request.user.role.role_name == "moderator"
        )
    ):
        show_moderator_comment = True
    progress_percentage = 0
    if startup.funding_goal and startup.funding_goal > 0:
        progress_percentage = (
            (startup.amount_raised / startup.funding_goal) * 100
            if startup.amount_raised
            else 0
        )
        progress_percentage = min(progress_percentage, 100)
    investors_count = startup.get_investors_count()
    timeline_events = StartupTimeline.objects.filter(startup=startup).order_by(
        "step_number"
    )
    try:
        proof_file_type = FileTypes.objects.get(type_name="proof")
        startup_documents = FileStorage.objects.filter(
            startup=startup, file_type=proof_file_type
        ).order_by("-uploaded_at")
    except FileTypes.DoesNotExist:
        startup_documents = FileStorage.objects.none()
    context = {
        "startup": startup,
        "comments": comments_with_rating,
        "form": form,
        "average_rating": average_rating,
        "total_votes_count": total_votes,
        "user_has_voted": user_has_voted,
        "rating_distribution": rating_distribution,
        "similar_startups": similar_startups,
        "logo_urls": logo_urls,
        "creatives_urls": creatives_urls,
        "video_urls": video_urls,
        "show_moderator_comment": show_moderator_comment,
        "progress_percentage": progress_percentage,
        "investors_count": investors_count,
        "timeline_events": timeline_events,
        "startup_documents": startup_documents,
    }
    return render(request, "accounts/startup_detail.html", context)
def load_similar_startups(request, startup_id: int):
    current_startup_id = startup_id
    similar_startups = (
        Startups.objects.filter(status="approved")
        .exclude(startup_id=current_startup_id)
        .order_by("?")[:4]
    )
    similar_startups = similar_startups.annotate(
        average_rating_calc=Avg(
            models.ExpressionWrapper(
                models.F("sum_votes") * 1.0 / models.F("total_voters"),
                output_field=FloatField(),
            ),
            filter=models.Q(total_voters__gt=0),
        )
    ).annotate(average_rating=Coalesce("average_rating_calc", 0.0))
    html = render_to_string(
        "accounts/_similar_startup_cards.html",
        {"similar_startups": similar_startups, "request": request},
    )
    return HttpResponse(html)
@login_required
def investments(request):
    if not hasattr(request.user, "role") or request.user.role.role_name != "investor":
        messages.error(request, "Доступ к этой странице разрешен только инвесторам.")
        return redirect("profile")
    default_month_labels = [
        "Янв",
        "Фев",
        "Мар",
        "Апр",
        "Май",
        "Июн",
        "Июл",
        "Авг",
        "Сен",
        "Окт",
        "Ноя",
        "Дек",
    ]
    safe_context = {
        "startups_count": 0,
        "total_investment": Decimal("0"),
        "max_investment": Decimal("0"),
        "min_investment": Decimal("0"),
        "investment_categories": [],
        "month_labels": default_month_labels,
        "chart_monthly_category_data": [],
        "chart_categories": [],
        "all_directions": [],
        "invested_category_data": {},
        "user_investments": [],
        "user_owned_startups": [],
        "current_sort": "newest",
        "planetary_investments": [],
        "planetary_investments_json": [],
        "investor_logo_url": request.user.get_profile_picture_url() or "https://via.placeholder.com/60",
    }
    try:
        base_tx_qs = InvestmentTransactions.objects.filter(
            investor=request.user
        )
        user_investments_qs = (
            base_tx_qs.filter(amount__gt=0)
            .select_related("startup", "startup__direction", "startup__owner")
            .defer("franchise")
        )
        logger.info(
            f"[investments] tx count for {request.user.email}: {user_investments_qs.count()} (base={base_tx_qs.count()})"
        )
        total_investment_data = user_investments_qs.aggregate(
            total_investment=Sum("amount"),
            max_investment=Max("amount"),
            startups_count=Count("startup", distinct=True),
        )
        total_investment = total_investment_data.get("total_investment") or Decimal("0")
        max_investment = total_investment_data.get("max_investment") or Decimal("0")
        
        investments_with_amount = user_investments_qs.filter(amount__gt=0)
        min_investment_data = investments_with_amount.aggregate(
            min_investment=Min("amount")
        )
        min_investment = min_investment_data.get("min_investment") or Decimal("0")
        startups_count = total_investment_data.get("startups_count", 0)
        logger.info(
            f"[investments] User: {request.user.email}, Total Investment: {total_investment}"
        )
        category_data_raw = (
            user_investments_qs.values("startup__direction__direction_name")
            .annotate(category_total=Sum("amount"))
            .order_by("-category_total")
        )
        investment_categories = []
        invested_category_data_dict = {}
        total_for_category_percentage = (
            total_investment if total_investment > 0 else Decimal("1")
        )
        for cat_data in category_data_raw:
            percentage = 0
            category_sum = cat_data.get("category_total")
            category_name = (
                cat_data.get("startup__direction__direction_name") or "Без категории"
            )
            if category_sum and total_for_category_percentage > 0:
                try:
                    percentage = round(
                        (Decimal(category_sum) / total_for_category_percentage) * 100
                    )
                    percentage = min(percentage, 100)
                except Exception as e:
                    logger.error(
                        f"Ошибка расчета процента для категории '{category_name}': {e}"
                    )
                    percentage = 0
            investment_categories.append(
                {"name": category_name, "percentage": percentage}
            )
            invested_category_data_dict[category_name] = percentage
        end_dt = timezone.now()
        start_dt = (end_dt - relativedelta(months=11)).replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        logger.info(
            f"[investments] Preparing chart data for user {request.user.email}, range: {start_dt.date()}..{end_dt.date()}"
        )
        monthly_data_direct = (
            user_investments_qs.filter(created_at__date__gte=start_dt.date(), created_at__date__lte=end_dt.date())
            .annotate(month=TruncMonth("created_at"))
            .values("month")
            .annotate(monthly_total=Sum(Coalesce("amount", Decimal(0))))
            .order_by("month")
        )
        month_labels = []
        month_cursor = start_dt
        for _ in range(12):
            month_labels.append(month_cursor.strftime("%b"))
            month_cursor = month_cursor + relativedelta(months=1)
        monthly_totals = [0] * 12
        month_to_index = {}
        month_cursor = start_dt
        for idx in range(12):
            month_to_index[month_cursor.strftime("%Y-%m")] = idx
            month_cursor = month_cursor + relativedelta(months=1)
        for data in monthly_data_direct:
            month_key = data["month"].strftime("%Y-%m")
            idx = month_to_index.get(month_key)
            if idx is not None:
                monthly_total_decimal = data.get("monthly_total", Decimal(0)) or Decimal(0)
                monthly_totals[idx] = float(monthly_total_decimal)
        monthly_category_data_raw = (
            user_investments_qs.filter(
                created_at__date__gte=start_dt.date(),
                created_at__date__lte=end_dt.date(),
                startup__direction__isnull=False,
            )
            .annotate(month=TruncMonth("created_at"))
            .values("month", "startup__direction__direction_name")
            .annotate(monthly_category_total=Sum(Coalesce("amount", Decimal(0))))
            .order_by("month", "startup__direction__direction_name")
        )
        logger.info(
            f"[investments] Raw monthly category data from DB: {list(monthly_category_data_raw)}"
        )
        structured_monthly_data = collections.defaultdict(
            lambda: collections.defaultdict(float)
        )
        unique_categories = set()
        for data in monthly_category_data_raw:
            month_dt = data["month"]
            category_name = data["startup__direction__direction_name"]
            amount = float(data.get("monthly_category_total", 0) or 0)
            month_key = month_dt.strftime("%Y-%m-01")
            structured_monthly_data[month_key][category_name] += amount
            unique_categories.add(category_name)
        sorted_categories = sorted(list(unique_categories))
        logger.info(
            f"[investments] Unique categories found for chart: {sorted_categories}"
        )
        chart_data_list = []
        rolling_start = start_dt.date()
        for i in range(12):
            month_key = (rolling_start + relativedelta(months=i)).strftime("%Y-%m-01")
            month_data = {
                "month_key": month_key,
                "category_data": dict(structured_monthly_data[month_key]),
            }
            chart_data_list.append(month_data)
        logger.info(
            f"[investments] Final structured chart data list: {chart_data_list}"
        )
        try:
            s3_client = client(
                "s3",
                aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
                aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
                endpoint_url=settings.AWS_S3_ENDPOINT_URL,
                region_name=settings.AWS_S3_REGION_NAME,
            )
        except Exception as s3_init_err:
            logger.error(f"[investments] S3 client init failed: {s3_init_err}")
            s3_client = None
        invested_startups_qs = (
            user_investments_qs.select_related("startup").defer("franchise")
            .annotate(
                average_rating=Avg(
                    ExpressionWrapper(
                        Coalesce(F("startup__sum_votes"), 0)
                        * 1.0
                        / Coalesce(F("startup__total_voters"), 1),
                        output_field=FloatField(),
                    ),
                    filter=Q(startup__total_voters__gt=0),
                    default=0.0,
                ),
                comment_count=Count("startup__comments", distinct=True),
                investors_count=Count(
                    "startup__investmenttransactions__investor", distinct=True
                ),
            )
            .order_by("-amount")[:5]
        )
        owned_startups_qs = (
            Startups.objects.filter(owner_id=request.user.user_id, status="approved")
            .select_related("direction")
            .annotate(
                average_rating=Avg(
                    ExpressionWrapper(
                        Coalesce(F("sum_votes"), 0)
                        * 1.0
                        / Coalesce(F("total_voters"), 1),
                        output_field=FloatField(),
                    ),
                    filter=Q(total_voters__gt=0),
                    default=0.0,
                ),
                comment_count=Count("comments", distinct=True),
                investors_count=Count(
                    "investmenttransactions__investor", distinct=True
                ),
            )
            .order_by("-amount_raised")[:5]
        )
        planetary_investments = []
        min_orbit_size = 200
        max_orbit_size = 800
        orbit_step = 50
        available_sizes = list(
            range(min_orbit_size, max_orbit_size + orbit_step, orbit_step)
        )
        shuffle(available_sizes)
        for idx, startup in enumerate(
            list(invested_startups_qs) + list(owned_startups_qs), 1
        ):
            if hasattr(startup, "startup"):
                startup_obj = startup.startup
            else:
                startup_obj = startup
            if (
                not startup_obj.logo_urls
                or not isinstance(startup_obj.logo_urls, list)
                or len(startup_obj.logo_urls) == 0
            ):
                logger.warning(
                    f"Стартап {startup_obj.startup_id} ({startup_obj.title}) не имеет логотипа в logo_urls"
                )
                logo_url = "https://via.placeholder.com/150"
            else:
                if s3_client is not None:
                    try:
                        prefix = f"startups/{startup_obj.startup_id}/logos/"
                        response = s3_client.list_objects_v2(
                            Bucket=settings.AWS_STORAGE_BUCKET_NAME, Prefix=prefix
                        )
                        if "Contents" in response and len(response["Contents"]) > 0:
                            file_key = response["Contents"][0]["Key"]
                            logo_url = f"{settings.AWS_S3_ENDPOINT_URL}/{settings.AWS_STORAGE_BUCKET_NAME}/{file_key}"
                            logger.info(
                                f"Сгенерирован URL для логотипа стартапа {startup_obj.startup_id}: {logo_url}"
                            )
                        else:
                            logger.warning(
                                f"Файл для логотипа стартапа {startup_obj.startup_id} не найден в бакете по префиксу {prefix}"
                            )
                            logo_url = "https://via.placeholder.com/150"
                    except Exception as e:
                        logger.error(
                            f"Ошибка при генерации URL для логотипа стартапа {startup_obj.startup_id}: {str(e)}"
                        )
                        logo_url = "https://via.placeholder.com/150"
                else:
                    logo_url = "https://via.placeholder.com/150"
            orbit_size = (idx * 100) + 100
            orbit_time = (idx * 20) + 60
            planet_size = (idx * 2) + 50
            investment_type = (
                "Инвестирование"
                if startup_obj.only_invest
                else "Выкуп"
                if startup_obj.only_buy
                else "Выкуп+инвестирование"
                if startup_obj.both_mode
                else "Не указано"
            )
            planet_data = {
                "id": str(idx),
                "startup_id": startup_obj.startup_id,
                "name": startup_obj.title or "Без названия",
                "description": startup_obj.description or "Описание отсутствует",
                "rating": f"{(startup.average_rating or 0):.1f}/5 ({startup_obj.total_voters or 0})",
                "comment_count": startup.comment_count or 0,
                "progress": f"{(startup_obj.amount_raised / startup_obj.funding_goal * 100 if startup_obj.funding_goal else 0):.0f}%",
                "direction": startup_obj.direction.direction_name
                if startup_obj.direction
                else "Не указано",
                "investment_type": investment_type,
                "funding": f"{int(startup_obj.amount_raised or 0):,d} ₽".replace(
                    ",", " "
                ),
                "funding_goal": f"{int(startup_obj.funding_goal or 0):,d} ₽".replace(
                    ",", " "
                ),
                "investors": f"Инвесторов: {startup.investors_count or 0}",
                "image": logo_url,
                "orbit_size": orbit_size,
                "orbit_time": orbit_time,
                "planet_size": planet_size,
            }
            planetary_investments.append(planet_data)
        logger.info(
            f"[investments] Planetary investments for user {request.user.email}: {planetary_investments}"
        )
        user_investments = (
            user_investments_qs.select_related("startup")
            .annotate(
                startup_average_rating=Avg(
                    ExpressionWrapper(
                        F("startup__sum_votes") * 1.0 / F("startup__total_voters"),
                        output_field=FloatField(),
                    ),
                    filter=Q(startup__total_voters__gt=0),
                    default=0.0,
                ),
                startup_comment_count=Count("startup__comments", distinct=True),
            )
            .order_by("-created_at")
        )
        user_owned_startups = (
            Startups.objects.filter(owner_id=request.user.user_id)
            .select_related("direction", "stage", "status_id")
            .annotate(
                average_rating=Avg(
                    ExpressionWrapper(
                        Coalesce(F("sum_votes"), 0)
                        * 1.0
                        / Coalesce(F("total_voters"), 1),
                        output_field=FloatField(),
                    ),
                    filter=Q(total_voters__gt=0),
                    default=0.0,
                ),
                comment_count=Count("comments"),
            )
            .order_by("-created_at")
        )
        all_directions_qs = Directions.objects.all().order_by("direction_name")
        all_directions_list = list(all_directions_qs.values("pk", "direction_name"))
        context = {
            "startups_count": startups_count,
            "total_investment": total_investment,
            "max_investment": max_investment,
            "min_investment": min_investment,
            "investment_categories": investment_categories[:7],
            "month_labels": month_labels,
            "chart_monthly_category_data": chart_data_list,
            "chart_categories": sorted_categories,
            "all_directions": all_directions_list,
            "invested_category_data": invested_category_data_dict,
            "user_investments": user_investments,
            "user_owned_startups": user_owned_startups,
            "current_sort": "newest",
            "planetary_investments": planetary_investments,
            "planetary_investments_json": planetary_investments,
            "investor_logo_url": request.user.get_profile_picture_url()
            or "https://via.placeholder.com/60",
        }
        return render(request, "accounts/investments.html", context)
    except Exception as e:
        logger.error(f"Произошла ошибка в investments: {str(e)}", exc_info=True)
        try:
            user_investments_qs = InvestmentTransactions.objects.filter(
                investor=request.user, transaction_type__type_name__iexact="investment"
            ).select_related("startup", "startup__direction")
            total_investment_data = user_investments_qs.aggregate(
                total_investment=Sum("amount"),
                max_investment=Max("amount"),
                startups_count=Count("startup", distinct=True),
            )
            total_investment = total_investment_data.get("total_investment") or Decimal("0")
            max_investment = total_investment_data.get("max_investment") or Decimal("0")
            investments_with_amount = user_investments_qs.filter(amount__gt=0)
            min_investment = investments_with_amount.aggregate(m=Min("amount")).get("m") or Decimal("0")
            category_data_raw = (
                user_investments_qs.values("startup__direction__direction_name").annotate(category_total=Sum("amount")).order_by("-category_total")
            )
            investment_categories = []
            invested_category_data_dict = {}
            denom = total_investment if total_investment > 0 else Decimal("1")
            for c in category_data_raw:
                name = c.get("startup__direction__direction_name") or "Без категории"
                s = c.get("category_total") or Decimal("0")
                pct = int(round((Decimal(s) / denom) * 100)) if s else 0
                pct = max(0, min(pct, 100))
                investment_categories.append({"name": name, "percentage": pct})
                invested_category_data_dict[name] = pct
            end_dt = timezone.now()
            start_dt = (end_dt - relativedelta(months=11)).replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            month_labels = []
            month_cursor = start_dt
            for _ in range(12):
                month_labels.append(month_cursor.strftime("%b"))
                month_cursor = month_cursor + relativedelta(months=1)
            monthly_category_data_raw = (
                user_investments_qs.filter(created_at__date__gte=start_dt.date(), created_at__date__lte=end_dt.date(), amount__gt=0, startup__direction__isnull=False)
                .annotate(month=TruncMonth("created_at"))
                .values("month", "startup__direction__direction_name")
                .annotate(monthly_category_total=Sum(Coalesce("amount", Decimal(0))))
                .order_by("month", "startup__direction__direction_name")
            )
            structured = collections.defaultdict(lambda: collections.defaultdict(float))
            cats = set()
            for row in monthly_category_data_raw:
                mkey = row["month"].strftime("%Y-%m-01") if row.get("month") else None
                if not mkey:
                    continue
                cname = row["startup__direction__direction_name"]
                val = float(row.get("monthly_category_total") or 0)
                structured[mkey][cname] += val
                cats.add(cname)
            chart_data_list = []
            rolling_start = start_dt.date()
            for i in range(12):
                k = (rolling_start + relativedelta(months=i)).strftime("%Y-%m-01")
                chart_data_list.append({"month_key": k, "category_data": dict(structured[k])})
            sorted_categories = sorted(list(cats))
            context = {
                "startups_count": total_investment_data.get("startups_count", 0),
                "total_investment": total_investment,
                "max_investment": max_investment,
                "min_investment": min_investment,
                "investment_categories": investment_categories[:7],
                "month_labels": month_labels,
                "chart_monthly_category_data": chart_data_list,
                "chart_categories": sorted_categories,
                "all_directions": list(Directions.objects.values("pk", "direction_name")),
                "invested_category_data": invested_category_data_dict,
                "user_investments": user_investments_qs.order_by("-created_at")[:12],
                "user_owned_startups": Startups.objects.filter(owner_id=request.user.user_id)[:12],
                "current_sort": "newest",
                "planetary_investments": [],
                "planetary_investments_json": [],
                "investor_logo_url": request.user.get_profile_picture_url() or "https://via.placeholder.com/60",
            }
            return render(request, "accounts/investments.html", context)
        except Exception as e2:
            logger.error(f"[investments] Fallback building failed: {e2}", exc_info=True)
            return render(request, "accounts/investments.html", safe_context)
def legal(request):
    return render(request, "accounts/legal.html")
@login_required
def profile(request, user_id=None):
    if not user_id:
        user_id_param = request.GET.get("user_id")
        if user_id_param:
            try:
                user_id = int(user_id_param)
            except ValueError:
                user_id = None
    if user_id:
        user = get_object_or_404(Users, user_id=user_id)
        is_own_profile = request.user.user_id == user.user_id
    else:
        user = request.user
        is_own_profile = True
    if request.headers.get("x-requested-with") == "XMLHttpRequest" and request.method == "GET":
        user_data = {
            "user_id": user.user_id,
            "first_name": user.first_name,
            "last_name": user.last_name,
            "role": user.role.role_name if user.role else "",
            "profile_picture_url": user.get_profile_picture_url() if hasattr(user, "get_profile_picture_url") else "",
            "rating": getattr(user, "rating", None),
            "bio": getattr(user, "bio", ""),
        }
        return JsonResponse(user_data)
    show_role_selection = (not user.role_id or user.role_id == 4) and is_own_profile
    if request.method == "POST" and is_own_profile:
        if "select_role" in request.POST:
            role_id = request.POST.get("role_id")
            if role_id in ["1", "2"]:
                user.role_id = int(role_id)
                user.save(update_fields=["role"])
                messages.success(request, "Роль успешно выбрана!")
                return redirect("profile")
            else:
                messages.error(request, "Выбрана неверная роль.")
                return redirect("profile")
        elif "edit_profile" in request.POST:
            form = ProfileEditForm(request.POST, instance=user)
            if form.is_valid():
                form.save()
                if request.headers.get("x-requested-with") == "XMLHttpRequest":
                    return JsonResponse(
                        {"success": True, "message": "Профиль успешно обновлен!"}
                    )
                messages.success(request, "Профиль успешно обновлен!")
                return redirect("profile")
            else:
                if request.headers.get("x-requested-with") == "XMLHttpRequest":
                    return JsonResponse({"success": False, "errors": form.errors})
                messages.error(request, "Пожалуйста, исправьте ошибки.")
        elif "avatar" in request.FILES:
            user.profile_picture_url = request.FILES["avatar"]
            user.save(update_fields=["profile_picture_url"])
            messages.success(request, "Аватар успешно обновлен!")
            return redirect("profile")
    form = ProfileEditForm(instance=user)
    startups_list = Startups.objects.filter(owner=user).order_by("-created_at")
    startups_paginator = Paginator(startups_list, 5)
    startups_page_number = request.GET.get("startups_page")
    startups_page_obj = startups_paginator.get_page(startups_page_number)
    news_list = NewsArticles.objects.filter(author=user).order_by("-published_at")
    news_paginator = Paginator(news_list, 6)
    news_page_number = request.GET.get("news_page")
    news_page_obj = news_paginator.get_page(news_page_number)
    context = {
        "user": user,
        "is_own_profile": is_own_profile,
        "show_role_selection": show_role_selection,
        "form": form,
        "startups_page": startups_page_obj,
        "news_page": news_page_obj,
    }
    return render(request, "accounts/profile.html", context)
@login_required
def delete_avatar(request):
    if request.method == "POST":
        user = request.user
        if "avatar" in request.FILES:
            avatar = request.FILES["avatar"]
            allowed_mimes = ["image/jpeg", "image/png"]
            if avatar.content_type not in allowed_mimes:
                messages.error(request, "Допустимы только файлы PNG или JPEG.")
                if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                    return JsonResponse(
                        {
                            "success": False,
                            "error": "Допустимы только файлы PNG или JPEG.",
                        }
                    )
                return render(
                    request,
                    "accounts/profile.html",
                    {
                        "user": user,
                        "is_own_profile": True,
                        "form": form,
                        "startups_page": startups_page,
                        "news_page": news_page,
                        "show_role_selection": show_role_selection,
                    },
                )
            max_size = 5 * 1024 * 1024
            if avatar.size > max_size:
                messages.error(request, "Размер файла не должен превышать 5 МБ.")
                if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                    return JsonResponse(
                        {
                            "success": False,
                            "error": "Размер файла не должен превышать 5 МБ.",
                        }
                    )
                return render(
                    request,
                    "accounts/profile.html",
                    {
                        "user": user,
                        "is_own_profile": True,
                        "form": form,
                        "startups_page": startups_page,
                        "news_page": news_page,
                        "show_role_selection": show_role_selection,
                    },
                )
            avatar_id = str(uuid.uuid4())
            file_path = f"users/{request.user.user_id}/avatar/{avatar_id}_{avatar.name}"
            try:
                s3_client = boto3.client(
                    "s3",
                    endpoint_url=settings.AWS_S3_ENDPOINT_URL,
                    aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
                    aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
                    region_name=settings.AWS_S3_REGION_NAME,
                )
                bucket_name = settings.AWS_STORAGE_BUCKET_NAME
                prefix = f"users/{request.user.user_id}/avatar/"
                response = s3_client.list_objects_v2(Bucket=bucket_name, Prefix=prefix)
                if "Contents" in response:
                    for obj in response["Contents"]:
                        s3_client.delete_object(Bucket=bucket_name, Key=obj["Key"])
                        logger.info(f"Удалён старый аватар: {obj['Key']}")
                FileStorage.objects.filter(
                    entity_type__type_name="user",
                    entity_id=request.user.user_id,
                    file_type__type_name="avatar",
                ).delete()
                default_storage.save(file_path, avatar)
                request.user.profile_picture_url = avatar_id
                request.user.save()
                entity_type, _ = EntityTypes.objects.get_or_create(type_name="user")
                file_type, _ = FileTypes.objects.get_or_create(type_name="avatar")
                FileStorage.objects.create(
                    entity_type=entity_type,
                    entity_id=request.user.user_id,
                    file_url=avatar_id,
                    file_type=file_type,
                    uploaded_at=timezone.now(),
                )
                logger.info(
                    f"Аватар сохранён для user_id {request.user.user_id} по пути: {file_path}, UUID: {avatar_id}"
                )
                messages.success(request, "Аватарка успешно загружена!")
                if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                    return JsonResponse(
                        {"success": True, "message": "Аватарка успешно загружена!"}
                    )
            except Exception as e:
                logger.error(
                    f"Ошибка при сохранении аватара для user_id {request.user.user_id}: {str(e)}"
                )
                messages.error(request, "Ошибка при загрузке аватара.")
                if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                    return JsonResponse(
                        {"success": False, "error": "Ошибка при загрузке аватара."}
                    )
            return redirect("profile")
    return render(
        request,
        "accounts/profile.html",
        {
            "user": user,
            "is_own_profile": profile_user == request.user,
            "form": form,
            "startups_page": startups_page,
            "news_page": news_page,
            "show_role_selection": show_role_selection,
        },
    )
@login_required
def start_deal(request, chat_id):
    if request.method != "POST":
        return JsonResponse(
            {"success": False, "error": "Неверный метод запроса"}, status=405
        )
    chat = get_object_or_404(ChatConversations, conversation_id=chat_id)
    if not chat.chatparticipants_set.filter(user=request.user).exists():
        return JsonResponse(
            {"success": False, "error": "У вас нет доступа к этому чату"}, status=403
        )
    logger.info(
        f"Starting deal check for chat {chat_id}, participants: {chat.chatparticipants_set.count()}"
    )
    if chat.is_group_chat or chat.is_deal:
        logger.error(f"Chat {chat_id} is group or already a deal")
        return JsonResponse(
            {
                "success": False,
                "error": "Сделку можно начать только в личном чате, который ещё не является сделкой",
            },
            status=400,
        )
    participants = chat.chatparticipants_set.all()
    if (
        participants.count() < 2
    ):
        logger.error(
            f"Chat {chat_id} has {participants.count()} participants, expected at least 2"
        )
        return JsonResponse(
            {"success": False, "error": "В чате должно быть как минимум два участника"},
            status=400,
        )
    roles = {
        p.user.role.role_name.lower() for p in participants if p.user and p.user.role
    }
    if not {"startuper", "investor"}.issubset(roles):
        logger.error(
            f"Chat {chat_id} roles: {roles}, expected 'startuper' and 'investor'"
        )
        return JsonResponse(
            {
                "success": False,
                "error": "Чат должен включать одного стартапера и одного инвестора",
            },
            status=400,
        )
    try:
        data = json.loads(request.body)
        initiator_name = data.get(
            "initiator_name", request.user.get_full_name() or "Пользователь"
        )
    except json.JSONDecodeError:
        initiator_name = request.user.get_full_name() or "Пользователь"
    with transaction.atomic():
        chat.is_deal = True
        chat.deal_status = "pending"
        chat.updated_at = timezone.now()
        chat.save()
        moderators = Users.objects.filter(role__role_name="moderator")
        if not moderators.exists():
            return JsonResponse(
                {"success": False, "error": "Нет доступных модераторов"}, status=500
            )
        moderator = choice(list(moderators))
        moderator_participant, created = ChatParticipants.objects.get_or_create(
            conversation=chat, user=moderator
        )
        if not created and not moderator_participant:
            logger.error(
                f"Failed to create or find moderator {moderator.user_id} for chat {chat_id}"
            )
            return JsonResponse(
                {"success": False, "error": "Ошибка назначения модератора"}, status=500
            )
        logger.info(
            f"Moderator {moderator.user_id} added to chat {chat_id}, created: {created}"
        )
        message = Messages(
            conversation=chat,
            sender=None,
            message_text=f"Сделку начал {initiator_name}. Назначен модератор: {moderator.get_full_name()}",
            status=MessageStatuses.objects.get(status_name="sent"),
            created_at=timezone.now(),
            updated_at=timezone.now(),
        )
        message.save()
    participants_data = [
        {
            "user_id": p.user.user_id,
            "name": p.user.get_full_name(),
            "role": p.user.role.role_name if p.user.role else "unknown",
        }
        for p in chat.chatparticipants_set.all()
    ]
    logger.info(
        f"Сделка начата в чате {chat_id}, модератор {moderator.user_id} назначен"
    )
    return JsonResponse(
        {
            "success": True,
            "message": "Сделка начата, модератор назначен",
            "moderator": {
                "user_id": moderator.user_id,
                "name": moderator.get_full_name(),
            },
            "participants": participants_data,
        }
    )
@login_required
def deals_view(request):
    if not hasattr(request.user, "role") or (request.user.role.role_name or "").lower() != "moderator":
        messages.error(request, "Доступ к этой странице разрешен только модераторам.")
        logger.warning(
            f"Access denied for user {request.user.user_id} - not a moderator"
        )
        return redirect("home")
    status_filter = request.GET.get("status", "pending")
    valid_statuses = ["pending", "approved", "rejected"]
    if status_filter not in valid_statuses:
        status_filter = "pending"
    logger.info(
        f"Processing deals_view for user_id={request.user.user_id}, status_filter={status_filter}"
    )
    try:
        deals_query = (
            ChatConversations.objects.filter(is_deal=True, deal_status=status_filter)
            .prefetch_related("chatparticipants_set__user")
            .order_by("-updated_at")
        )
        logger.info(f"Initial query returned {deals_query.count()} deals")
    except Exception as e:
        logger.error(f"Error in initial query: {str(e)}")
        return JsonResponse({"error": f"Database query failed: {str(e)}"}, status=500)
    deals = deals_query.filter(chatparticipants__user=request.user)
    logger.info(f"Filtered deals for moderator {request.user.user_id}: {deals.count()}")
    for deal in deals:
        try:
            participants = deal.chatparticipants_set.all()
            logger.debug(
                f"Deal {deal.conversation_id}: Participants {[(p.user.user_id, p.user.role.role_name if p.user.role else 'None') for p in participants]}, Status: {deal.deal_status}"
            )
        except Exception as e:
            logger.error(f"Error processing deal {deal.conversation_id}: {str(e)}")
    deal_data = []
    selected_chat = None
    chat_id = request.GET.get("chat_id")
    if chat_id:
        try:
            selected_chat = get_object_or_404(
                ChatConversations, conversation_id=chat_id, is_deal=True
            )
            if not selected_chat.chatparticipants_set.filter(
                user=request.user
            ).exists():
                messages.error(request, "У вас нет доступа к этому чату.")
                logger.warning(
                    f"No access to chat {chat_id} for user {request.user.user_id}"
                )
                selected_chat = None
            else:
                messages = Messages.objects.filter(conversation=selected_chat).order_by(
                    "created_at"
                )
                messages_data = [
                    {
                        "message_id": msg.message_id,
                        "sender_name": msg.sender.get_full_name()
                        if msg.sender
                        else "Система",
                        "message_text": msg.message_text,
                        "created_at": msg.created_at.strftime("%H:%M %d/%m/%Y")
                        if msg.created_at
                        else "",
                        "is_own": msg.sender == request.user if msg.sender else False,
                    }
                    for msg in messages
                ]
                selected_chat_messages = messages_data
                logger.info(f"Loaded {len(messages_data)} messages for chat {chat_id}")
        except Exception as e:
            logger.error(f"Error loading chat {chat_id}: {str(e)}")
            messages.error(request, "Ошибка загрузки чата.")
            selected_chat = None
    for deal in deals:
        try:
            participants = deal.chatparticipants_set.all()
            moderator = next(
                (
                    p.user
                    for p in participants
                    if p.user.role and p.user.role.role_name == "moderator"
                ),
                None,
            )
            other_participants = [
                p.user for p in participants if p.user and p.user != moderator
            ]
            deal_data.append(
                {
                    "conversation_id": deal.conversation_id,
                    "name": deal.name or f"Сделка {deal.conversation_id}",
                    "participants": [
                        f"{p.first_name} {p.last_name}" for p in other_participants
                    ],
                    "moderator": moderator.get_full_name()
                    if moderator
                    else "Не назначен",
                    "last_message": deal.get_last_message().message_text
                    if deal.get_last_message()
                    else "Нет сообщений",
                    "created_at": deal.created_at.strftime("%H:%M")
                    if deal.created_at
                    else "",
                    "date": deal.created_at.strftime("%d/%m/%Y")
                    if deal.created_at
                    else "",
                    "unread_count": Messages.objects.filter(
                        conversation=deal, status__status_name="sent"
                    )
                    .exclude(sender=moderator)
                    .count()
                    if moderator
                    else 0,
                    "deal_status": deal.deal_status,
                }
            )
        except Exception as e:
            logger.error(
                f"Error processing deal data for {deal.conversation_id}: {str(e)}"
            )
    context = {
        "deals": deal_data,
        "current_status": status_filter,
        "selected_chat": selected_chat,
        "chat_messages": selected_chat_messages if selected_chat else [],
    }
    logger.info(f"Rendering deals.html with {len(deal_data)} deals")
    return render(request, "accounts/deals.html", context)
@login_required
def send_message(request):
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Неверный метод запроса"})
    form = MessageForm(request.POST)
    if not form.is_valid():
        return JsonResponse({"success": False, "error": "Неверные данные формы"})
    chat_id = request.POST.get("chat_id")
    chat = get_object_or_404(ChatConversations, conversation_id=chat_id)
    if not chat.chatparticipants_set.filter(user=request.user).exists():
        return JsonResponse(
            {"success": False, "error": "У вас нет доступа к этому чату"}
        )
    if not getattr(request.user, "role", None) or (request.user.role.role_name or "").lower() != "moderator":
        return JsonResponse(
            {
                "success": False,
                "error": "Только модератор может отправлять сообщения здесь",
            }
        )
    message = Messages(
        conversation=chat,
        sender=request.user,
        message_text=form.cleaned_data["message_text"],
        status=MessageStatuses.objects.get(status_name="sent"),
        created_at=timezone.now(),
        updated_at=timezone.now(),
    )
    message.save()
    chat.updated_at = timezone.now()
    chat.save()
    return JsonResponse(
        {
            "success": True,
            "message": {
                "message_id": message.message_id,
                "sender_name": request.user.get_full_name(),
                "message_text": message.message_text,
                "created_at": message.created_at.strftime("%H:%M %d/%m/%Y"),
                "is_own": True,
            },
        }
    )
@login_required
def approve_deal(request, chat_id):
    if request.method != "POST":
        return JsonResponse(
            {"success": False, "error": "Неверный метод запроса"}, status=405
        )
    chat = get_object_or_404(ChatConversations, conversation_id=chat_id)
    if not chat.chatparticipants_set.filter(user=request.user).exists() or (
        request.user.role and (request.user.role.role_name or "").lower() != "moderator"
    ):
        return JsonResponse(
            {"success": False, "error": "У вас нет прав для этого действия"}, status=403
        )
    if not chat.is_deal:
        return JsonResponse({"success": False, "error": "Это не сделка"}, status=400)
    with transaction.atomic():
        chat.deal_status = "approved"
        chat.updated_at = timezone.now()
        chat.save()
        message = Messages(
            conversation=chat,
            sender=None,
            message_text=f"Сделка #{chat.conversation_id} одобрена модератором {request.user.get_full_name()}",
            status=MessageStatuses.objects.get(status_name="sent"),
            created_at=timezone.now(),
            updated_at=timezone.now(),
        )
        message.save()
    logger.info(f"Сделка {chat_id} одобрена модератором {request.user.user_id}")
    return JsonResponse({"success": True, "message": "Сделка одобрена"})
@login_required
def reject_deal(request, chat_id):
    if request.method != "POST":
        return JsonResponse(
            {"success": False, "error": "Неверный метод запроса"}, status=405
        )
    chat = get_object_or_404(ChatConversations, conversation_id=chat_id)
    if not chat.chatparticipants_set.filter(user=request.user).exists() or (
        request.user.role and (request.user.role.role_name or "").lower() != "moderator"
    ):
        return JsonResponse(
            {"success": False, "error": "У вас нет прав для этого действия"}, status=403
        )
    if not chat.is_deal:
        return JsonResponse({"success": False, "error": "Это не сделка"}, status=400)
    with transaction.atomic():
        chat.deal_status = "rejected"
        chat.updated_at = timezone.now()
        chat.save()
        message = Messages(
            conversation=chat,
            sender=None,
            message_text=f"Сделка #{chat.conversation_id} отклонена модератором {request.user.get_full_name()}",
            status=MessageStatuses.objects.get(status_name="sent"),
            created_at=timezone.now(),
            updated_at=timezone.now(),
        )
        message.save()
    logger.info(f"Сделка {chat_id} отклонена модератором {request.user.user_id}")
    return JsonResponse({"success": True, "message": "Сделка отклонена"})
@login_required
def create_startup(request):
    allowed_roles = ["startuper", "moderator"]
    if not hasattr(request.user, "role") or request.user.role.role_name.lower() not in allowed_roles:
        messages.error(request, "Доступ к созданию стартапа разрешён только пользователям с ролью 'Стартаппер' или 'Модератор'.")
        return redirect("home")
    if request.method == "POST":
        form = StartupForm(request.POST, request.FILES)
        if form.is_valid():
            startup = form.save(commit=False)
            startup.owner = request.user
            startup.created_at = timezone.now()
            startup.updated_at = timezone.now()
            startup.status = "pending"
            try:
                startup.status_id = ReviewStatuses.objects.get(status_name="Pending")
            except ReviewStatuses.DoesNotExist:
                logger.error("Статус 'Pending' не найден в базе данных.")
                messages.error(request, "Статус 'Pending' не найден в базе данных.")
                return render(
                    request,
                    "accounts/create_startup.html",
                    {"form": form, "timeline_steps": request.POST},
                )
            investment_type = form.cleaned_data.get("investment_type")
            if investment_type == "invest":
                startup.only_invest = True
                startup.only_buy = False
                startup.both_mode = False
            elif investment_type == "buy":
                startup.only_invest = False
                startup.only_buy = True
                startup.both_mode = False
            elif investment_type == "both":
                startup.only_invest = False
                startup.only_buy = False
                startup.both_mode = True
            startup.step_number = int(request.POST.get("step_number", 1))
            startup.planet_image = form.cleaned_data.get("planet_image")
            logger.info("Сохранение стартапа перед обработкой файлов...")
            startup.save()
            logger.info(f"Стартап сохранен, startup_id: {startup.startup_id}")
            if not startup.startup_id:
                logger.error("Ошибка: startup_id не сгенерирован после сохранения!")
                messages.error(
                    request,
                    "Произошла ошибка при создании стартапа: ID не сгенерирован.",
                )
                return render(
                    request,
                    "accounts/create_startup.html",
                    {"form": form, "timeline_steps": request.POST},
                )
            for i in range(1, 6):
                description = request.POST.get(f"step_description_{i}", "").strip()
                if description:
                    StartupTimeline.objects.create(
                        startup=startup,
                        step_number=i,
                        title=f"Этап {i}",
                        description=description,
                    )
            logo_ids = []
            creatives_ids = []
            proofs_ids = []
            video_ids = []
            logo = form.cleaned_data.get("logo")
            if logo:
                logo_id = str(uuid.uuid4())
                base_name = os.path.splitext(logo.name)[0]
                ext = os.path.splitext(logo.name)[1]
                safe_base_name = "".join(
                    c for c in base_name if c.isalnum() or c in ("-", "_")
                )
                safe_name = slugify(safe_base_name) + ext
                file_path = f"startups/{startup.startup_id}/logos/{logo_id}_{safe_name}"
                logo_type, _ = FileTypes.objects.get_or_create(type_name="logo")
                entity_type, _ = EntityTypes.objects.get_or_create(type_name="startup")
                try:
                    logger.info(f"Попытка сохранить логотип по пути: {file_path}")
                    default_storage.save(file_path, logo)
                    logger.info(f"Логотип успешно сохранён по пути: {file_path}")
                    logo_ids.append(logo_id)
                    FileStorage.objects.create(
                        entity_type=entity_type,
                        entity_id=startup.startup_id,
                        file_type=logo_type,
                        file_url=logo_id,
                        uploaded_at=timezone.now(),
                        startup=startup,
                    )
                    logger.info(f"Логотип сохранён: {file_path}")
                except Exception as e:
                    logger.error(f"Ошибка сохранения логотипа: {e}", exc_info=True)
                    messages.warning(
                        request, "Не удалось сохранить логотип, но стартап создан."
                    )
            creatives = form.cleaned_data.get("creatives", [])
            if creatives:
                creative_type, _ = FileTypes.objects.get_or_create(type_name="creative")
                entity_type, _ = EntityTypes.objects.get_or_create(type_name="startup")
                for creative_file in creatives:
                    if not hasattr(creative_file, "name"):
                        logger.warning(f"Пропущен креатив: {creative_file}")
                        continue
                    unique_filename = get_unique_filename(creative_file.name, startup.startup_id, "creative")
                    creative_id = str(uuid.uuid4())
                    base_name = os.path.splitext(creative_file.name)[0]
                    ext = os.path.splitext(creative_file.name)[1]
                    safe_base_name = "".join(
                        c for c in base_name if c.isalnum() or c in ("-", "_")
                    )
                    safe_name = slugify(safe_base_name) + ext
                    file_path = f"startups/{startup.startup_id}/creatives/{creative_id}_{safe_name}"
                    try:
                        logger.info(f"Попытка сохранить креатив по пути: {file_path}")
                        default_storage.save(file_path, creative_file)
                        logger.info(f"Креатив успешно сохранён по пути: {file_path}")
                        creatives_ids.append(creative_id)
                        safe_create_file_storage(
                            entity_type=entity_type,
                            entity_id=startup.startup_id,
                            file_type=creative_type,
                            file_url=creative_id,
                            uploaded_at=timezone.now(),
                            startup=startup,
                            original_file_name=unique_filename,
                        )
                        logger.info(f"Креатив сохранён: {file_path}")
                    except Exception as e:
                        logger.error(f"Ошибка сохранения креатива: {e}", exc_info=True)
                        messages.warning(
                            request,
                            "Не удалось сохранить один из креативов, но стартап создан.",
                        )
            proofs = form.cleaned_data.get("proofs", [])
            if proofs:
                proof_type, _ = FileTypes.objects.get_or_create(type_name="proof")
                entity_type, _ = EntityTypes.objects.get_or_create(type_name="startup")
                for proof_file in proofs:
                    if not hasattr(proof_file, "name"):
                        logger.warning(f"Пропущен пруф: {proof_file}")
                        continue
                    unique_filename = get_unique_filename(proof_file.name, startup.startup_id, "proof")
                    proof_id = str(uuid.uuid4())
                    base_name = os.path.splitext(proof_file.name)[0]
                    ext = os.path.splitext(proof_file.name)[1]
                    safe_base_name = "".join(
                        c for c in base_name if c.isalnum() or c in ("-", "_")
                    )
                    safe_name = slugify(safe_base_name) + ext
                    file_path = (
                        f"startups/{startup.startup_id}/proofs/{proof_id}_{safe_name}"
                    )
                    try:
                        logger.info(f"Попытка сохранить пруф по пути: {file_path}")
                        default_storage.save(file_path, proof_file)
                        logger.info(f"Пруф успешно сохранён по пути: {file_path}")
                        proofs_ids.append(proof_id)
                        safe_create_file_storage(
                            entity_type=entity_type,
                            entity_id=startup.startup_id,
                            file_type=proof_type,
                            file_url=proof_id,
                            uploaded_at=timezone.now(),
                            startup=startup,
                            original_file_name=unique_filename,
                        )
                        logger.info(f"Пруф сохранён: {file_path}, оригинальное название: {unique_filename}")
                    except Exception as e:
                        logger.error(f"Ошибка сохранения пруфа: {e}", exc_info=True)
                        messages.warning(
                            request,
                            "Не удалось сохранить один из пруфов, но стартап создан.",
                        )
            video = form.cleaned_data.get("video")
            if video:
                unique_filename = get_unique_filename(video.name, startup.startup_id, "video")
                video_id = str(uuid.uuid4())
                base_name = os.path.splitext(video.name)[0]
                ext = os.path.splitext(video.name)[1]
                safe_base_name = "".join(
                    c for c in base_name if c.isalnum() or c in ("-", "_")
                )
                safe_name = slugify(safe_base_name) + ext
                file_path = (
                    f"startups/{startup.startup_id}/videos/{video_id}_{safe_name}"
                )
                video_type, _ = FileTypes.objects.get_or_create(type_name="video")
                entity_type, _ = EntityTypes.objects.get_or_create(type_name="startup")
                try:
                    logger.info(f"Попытка сохранить видео по пути: {file_path}")
                    default_storage.save(file_path, video)
                    logger.info(f"Видео успешно сохранено по пути: {file_path}")
                    video_ids.append(video_id)
                    safe_create_file_storage(
                        entity_type=entity_type,
                        entity_id=startup.startup_id,
                        file_type=video_type,
                        file_url=video_id,
                        uploaded_at=timezone.now(),
                        startup=startup,
                        original_file_name=unique_filename,
                    )
                    logger.info(f"Видео сохранено: {file_path}")
                except Exception as e:
                    logger.error(f"Ошибка сохранения видео: {e}", exc_info=True)
                    messages.warning(
                        request, "Не удалось сохранить видео, но стартап создан."
                    )
            startup.logo_urls = logo_ids
            startup.creatives_urls = creatives_ids
            startup.proofs_urls = proofs_ids
            startup.video_urls = video_ids
            startup.save()
            logger.info(
                f"Стартап создан: ID={startup.startup_id}, Planet={startup.planet_image}"
            )
            messages.success(
                request,
                f'Стартап "{startup.title}" успешно создан и отправлен на модерацию!',
            )
            return redirect("startup_creation_success")
        else:
            messages.error(request, "Форма содержит ошибки.")
            return render(
                request,
                "accounts/create_startup.html",
                {"form": form, "timeline_steps": request.POST},
            )
    else:
        form = StartupForm()
    return render(request, "accounts/create_startup.html", {"form": form})

@login_required
def create_franchise(request):
    allowed_roles = ["startuper", "moderator"]
    if not hasattr(request.user, "role") or request.user.role.role_name.lower() not in allowed_roles:
        messages.error(request, "Доступ к созданию франшизы разрешён только пользователям с ролью 'Стартаппер' или 'Модератор'.")
        return redirect("home")
    if request.method == "POST":
        form = FranchiseForm(request.POST, request.FILES)
        if form.is_valid():
            franchise = form.save(commit=False)
            franchise.owner = request.user
            franchise.created_at = timezone.now()
            franchise.updated_at = timezone.now()
            franchise.status = "pending"
            try:
                franchise.status_id = ReviewStatuses.objects.get(status_name="Pending")
            except ReviewStatuses.DoesNotExist:
                messages.error(request, "Статус 'Pending' не найден в базе данных.")
                return render(request, "accounts/create_franchise.html", {"form": form})
            franchise.planet_image = form.cleaned_data.get("planet_image")
            franchise.save()

            logo_ids, creatives_ids, proofs_ids, video_ids = [], [], [], []
            logo = form.cleaned_data.get("logo")
            if logo:
                logo_id = str(uuid.uuid4())
                base_name = os.path.splitext(logo.name)[0]
                ext = os.path.splitext(logo.name)[1]
                safe_base_name = "".join(c for c in base_name if c.isalnum() or c in ("-", "_"))
                safe_name = slugify(safe_base_name) + ext
                file_path = f"franchises/{franchise.franchise_id}/logos/{logo_id}_{safe_name}"
                logo_type, _ = FileTypes.objects.get_or_create(type_name="logo")
                entity_type, _ = EntityTypes.objects.get_or_create(type_name="franchise")
                try:
                    default_storage.save(file_path, logo)
                    logo_ids.append(logo_id)
                    safe_create_file_storage(
                        entity_type=entity_type,
                        entity_id=franchise.franchise_id,
                        file_type=logo_type,
                        file_url=logo_id,
                        uploaded_at=timezone.now(),
                        startup=None,
                        original_file_name=os.path.basename(file_path),
                    )
                except Exception:
                    messages.warning(request, "Не удалось сохранить логотип, но франшиза создана.")

            creatives = form.cleaned_data.get("creatives", [])
            if creatives:
                creative_type, _ = FileTypes.objects.get_or_create(type_name="creative")
                entity_type, _ = EntityTypes.objects.get_or_create(type_name="franchise")
                for creative_file in creatives:
                    if not hasattr(creative_file, "name"):
                        continue
                    creative_id = str(uuid.uuid4())
                    base_name = os.path.splitext(creative_file.name)[0]
                    ext = os.path.splitext(creative_file.name)[1]
                    safe_base_name = "".join(c for c in base_name if c.isalnum() or c in ("-", "_"))
                    safe_name = slugify(safe_base_name) + ext
                    file_path = f"franchises/{franchise.franchise_id}/creatives/{creative_id}_{safe_name}"
                    try:
                        default_storage.save(file_path, creative_file)
                        creatives_ids.append(creative_id)
                        safe_create_file_storage(
                            entity_type=entity_type,
                            entity_id=franchise.franchise_id,
                            file_type=creative_type,
                            file_url=creative_id,
                            uploaded_at=timezone.now(),
                            startup=None,
                            original_file_name=os.path.basename(file_path),
                        )
                    except Exception:
                        messages.warning(request, "Не удалось сохранить один из креативов, но франшиза создана.")

            proofs = form.cleaned_data.get("proofs", [])
            if proofs:
                proof_type, _ = FileTypes.objects.get_or_create(type_name="proof")
                entity_type, _ = EntityTypes.objects.get_or_create(type_name="franchise")
                for proof_file in proofs:
                    if not hasattr(proof_file, "name"):
                        continue
                    proof_id = str(uuid.uuid4())
                    base_name = os.path.splitext(proof_file.name)[0]
                    ext = os.path.splitext(proof_file.name)[1]
                    safe_base_name = "".join(c for c in base_name if c.isalnum() or c in ("-", "_"))
                    safe_name = slugify(safe_base_name) + ext
                    file_path = f"franchises/{franchise.franchise_id}/proofs/{proof_id}_{safe_name}"
                    try:
                        default_storage.save(file_path, proof_file)
                        proofs_ids.append(proof_id)
                        safe_create_file_storage(
                            entity_type=entity_type,
                            entity_id=franchise.franchise_id,
                            file_type=proof_type,
                            file_url=proof_id,
                            uploaded_at=timezone.now(),
                            startup=None,
                            original_file_name=os.path.basename(file_path),
                        )
                    except Exception:
                        messages.warning(request, "Не удалось сохранить один из документов, но франшиза создана.")

            video = form.cleaned_data.get("video")
            if video:
                video_type, _ = FileTypes.objects.get_or_create(type_name="video")
                entity_type, _ = EntityTypes.objects.get_or_create(type_name="franchise")
                video_id = str(uuid.uuid4())
                base_name = os.path.splitext(video.name)[0]
                ext = os.path.splitext(video.name)[1]
                safe_base_name = "".join(c for c in base_name if c.isalnum() or c in ("-", "_"))
                safe_name = slugify(safe_base_name) + ext
                file_path = f"franchises/{franchise.franchise_id}/videos/{video_id}_{safe_name}"
                try:
                    default_storage.save(file_path, video)
                    video_ids.append(video_id)
                    safe_create_file_storage(
                        entity_type=entity_type,
                        entity_id=franchise.franchise_id,
                        file_type=video_type,
                        file_url=video_id,
                        uploaded_at=timezone.now(),
                        startup=None,
                        original_file_name=os.path.basename(file_path),
                    )
                except Exception:
                    messages.warning(request, "Не удалось сохранить видео, но франшиза создана.")

            franchise.logo_urls = logo_ids
            franchise.creatives_urls = creatives_ids
            franchise.proofs_urls = proofs_ids
            franchise.video_urls = video_ids
            franchise.save()
            messages.success(request, f'Франшиза "{franchise.title}" успешно создана и отправлена на модерацию!')
            return redirect("franchises_list")
        else:
            messages.error(request, "Форма содержит ошибки.")
            return render(request, "accounts/create_franchise.html", {"form": form})
    else:
        form = FranchiseForm()
    return render(request, "accounts/create_franchise.html", {"form": form})

@login_required
def create_agency(request):
    allowed_roles = ["startuper", "moderator"]
    if not hasattr(request.user, "role") or request.user.role.role_name.lower() not in allowed_roles:
        messages.error(request, "Доступ к созданию агентства разрешён только пользователям с ролью 'Стартаппер' или 'Модератор'.")
        return redirect("home")
    if request.method == "POST":
        form = AgencyForm(request.POST, request.FILES)
        if form.is_valid():
            agency = form.save(commit=False)
            agency.owner = request.user
            agency.created_at = timezone.now()
            agency.updated_at = timezone.now()
            agency.status = "pending"
            agency.planet_image = form.cleaned_data.get("planet_image")
            agency.save()
            cat = form.cleaned_data.get("direction")
            if cat:
                data = agency.customization_data or {}
                data["agency_category"] = cat
                agency.customization_data = data
                agency.save(update_fields=["customization_data"])

            logo_ids, creatives_ids, proofs_ids, video_ids = [], [], [], []
            # Повторное использование логики сохранения файлов
            def save_file_set(files, type_name, subdir, ids_collector):
                file_type, _ = FileTypes.objects.get_or_create(type_name=type_name)
                entity_type, _ = EntityTypes.objects.get_or_create(type_name="agency")
                for f in files:
                    if not hasattr(f, "name"):
                        continue
                    file_id = str(uuid.uuid4())
                    base_name = os.path.splitext(f.name)[0]
                    ext = os.path.splitext(f.name)[1]
                    safe_base = "".join(c for c in base_name if c.isalnum() or c in ("-", "_"))
                    safe_name = slugify(safe_base) + ext
                    file_path = f"agencies/{agency.agency_id}/{subdir}/{file_id}_{safe_name}"
                    try:
                        default_storage.save(file_path, f)
                        ids_collector.append(file_id)
                        safe_create_file_storage(
                            entity_type=entity_type,
                            entity_id=agency.agency_id,
                            file_type=file_type,
                            file_url=file_id,
                            uploaded_at=timezone.now(),
                            startup=None,
                            original_file_name=os.path.basename(file_path),
                        )
                    except Exception:
                        pass

            logo = form.cleaned_data.get("logo")
            if logo:
                save_file_set([logo], "logo", "logos", logo_ids)
            save_file_set(form.cleaned_data.get("creatives", []), "creative", "creatives", creatives_ids)
            save_file_set(form.cleaned_data.get("proofs", []), "proof", "proofs", proofs_ids)
            video = form.cleaned_data.get("video")
            if video:
                save_file_set([video], "video", "videos", video_ids)

            agency.logo_urls = logo_ids
            agency.creatives_urls = creatives_ids
            agency.proofs_urls = proofs_ids
            agency.video_urls = video_ids
            agency.save()
            messages.success(request, f'Агентство "{agency.title}" успешно создано и отправлено на модерацию!')
            return redirect("agencies_list")
        else:
            messages.error(request, "Форма содержит ошибки.")
            return render(request, "accounts/create_agency.html", {"form": form})
    else:
        form = AgencyForm()
    return render(request, "accounts/create_agency.html", {"form": form})

@login_required
def create_specialist(request):
    allowed_roles = ["startuper", "moderator"]
    if not hasattr(request.user, "role") or request.user.role.role_name.lower() not in allowed_roles:
        messages.error(request, "Доступ к созданию профиля специалиста разрешён только пользователям с ролью 'Стартаппер' или 'Модератор'.")
        return redirect("home")
    if request.method == "POST":
        form = SpecialistForm(request.POST, request.FILES)
        if form.is_valid():
            spec = form.save(commit=False)
            spec.owner = request.user
            spec.created_at = timezone.now()
            spec.updated_at = timezone.now()
            spec.status = "pending"
            spec.planet_image = form.cleaned_data.get("planet_image")
            spec.save()
            cat = form.cleaned_data.get("direction")
            if cat:
                data = spec.customization_data or {}
                data["specialist_category"] = cat
                spec.customization_data = data
                spec.save(update_fields=["customization_data"])

            logo_ids, creatives_ids, proofs_ids, video_ids = [], [], [], []
            def save_file_set(files, type_name, subdir, ids_collector):
                file_type, _ = FileTypes.objects.get_or_create(type_name=type_name)
                entity_type, _ = EntityTypes.objects.get_or_create(type_name="specialist")
                for f in files:
                    if not hasattr(f, "name"):
                        continue
                    file_id = str(uuid.uuid4())
                    base_name = os.path.splitext(f.name)[0]
                    ext = os.path.splitext(f.name)[1]
                    safe_base = "".join(c for c in base_name if c.isalnum() or c in ("-", "_"))
                    safe_name = slugify(safe_base) + ext
                    file_path = f"specialists/{spec.specialist_id}/{subdir}/{file_id}_{safe_name}"
                    try:
                        default_storage.save(file_path, f)
                        ids_collector.append(file_id)
                        safe_create_file_storage(
                            entity_type=entity_type,
                            entity_id=spec.specialist_id,
                            file_type=file_type,
                            file_url=file_id,
                            uploaded_at=timezone.now(),
                            startup=None,
                            original_file_name=os.path.basename(file_path),
                        )
                    except Exception:
                        pass
            logo = form.cleaned_data.get("logo")
            if logo:
                save_file_set([logo], "logo", "logos", logo_ids)
            save_file_set(form.cleaned_data.get("creatives", []), "creative", "creatives", creatives_ids)
            save_file_set(form.cleaned_data.get("proofs", []), "proof", "proofs", proofs_ids)
            video = form.cleaned_data.get("video")
            if video:
                save_file_set([video], "video", "videos", video_ids)

            spec.logo_urls = logo_ids
            spec.creatives_urls = creatives_ids
            spec.proofs_urls = proofs_ids
            spec.video_urls = video_ids
            spec.save()
            messages.success(request, f'Профиль специалиста "{spec.title}" успешно создан и отправлен на модерацию!')
            return redirect("specialists_list")
        else:
            messages.error(request, "Форма содержит ошибки.")
            return render(request, "accounts/create_specialist.html", {"form": form})
    else:
        form = SpecialistForm()
    return render(request, "accounts/create_specialist.html", {"form": form})
@login_required
def startup_creation_success(request):
    return render(request, "accounts/startup_creation_success.html")
@login_required
def delete_message(request, message_id):
    message = get_object_or_404(Messages, message_id=message_id)
    chat = message.conversation
    if not chat.chatparticipants_set.filter(user=request.user).exists():
        return JsonResponse(
            {"success": False, "error": "У вас нет доступа к этому чату"}, status=403
        )
    if request.user.role and request.user.role.role_name.lower() == "moderator":
        message.is_deleted = True
        message.save()
        return JsonResponse({"success": True})
    return JsonResponse(
        {"success": False, "error": "Только модератор может удалить сообщение"},
        status=403,
    )
@login_required
def remove_participant(request, chat_id):
    chat = get_object_or_404(ChatConversations, conversation_id=chat_id)
    if not chat.chatparticipants_set.filter(user=request.user).exists():
        return JsonResponse(
            {"success": False, "error": "У вас нет доступа к этому чату"}, status=403
        )
    if request.method != "POST":
        return JsonResponse(
            {"success": False, "error": "Неверный метод запроса"}, status=405
        )
    user_id = request.POST.get("user_id")
    if not user_id:
        return JsonResponse(
            {"success": False, "error": "Не указан пользователь"}, status=400
        )
    try:
        user_to_remove = Users.objects.get(user_id=user_id)
    except Users.DoesNotExist:
        return JsonResponse(
            {"success": False, "error": "Пользователь не найден"}, status=404
        )
    if (
        request.user.role
        and request.user.role.role_name.lower() == "moderator"
        and chat.is_group_chat
    ):
        participant = chat.chatparticipants_set.filter(user=user_to_remove).first()
        if participant:
            participant.delete()
            chat.updated_at = timezone.now()
            chat.save()
            return JsonResponse({"success": True})
    return JsonResponse(
        {
            "success": False,
            "error": "Только модератор может исключить участника из группового чата",
        },
        status=403,
    )
@login_required
def edit_startup(request, startup_id):
    logger.debug(f"Request method: {request.method}")
    logger.debug(f"Request POST: {request.POST}")
    logger.debug(f"Request FILES: {dict(request.FILES)}")
    startup = get_object_or_404(Startups, startup_id=startup_id)
    if not (
        request.user == startup.owner
        or (
            hasattr(request.user, "role")
            and request.user.role
            and request.user.role.role_name == "moderator"
        )
    ):
        messages.error(request, "У вас нет прав для редактирования этого стартапа.")
        return redirect("startup_detail", startup_id=startup_id)
    timeline = StartupTimeline.objects.filter(startup=startup)
    timeline_steps = timeline
    if request.method == "POST":
        form = StartupForm(request.POST, request.FILES, instance=startup)
        if form.is_valid():
            startup = form.save(commit=False)
            startup.status = "pending"
            startup.is_edited = True
            startup.updated_at = timezone.now()
            if "step_number" in request.POST:
                new_step = int(request.POST.get("step_number"))
                startup.step_number = new_step
            investment_type = form.cleaned_data.get("investment_type")
            if investment_type == "invest":
                startup.only_invest = True
                startup.only_buy = False
                startup.both_mode = False
            elif investment_type == "buy":
                startup.only_invest = False
                startup.only_buy = True
                startup.both_mode = False
            elif investment_type == "both":
                startup.only_invest = False
                startup.only_buy = False
                startup.both_mode = True
            startup.save()
            deleted_files_json = request.POST.get('deleted_files', '[]')
            try:
                deleted_files = json.loads(deleted_files_json)
                for deleted_file in deleted_files:
                    file_id = deleted_file.get('id')
                    file_type = deleted_file.get('type')
                    if file_id and file_type:
                        FileStorage.objects.filter(
                            startup=startup,
                            file_url=file_id
                        ).delete()
                        if file_type == 'creative' and startup.creatives_urls:
                            startup.creatives_urls = [url for url in startup.creatives_urls if url != file_id]
                        elif file_type == 'proof' and startup.proofs_urls:
                            startup.proofs_urls = [url for url in startup.proofs_urls if url != file_id]
                        elif file_type == 'video' and startup.video_urls:
                            startup.video_urls = [url for url in startup.video_urls if url != file_id]
                        logger.info(f"Удален файл {file_type}: {file_id}")
            except json.JSONDecodeError:
                logger.warning("Ошибка при разборе deleted_files JSON")
            for i in range(1, 6):
                description = request.POST.get(f"step_description_{i}", "").strip()
                if description:
                    timeline_entry, created = StartupTimeline.objects.get_or_create(
                        startup=startup,
                        step_number=i,
                        defaults={"title": f"Этап {i}", "description": description},
                    )
                    if not created and timeline_entry.description != description:
                        timeline_entry.description = description
                        timeline_entry.save()
            logo_ids = startup.logo_urls or []
            creatives_ids = startup.creatives_urls or []
            proofs_ids = startup.proofs_urls or []
            video_ids = startup.video_urls or []
            logo = form.cleaned_data.get("logo")
            if logo:
                logo_id = str(uuid.uuid4())
                file_path = f"startups/{startup.startup_id}/logos/{logo_id}_{logo.name}"
                default_storage.save(file_path, logo)
                logo_ids = [logo_id]
                logger.info(f"Логотип сохранён с ID: {logo_id}")
            creatives = form.cleaned_data.get("creatives", [])
            if creatives:
                creative_type = FileTypes.objects.get(type_name="creative")
                entity_type = EntityTypes.objects.get(type_name="startup")
                creatives_ids = []
                for creative_file in creatives:
                    if not hasattr(creative_file, "name"):
                        logger.warning(
                            f"Пропущен креатив, так как это не файл: {creative_file}"
                        )
                        continue
                    unique_filename = get_unique_filename(creative_file.name, startup.startup_id, "creative")
                    creative_id = str(uuid.uuid4())
                    file_path = f"startups/{startup.startup_id}/creatives/{creative_id}_{creative_file.name}"
                    default_storage.save(file_path, creative_file)
                    creatives_ids.append(creative_id)
                    safe_create_file_storage_instance(
                        entity_type=entity_type,
                        entity_id=startup.startup_id,
                        file_type=creative_type,
                        file_url=creative_id,
                        uploaded_at=timezone.now(),
                        startup=startup,
                        original_file_name=unique_filename,
                    )
                    logger.info(f"Креатив сохранён с ID: {creative_id}")
            proofs = form.cleaned_data.get("proofs", [])
            if proofs:
                proof_type = FileTypes.objects.get(type_name="proof")
                entity_type = EntityTypes.objects.get(type_name="startup")
                proofs_ids = []
                for proof_file in proofs:
                    if not hasattr(proof_file, "name"):
                        logger.warning(
                            f"Пропущен пруф, так как это не файл: {proof_file}"
                        )
                        continue
                    unique_filename = get_unique_filename(proof_file.name, startup.startup_id, "proof")
                    proof_id = str(uuid.uuid4())
                    file_path = f"startups/{startup.startup_id}/proofs/{proof_id}_{proof_file.name}"
                    default_storage.save(file_path, proof_file)
                    proofs_ids.append(proof_id)
                    safe_create_file_storage_instance(
                        entity_type=entity_type,
                        entity_id=startup.startup_id,
                        file_type=proof_type,
                        file_url=proof_id,
                        uploaded_at=timezone.now(),
                        startup=startup,
                        original_file_name=unique_filename,
                    )
                    logger.info(f"Пруф сохранён с ID: {proof_id}")
            video = form.cleaned_data.get("video")
            if video:
                unique_filename = get_unique_filename(video.name, startup.startup_id, "video")
                video_id = str(uuid.uuid4())
                file_path = (
                    f"startups/{startup.startup_id}/videos/{video_id}_{video.name}"
                )
                default_storage.save(file_path, video)
                video_ids = [video_id]
                video_type, _ = FileTypes.objects.get_or_create(type_name="video")
                entity_type = EntityTypes.objects.get(type_name="startup")
                safe_create_file_storage_instance(
                    entity_type=entity_type,
                    entity_id=startup.startup_id,
                    file_type=video_type,
                    file_url=video_id,
                    uploaded_at=timezone.now(),
                    startup=startup,
                    original_file_name=unique_filename,
                )
                logger.info(f"Видео сохранено с ID: {video_id}")
            startup.logo_urls = logo_ids
            startup.creatives_urls = creatives_ids
            startup.proofs_urls = proofs_ids
            startup.video_urls = video_ids
            startup.save()
            logger.info("=== Обновление стартапа ===")
            logger.info(f"Стартап ID: {startup.startup_id}")
            if logo:
                logger.info(f"Логотип: {logo.name}, размер: {logo.size} байт")
                logger.info(
                    f"ID логотипа: {logo_ids[0] if logo_ids else 'Не сохранён'}"
                )
            else:
                logger.info("Логотип не загружен")
            if creatives:
                logger.info(f"Креативы: {len(creatives)} файлов")
                for i, creative_file in enumerate(creatives, 1):
                    if hasattr(creative_file, "name"):
                        logger.info(
                            f"Креатив {i}: {creative_file.name}, размер: {creative_file.size} байт"
                        )
                    else:
                        logger.info(
                            f"Креатив {i}: Неверный формат (не файл): {creative_file}"
                        )
            else:
                logger.info("Креативы не загружены")
            if proofs:
                logger.info(f"Пруфы: {len(proofs)} файлов")
                for i, proof_file in enumerate(proofs, 1):
                    if hasattr(proof_file, "name"):
                        logger.info(
                            f"Пруф {i}: {proof_file.name}, размер: {proof_file.size} байт"
                        )
                    else:
                        logger.info(
                            f"Пруф {i}: Неверный формат (не файл): {proof_file}"
                        )
            else:
                logger.info("Пруфы не загружены")
            if video:
                logger.info(f"Видео: {video.name}, размер: {video.size} байт")
                logger.info(f"ID видео: {video_ids[0] if video_ids else 'Не сохранён'}")
            else:
                logger.info("Видео не загружено")
            logger.info("=== Переменные окружения ===")
            for key, value in os.environ.items():
                logger.info(f"{key}: {value}")
            logger.info("=== Настройки Yandex Object Storage ===")
            logger.info(
                f"AWS_ACCESS_KEY_ID: {getattr(settings, 'AWS_ACCESS_KEY_ID', 'Не задано')}"
            )
            logger.info(
                f"AWS_SECRET_ACCESS_KEY: {getattr(settings, 'AWS_SECRET_ACCESS_KEY', 'Не задано')}"
            )
            logger.info(
                f"AWS_STORAGE_BUCKET_NAME: {getattr(settings, 'AWS_STORAGE_BUCKET_NAME', 'Не задано')}"
            )
            logger.info(
                f"AWS_S3_ENDPOINT_URL: {getattr(settings, 'AWS_S3_ENDPOINT_URL', 'Не задано')}"
            )
            logger.info(
                f"AWS_DEFAULT_ACL: {getattr(settings, 'AWS_DEFAULT_ACL', 'Не задано')}"
            )
            logger.info("=== Проверка STORAGES ===")
            logger.info(
                f"STORAGES['default']['BACKEND']: {settings.STORAGES['default']['BACKEND']}"
            )
            logger.info(f"default_storage: {default_storage.__class__.__name__}")
            logger.info("=== Проверка подключения к Yandex Object Storage ===")
            try:
                from django.core.files.base import ContentFile
                from storages.backends.s3boto3 import S3Boto3Storage
                storage = S3Boto3Storage()
                test_file_name = f"test/test_file_{startup.startup_id}.txt"
                test_content = (
                    "This is a test file to check Yandex Object Storage connection."
                )
                test_file = ContentFile(test_content.encode("utf-8"))
                storage.save(test_file_name, test_file)
                logger.info(f"Тестовый файл успешно сохранён: {test_file_name}")
                test_file_url = storage.url(test_file_name)
                logger.info(f"URL тестового файла: {test_file_url}")
                storage.delete(test_file_name)
                logger.info(f"Тестовый файл удалён: {test_file_name}")
            except Exception as e:
                logger.error(
                    f"Ошибка подключения к Yandex Object Storage: {str(e)}",
                    exc_info=True,
                )
            messages.success(
                request,
                f'Стартап "{startup.title}" успешно отредактирован и отправлен на модерацию!',
            )
            return redirect("profile")
        else:
            messages.error(request, "Форма содержит ошибки.")
            return render(
                request,
                "accounts/edit_startup.html",
                {
                    "form": form,
                    "startup": startup,
                    "timeline_steps": timeline_steps,
                },
            )
    else:
        form = StartupForm(instance=startup)
    return render(
        request,
        "accounts/edit_startup.html",
        {
            "form": form,
            "startup": startup,
            "timeline_steps": timeline_steps,
        },
    )
@login_required
def main_page_moderator(request):
    """
    Отображает главную страницу для модератора.
    """
    if not getattr(request.user, "role", None) or (request.user.role.role_name or "").lower() != "moderator":
        return redirect("home")
    return render(request, "accounts/moderator_main.html")
@login_required
def investor_main(request):
    """
    Отображает главную страницу инвестора с планетарной системой стартапов.
    """
    directions_data_json = FIXED_CATEGORIES.copy()
    selected_direction_name = request.GET.get("direction", "All")
    startups_query = Startups.objects.filter(status="approved").annotate(
        rating_avg=Coalesce(Avg("uservotes__rating"), 0.0, output_field=FloatField()),
        voters_count=Count("uservotes", distinct=True),
        total_investors=Count("investmenttransactions", distinct=True),
        current_funding=Coalesce(
            Sum("investmenttransactions__amount"), 0, output_field=DecimalField()
        ),
        comment_count=Count("comments", distinct=True),
    )
    if selected_direction_name != "All" and selected_direction_name != "Все":
        from django.db.models import Q
        direction_filter = Q()
        for category in FIXED_CATEGORIES:
            if category['original_name'] == selected_direction_name or category['direction_name'] == selected_direction_name:
                direction_filter |= Q(direction__direction_name=category['direction_name'])
        if direction_filter:
            startups_query = startups_query.filter(direction_filter)
    startups_filtered = startups_query.annotate(
        progress=Case(
            When(funding_goal__gt=0, then=(F("amount_raised") * 100.0 / F("funding_goal"))),
            default=Value(0),
            output_field=FloatField(),
        )
    )[:6]
    planets_data_for_template = []
    fixed_orbit_sizes = [200, 300, 400, 500, 600, 700]
    orbit_times = [80, 95, 110, 125, 140, 160]
    planet_sizes = [60, 70, 56, 64, 50, 60]
    import random
    for idx, startup in enumerate(startups_filtered):
        folder_choice = random.choice(['planets_round', 'planets_ring'])
        if folder_choice == 'planets_round':
            random_planet_num = random.randint(1, 15)
            image_path = f"accounts/images/planetary_system/planets_round/{random_planet_num}.png"
        else:  # planets_ring
            random_planet_num = random.randint(1, 6)
            image_path = f"accounts/images/planetary_system/planets_ring/{random_planet_num}.png"
        planets_data_for_template.append(
            {
                "id": startup.startup_id,
                "image": static(image_path),
                "orbit_size": fixed_orbit_sizes[idx % len(fixed_orbit_sizes)],
                "orbit_time": orbit_times[idx % len(orbit_times)],
                "planet_size": planet_sizes[idx % len(planet_sizes)],
            }
        )
    planets_data_json = []
    for startup in startups_filtered:
        investment_type = (
            "Инвестирование"
            if startup.only_invest
            else "Выкуп"
            if startup.only_buy
            else "Выкуп+инвестирование"
            if startup.both_mode
            else "Не указано"
        )
        random_planet_num = random.randint(1, 8)
        planet_image_url = static(f"accounts/images/planetary_system/planets_round/{random_planet_num}.png")
        planets_data_json.append({
            "id": startup.startup_id,
            "name": startup.title,
            "image": planet_image_url,
            "rating": round(startup.rating_avg, 2),
            "progress": f"{startup.progress:.2f}%" if startup.progress is not None else "0%",
            "direction": startup.direction.direction_name if startup.direction else "Не указано",
            "investors": startup.total_investors,
            "funding_goal": f"{startup.funding_goal:,.0f} ₽".replace(",", " ") if startup.funding_goal else "Не определена",
            "comment_count": startup.comment_count,
            "startup_id": startup.startup_id,
            "description": startup.short_description,
            "investment_type": investment_type,
        })
    is_authenticated = request.user.is_authenticated
    is_startuper = is_authenticated and hasattr(request.user, 'role') and request.user.role and request.user.role.role_name == 'startuper'
    logo_data = {"image": static("accounts/images/planetary_system/gi.svg")}
    all_startups_query = Startups.objects.filter(status="approved").annotate(
        rating_avg=Coalesce(Avg("uservotes__rating"), 0.0, output_field=FloatField()),
        voters_count=Count("uservotes", distinct=True),
        total_investors=Count("investmenttransactions", distinct=True),
        current_funding=Coalesce(
            Sum("investmenttransactions__amount"), 0, output_field=DecimalField()
        ),
        comment_count=Count("comments", distinct=True),
        progress=Case(
            When(funding_goal__gt=0, then=(F("amount_raised") * 100.0 / F("funding_goal"))),
            default=Value(0),
            output_field=FloatField(),
        )
    )
    all_startups_data = []
    for startup in all_startups_query:
        investment_type = (
            "Инвестирование"
            if startup.only_invest
            else "Выкуп"
            if startup.only_buy
            else "Выкуп+инвестирование"
            if startup.both_mode
            else "Не указано"
        )
        folder_choice = random.choice(['planets_round', 'planets_ring'])
        if folder_choice == 'planets_round':
            random_planet_num = random.randint(1, 15)
            planet_image_url = static(f"accounts/images/planetary_system/planets_round/{random_planet_num}.png")
        else:  # planets_ring
            random_planet_num = random.randint(1, 6)
            planet_image_url = static(f"accounts/images/planetary_system/planets_ring/{random_planet_num}.png")
        direction_name = startup.direction.direction_name if startup.direction else "Не указано"
        original_direction = None
        for category in FIXED_CATEGORIES:
            if category['direction_name'] == direction_name:
                original_direction = category['original_name']
                break
        if not original_direction:
            original_direction = direction_name  # Если не найдено, используем как есть
        all_startups_data.append({
            "id": startup.startup_id,
            "name": startup.title,
            "image": planet_image_url,
            "rating": round(startup.rating_avg, 2),
            "voters_count": startup.voters_count,
            "progress": round(startup.progress, 2) if startup.progress is not None else 0,
            "direction": original_direction,  # Используем original_name для фильтрации
            "investors": startup.total_investors,
            "funding_goal": f"{startup.funding_goal:,.0f} ₽".replace(",", " ") if startup.funding_goal else "Не определена",
            "valuation": f"{startup.valuation:,.0f} ₽".replace(",", " ") if startup.valuation else "Не указана",
            "comment_count": startup.comment_count,
            "startup_id": startup.startup_id,
            "description": startup.short_description,
            "investment_type": investment_type,
        })
    context = {
        "planets_data": planets_data_for_template,
        "logo_data": logo_data,
        "directions": directions_data_json,
        "selected_galaxy": selected_direction_name,
        "planets_data_json": json.dumps(planets_data_json, cls=DjangoJSONEncoder),
        "directions_data_json": json.dumps(directions_data_json, cls=DjangoJSONEncoder),
        "all_startups_data_json": json.dumps(all_startups_data, cls=DjangoJSONEncoder),
        "is_startuper": is_startuper,
    }
    return render(request, "accounts/investor_main.html", context)
@login_required
def startuper_main(request):
    """
    Отображает главную страницу стартаппера с планетарной системой стартапов.
    """
    directions_data_json = FIXED_CATEGORIES.copy()
    selected_direction_name = request.GET.get("direction", "All")
    print(f"🔍 STARTUPPER_MAIN: Запрошено направление: '{selected_direction_name}'")
    startups_query = Startups.objects.filter(status="approved").annotate(
        rating_avg=Coalesce(Avg("uservotes__rating"), 0.0, output_field=FloatField()),
        voters_count=Count("uservotes", distinct=True),
        total_investors=Count("investmenttransactions", distinct=True),
        current_funding=Coalesce(
            Sum("investmenttransactions__amount"), 0, output_field=DecimalField()
        ),
        comment_count=Count("comments", distinct=True),
    )
    print(f"🔍 STARTUPPER_MAIN: Всего одобренных стартапов: {startups_query.count()}")
    
    from accounts.models import Directions
    all_directions = Directions.objects.all()
    print(f"🔍 STARTUPPER_MAIN: Все направления в БД:")
    for direction in all_directions:
        print(f"🔍   - {direction.direction_name}")
    
    if selected_direction_name != "All" and selected_direction_name != "Все":
        from django.db.models import Q
        direction_filter = Q()
        print(f"🔍 STARTUPPER_MAIN: Ищем направление '{selected_direction_name}'")
        for category in FIXED_CATEGORIES:
            if category['original_name'] == selected_direction_name or category['direction_name'] == selected_direction_name:
                direction_filter |= Q(direction__direction_name=category['direction_name'])
                print(f"🔍 STARTUPPER_MAIN: Найдено соответствие: {category['original_name']} -> {category['direction_name']}")
        if direction_filter:
            startups_query = startups_query.filter(direction_filter)
            print(f"🔍 STARTUPPER_MAIN: Применен фильтр, найдено стартапов: {startups_query.count()}")
        else:
            print(f"🔍 STARTUPPER_MAIN: Фильтр не найден для '{selected_direction_name}'")
    
    all_startups = Startups.objects.filter(status="approved").select_related("direction")
    print(f"🔍 STARTUPPER_MAIN: Все одобренные стартапы:")
    for startup in all_startups[:5]:  # Показываем первые 5
        direction_name = startup.direction.direction_name if startup.direction else "Нет направления"
        print(f"🔍   - {startup.title} -> {direction_name}")
    
    startups_filtered = startups_query.annotate(
        progress=Case(
            When(funding_goal__gt=0, then=(F("amount_raised") * 100.0 / F("funding_goal"))),
            default=Value(0),
            output_field=FloatField(),
        )
    )[:6]
    planets_data_for_template = []
    fixed_orbit_sizes = [200, 300, 400, 500, 600, 700]
    orbit_times = [80, 95, 110, 125, 140, 160]
    planet_sizes = [60, 70, 56, 64, 50, 60]
    import random
    for idx, startup in enumerate(startups_filtered):
        folder_choice = random.choice(['planets_round', 'planets_ring'])
        if folder_choice == 'planets_round':
            random_planet_num = random.randint(1, 15)
            image_path = f"accounts/images/planetary_system/planets_round/{random_planet_num}.png"
        else:  # planets_ring
            random_planet_num = random.randint(1, 6)
            image_path = f"accounts/images/planetary_system/planets_ring/{random_planet_num}.png"
        planets_data_for_template.append(
            {
                "id": startup.startup_id,
                "image": static(image_path),
                "orbit_size": fixed_orbit_sizes[idx % len(fixed_orbit_sizes)],
                "orbit_time": orbit_times[idx % len(orbit_times)],
                "planet_size": planet_sizes[idx % len(planet_sizes)],
            }
        )
    planets_data_json = []
    for startup in startups_filtered:
        investment_type = (
            "Инвестирование"
            if startup.only_invest
            else "Выкуп"
            if startup.only_buy
            else "Выкуп+инвестирование"
            if startup.both_mode
            else "Не указано"
        )
        folder_choice = random.choice(['planets_round', 'planets_ring'])
        if folder_choice == 'planets_round':
            random_planet_num = random.randint(1, 15)
            planet_image_url = static(f"accounts/images/planetary_system/planets_round/{random_planet_num}.png")
        else:  # planets_ring
            random_planet_num = random.randint(1, 6)
            planet_image_url = static(f"accounts/images/planetary_system/planets_ring/{random_planet_num}.png")
        planets_data_json.append({
            "id": startup.startup_id,
            "name": startup.title,
            "image": planet_image_url,
            "rating": round(startup.rating_avg, 2),
            "progress": f"{startup.progress:.2f}%" if startup.progress is not None else "0%",
            "direction": startup.direction.direction_name if startup.direction else "Не указано",
            "investors": startup.total_investors,
            "funding_goal": f"{startup.funding_goal:,.0f} ₽".replace(",", " ") if startup.funding_goal else "Не определена",
            "comment_count": startup.comment_count,
            "startup_id": startup.startup_id,
            "description": startup.short_description,
            "investment_type": investment_type,
        })
    is_authenticated = request.user.is_authenticated
    is_startuper = is_authenticated and hasattr(request.user, 'role') and request.user.role and request.user.role.role_name == 'startuper'
    logo_data = {"image": static("accounts/images/planetary_system/gi.svg")}
    all_startups_query = Startups.objects.filter(status="approved").annotate(
        rating_avg=Coalesce(Avg("uservotes__rating"), 0.0, output_field=FloatField()),
        voters_count=Count("uservotes", distinct=True),
        total_investors=Count("investmenttransactions", distinct=True),
        current_funding=Coalesce(
            Sum("investmenttransactions__amount"), 0, output_field=DecimalField()
        ),
        comment_count=Count("comments", distinct=True),
        progress=Case(
            When(funding_goal__gt=0, then=(F("amount_raised") * 100.0 / F("funding_goal"))),
            default=Value(0),
            output_field=FloatField(),
        )
    )
    all_startups_data = []
    for startup in all_startups_query:
        investment_type = (
            "Инвестирование"
            if startup.only_invest
            else "Выкуп"
            if startup.only_buy
            else "Выкуп+инвестирование"
            if startup.both_mode
            else "Не указано"
        )
        folder_choice = random.choice(['planets_round', 'planets_ring'])
        if folder_choice == 'planets_round':
            random_planet_num = random.randint(1, 15)
            planet_image_url = static(f"accounts/images/planetary_system/planets_round/{random_planet_num}.png")
        else:  # planets_ring
            random_planet_num = random.randint(1, 6)
            planet_image_url = static(f"accounts/images/planetary_system/planets_ring/{random_planet_num}.png")
        direction_name = startup.direction.direction_name if startup.direction else "Не указано"
        russian_direction = DIRECTION_TRANSLATIONS.get(direction_name, direction_name)
        original_direction = None
        for category in FIXED_CATEGORIES:
            if category['direction_name'] == direction_name:
                original_direction = category['original_name']
                break
        if not original_direction:
            original_direction = direction_name  # Если не найдено, используем как есть
        all_startups_data.append({
            "id": startup.startup_id,
            "name": startup.title,
            "image": planet_image_url,
            "rating": round(startup.rating_avg, 2),
            "voters_count": startup.voters_count,
            "progress": round(startup.progress, 2) if startup.progress is not None else 0,
            "direction": original_direction,  # Используем original_name для фильтрации
            "investors": startup.total_investors,
            "funding_goal": f"{startup.funding_goal:,.0f} ₽".replace(",", " ") if startup.funding_goal else "Не определена",
            "valuation": f"{startup.valuation:,.0f} ₽".replace(",", " ") if startup.valuation else "Не указана",
            "comment_count": startup.comment_count,
            "startup_id": startup.startup_id,
            "description": startup.short_description,
            "investment_type": investment_type,
        })
    # Получаем 3 случайных стартапа для блока "Нет идей ДЛЯ СТАРТАПА?"
    try:
        random_startups = Startups.objects.filter(status="approved").order_by('?')[:3]
        random_startups_data = []
        
        for startup in random_startups:
            # Определяем изображение стартапа
            if startup.planet_image:
                startup_image = f"{settings.S3_PUBLIC_BASE_URL}/choosable_planets/{startup.planet_image}"
            else:
                startup_image = static('accounts/images/main_page/volt_forge.webp')  # fallback изображение
            
            # Получаем аватар владельца
            if hasattr(startup, 'owner') and startup.owner and hasattr(startup.owner, 'get_profile_picture_url'):
                owner_avatar = startup.owner.get_profile_picture_url() or static('accounts/images/default_icon.svg')
            else:
                owner_avatar = static('accounts/images/default_icon.svg')
            
            # Форматируем рейтинг
            rating = getattr(startup, 'rating_avg', 0.0)
            if rating:
                rating_formatted = f"{rating:.1f}/5"
            else:
                rating_formatted = "0.0/5"
            
            # Обрезаем описание
            description = getattr(startup, 'short_description', '') or getattr(startup, 'description', '')
            if description:
                if len(description) > 100:
                    description = description[:100] + "..."
            else:
                description = "Описание не указано"
            
            # Проверяем, что startup_id является валидным числом
            startup_id = getattr(startup, 'startup_id', None)
            if startup_id and str(startup_id).isdigit():
                startup_url = f"/startups/{startup_id}/"
            else:
                startup_url = "/startups_list/"
            
            startup_data = {
                'id': startup_id or 'Unknown',
                'name': getattr(startup, 'title', 'Unknown'),
                'rating': rating_formatted,
                'description': description,
                'image': startup_image,
                'owner_avatar': owner_avatar,
                'url': startup_url
            }
            random_startups_data.append(startup_data)
            
    except Exception as e:
        logger.error(f"Error getting random startups for startuper_main: {e}")
        # Fallback данные
        random_startups_data = [
            {
                'id': 'fallback1',
                'name': 'VoltForge Dynamics',
                'rating': '4.4/5',
                'description': 'VoltForge разрабатывает твердотельные батареи с графеновыми наноструктурами, которые заряжаются...',
                'image': static('accounts/images/main_page/volt_forge.webp'),
                'owner_avatar': static('accounts/images/default_icon.svg'),
                'url': '/startups_list/'
            },
            {
                'id': 'fallback2',
                'name': 'NeuroBloom',
                'rating': '4.7/5',
                'description': 'NeuroBloom предлагает носимый гаджет с ИИ, который анализирует нейронные паттерны для раннего выявления тревоги, депрессии и выгорания.',
                'image': static('accounts/images/main_page/neuro_bloom.webp'),
                'owner_avatar': static('accounts/images/default_icon.svg'),
                'url': '/startups_list/'
            },
            {
                'id': 'fallback3',
                'name': 'BioCrop Nexus',
                'rating': '4.2/5',
                'description': 'BioCrop Nexus создает генетически оптимизированные семена, устойчивые к экстремальным климатическим условиям и вредителям.',
                'image': static('accounts/images/main_page/biocrop_nexus.webp'),
                'owner_avatar': static('accounts/images/default_icon.svg'),
                'url': '/startups_list/'
            }
        ]

    # Получаем 3 случайных стартапера для блока "Последние обновления от стартаперов 🔥"
    try:
        random_startupers = Users.objects.filter(role__role_name='startuper').order_by('?')[:3]
        random_startupers_data = []
        
        for startuper in random_startupers:
            # Получаем аватар стартапера
            if hasattr(startuper, 'get_profile_picture_url'):
                avatar_url = startuper.get_profile_picture_url() or static('accounts/images/avatars/default_avatar_ufo.png')
            else:
                avatar_url = static('accounts/images/avatars/default_avatar_ufo.png')
            
            # Форматируем рейтинг
            rating = getattr(startuper, 'rating_avg', 0.0)
            if rating:
                rating_formatted = f"{rating:.1f}/5"
            else:
                rating_formatted = "0.0/5"
            
            # Получаем имя стартапера
            first_name = getattr(startuper, 'first_name', '') or ''
            last_name = getattr(startuper, 'last_name', '') or ''
            if first_name and last_name:
                full_name = f"{first_name} {last_name}"
            elif first_name:
                full_name = first_name
            elif last_name:
                full_name = last_name
            else:
                full_name = "Стартапер"
            
            startuper_data = {
                'id': getattr(startuper, 'user_id', 'Unknown'),
                'name': full_name,
                'rating': rating_formatted,
                'avatar': avatar_url
            }
            random_startupers_data.append(startuper_data)
            
    except Exception as e:
        logger.error(f"Error getting random startupers for startuper_main: {e}")
        # Fallback данные
        random_startupers_data = [
            {
                'id': 'fallback1',
                'name': 'Виктор Смирнов',
                'rating': '4.5/5',
                'avatar': static('accounts/images/avatars/default_avatar_ufo.png')
            },
            {
                'id': 'fallback2',
                'name': 'Анна Кузнецова',
                'rating': '4.9/5',
                'avatar': static('accounts/images/avatars/default_avatar_ufo.png')
            },
            {
                'id': 'fallback3',
                'name': 'Дмитрий Иванов',
                'rating': '4.3/5',
                'avatar': static('accounts/images/avatars/default_avatar_ufo.png')
            }
        ]

    context = {
        "planets_data": planets_data_for_template,
        "logo_data": logo_data,
        "directions": directions_data_json,
        "selected_galaxy": selected_direction_name,
        "planets_data_json": json.dumps(planets_data_json, cls=DjangoJSONEncoder),
        "directions_data_json": json.dumps(directions_data_json, cls=DjangoJSONEncoder),
        "all_startups_data_json": json.dumps(all_startups_data, cls=DjangoJSONEncoder),
        "is_startuper": is_startuper,
        "random_startups": random_startups_data,
        "random_startupers": random_startupers_data,
    }
    
    print(f"🔍 STARTUPPER_MAIN: Передаем в шаблон:")
    print(f"🔍   - directions: {directions_data_json}")
    print(f"🔍   - selected_galaxy: {selected_direction_name}")
    print(f"🔍   - planets_data_json: {len(planets_data_json)} элементов")
    print(f"🔍   - all_startups_data_json: {len(all_startups_data)} элементов")
    
    print(f"🔍 STARTUPPER_MAIN: Первые 3 стартапа из all_startups_data:")
    for i, startup in enumerate(all_startups_data[:3]):
        print(f"🔍   {i+1}. {startup.get('name', 'Нет названия')} -> direction: {startup.get('direction', 'Нет направления')}")
    
    return render(request, "accounts/startuper_main.html", context)
def moderator_dashboard(request):
    pending_startups_list = Startups.objects.filter(status="pending")
    pending_franchises_list = Franchises.objects.filter(status="pending")
    pending_agencies_list = Agencies.objects.filter(status="pending")
    pending_specialists_list = Specialists.objects.filter(status="pending")
    all_categories = Directions.objects.all().order_by("direction_name")
    selected_category_name = request.GET.get("category")
    sort_order = request.GET.get("sort")
    filter_type = request.GET.get("filter")
    if filter_type == "all":
        selected_category_name = None
        sort_order = None
    if selected_category_name:
        pending_startups_list = pending_startups_list.filter(
            direction__direction_name__iexact=selected_category_name
        )
    if sort_order == "newest":
        if hasattr(Startups, "created_at"):
            pending_startups_list = pending_startups_list.order_by("-created_at")
        else:
            pending_startups_list = pending_startups_list.order_by(
                "-startup_id"
            )
    else:
        if hasattr(Startups, "created_at"):
            pending_startups_list = pending_startups_list.order_by(
                "-created_at"
            )
        else:
            pending_startups_list = pending_startups_list.order_by("-startup_id")
    context = {
        "pending_startups": pending_startups_list,
        "pending_franchises": pending_franchises_list,
        "pending_agencies": pending_agencies_list,
        "pending_specialists": pending_specialists_list,
        "all_categories": all_categories,
        "selected_category_name": selected_category_name,
        "current_sort_order": sort_order,
        "filter_type": filter_type,
    }
    return render(request, "accounts/moderator_dashboard.html", context)
def approve_startup(request, startup_id):
    if not request.user.is_authenticated or (request.user.role.role_name or "").lower() != "moderator":
        messages.error(request, "У вас нет прав для этого действия.")
        return redirect("home")
    startup = get_object_or_404(Startups, startup_id=startup_id)
    if request.method == "POST":
        moderator_comment = request.POST.get("moderator_comment", "")
        startup.moderator_comment = moderator_comment
        startup.status = "approved"
        try:
            startup.status_id = ReviewStatuses.objects.get(status_name="Approved")
        except ReviewStatuses.DoesNotExist:
            raise ValueError("Статус 'Approved' не найден в базе данных.")
        startup.save()
        messages.success(request, "Стартап одобрен.")
    return redirect("moderator_dashboard")
def reject_startup(request, startup_id):
    if not request.user.is_authenticated or (request.user.role.role_name or "").lower() != "moderator":
        messages.error(request, "У вас нет прав для этого действия.")
        return redirect("home")
    startup = get_object_or_404(Startups, startup_id=startup_id)
    if request.method == "POST":
        moderator_comment = request.POST.get("moderator_comment", "")
        startup.moderator_comment = moderator_comment
        startup.status = "rejected"
        try:
            startup.status_id = ReviewStatuses.objects.get(status_name="Rejected")
        except ReviewStatuses.DoesNotExist:
            raise ValueError("Статус 'Rejected' не найден в базе данных.")
        startup.save()
        messages.success(request, "Стартап отклонен.")
    return redirect("moderator_dashboard")
@login_required
def vote_startup(request, startup_id):
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Неверный метод запроса"})
    startup = get_object_or_404(Startups, startup_id=startup_id)
    rating = int(request.POST.get("rating", 0))
    if not 1 <= rating <= 5:
        return JsonResponse(
            {"success": False, "error": "Недопустимое значение рейтинга"}
        )
    if UserVotes.objects.filter(user=request.user, startup=startup).exists():
        return JsonResponse(
            {"success": False, "error": "Вы уже голосовали за этот стартап"}
        )
    UserVotes.objects.create(
        user=request.user, startup=startup, rating=rating, created_at=timezone.now()
    )
    startup.total_voters += 1
    startup.sum_votes += rating
    startup.save()
    average_rating = (
        startup.sum_votes / startup.total_voters if startup.total_voters > 0 else 0
    )
    return JsonResponse({"success": True, "average_rating": average_rating})
@login_required
def invest(request, startup_id):
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Неверный метод запроса"})
    startup = get_object_or_404(Startups, startup_id=startup_id)
    if not request.user.is_authenticated or request.user.role.role_name != "investor":
        return JsonResponse(
            {"success": False, "error": "Только инвесторы могут инвестировать"}
        )
    if startup.status in ["blocked", "closed"]:
        return JsonResponse(
            {
                "success": False,
                "error": f"Инвестирование запрещено: стартап {startup.status}",
            }
        )
    try:
        amount = Decimal(request.POST.get("amount", "0"))
        if amount <= 0:
            return JsonResponse(
                {"success": False, "error": "Сумма должна быть больше 0"}
            )
        transaction = InvestmentTransactions(
            startup=startup,
            investor=request.user,
            amount=amount,
            is_micro=startup.micro_investment_available,
            transaction_type=TransactionTypes.objects.get(type_name="investment"),
            transaction_status="completed",
            payment_method=PaymentMethods.objects.get(method_name="default"),
            created_at=timezone.now(),
            updated_at=timezone.now(),
        )
        transaction.save()
        startup.amount_raised = (startup.amount_raised or Decimal("0")) + amount
        startup.total_invested = (startup.total_invested or Decimal("0")) + amount
        startup.save()
        investors_count = startup.get_investors_count()
        progress_percentage = startup.get_progress_percentage()
        return JsonResponse(
            {
                "success": True,
                "amount_raised": float(startup.amount_raised),
                "investors_count": investors_count,
                "progress_percentage": float(progress_percentage),
            }
        )
    except Exception as e:
        logger.error(f"Ошибка при инвестировании: {str(e)}")
        return JsonResponse(
            {"success": False, "error": "Произошла ошибка при инвестировании"}
        )
class NewsForm(forms.Form):
    title = forms.CharField(max_length=255, label="Заголовок")
    content = forms.CharField(widget=forms.Textarea, label="Текст новости")
    image = forms.ImageField(label="Картинка", required=False)
def news(request):
    if request.method == "POST":
        if (
            not request.user.is_authenticated
            or (request.user.role.role_name or "").lower() != "moderator"
        ):
            return JsonResponse(
                {"success": False, "error": "У вас нет прав для этого действия."}
            )
        form = NewsForm(request.POST, request.FILES)
        if form.is_valid():
            article = NewsArticles(
                title=form.cleaned_data["title"],
                content=form.cleaned_data["content"],
                author=request.user,
                published_at=timezone.now(),
                updated_at=timezone.now(),
                tags="Администрация",
            )
            article.save()
            image = form.cleaned_data.get("image")
            if image:
                image_id = str(uuid.uuid4())
                file_path = f"news/{article.article_id}/{image_id}_{image.name}"
                default_storage.save(file_path, image)
                article.image_url = file_path
                article.save()
            return JsonResponse({"success": True})
        else:
            return JsonResponse({"success": False, "error": "Форма содержит ошибки."})
    
    # Получаем параметры фильтрации
    articles = NewsArticles.objects.all()
    
    # Сортировка
    sort_order = request.GET.get("sort", "new")
    if sort_order == "old":
        articles = articles.order_by("published_at")
    elif sort_order == "rating":
        articles = articles.order_by("-rating_agg", "-published_at")
    else:  # new
        articles = articles.order_by("-published_at")
    
    # Фильтрация по категориям
    selected_categories = request.GET.getlist("category")
    if selected_categories:
        # Здесь можно добавить логику фильтрации по категориям
        # Пока оставляем как есть
        pass
    
    # Фильтрация по микроинвестициям
    micro_investment = request.GET.get("micro_investment")
    if micro_investment == "1":
        # Здесь можно добавить логику фильтрации по микроинвестициям
        # Пока оставляем как есть
        pass
    
    # Фильтрация по рейтингу
    min_rating = request.GET.get("min_rating")
    max_rating = request.GET.get("max_rating")
    if min_rating:
        try:
            min_rating_val = float(min_rating)
            articles = articles.filter(rating_agg__gte=min_rating_val)
        except ValueError:
            pass
    
    if max_rating:
        try:
            max_rating_val = float(max_rating)
            articles = articles.filter(rating_agg__lte=max_rating_val)
        except ValueError:
            pass
    
    # Поиск
    search_query = request.GET.get("search")
    if search_query:
        articles = articles.filter(
            Q(title__icontains=search_query) | 
            Q(content__icontains=search_query) |
            Q(tags__icontains=search_query)
        )
    
    # Добавляем пагинацию для новостей
    try:
        page_number = int(request.GET.get("page", 1))
        if page_number < 1:
            page_number = 1
    except (ValueError, TypeError):
        page_number = 1
    
    paginator = Paginator(articles, 12)  # 12 новостей на страницу
    
    # Проверяем, что номер страницы не превышает общее количество страниц
    if page_number > paginator.num_pages and paginator.num_pages > 0:
        page_number = paginator.num_pages
    
    page_obj = paginator.get_page(page_number)
    
    context = {
        "articles": page_obj,
        "page_obj": page_obj,
        "paginator": paginator,
        "sort_order": sort_order,
        "selected_categories": selected_categories,
        "micro_investment": micro_investment == "1",
        "min_rating": min_rating,
        "max_rating": max_rating,
        "search_query": search_query,
    }
    
    is_ajax = request.headers.get("x-requested-with") == "XMLHttpRequest"
    if is_ajax:
        html = render_to_string("accounts/partials/_news_cards.html", context, request=request)
        return JsonResponse({"html": html})
    
    return render(request, "accounts/news.html", context)
def news_detail(request, article_id):
    article = get_object_or_404(NewsArticles, article_id=article_id)
    user = request.user if request.user.is_authenticated else None
    if not NewsViews.objects.filter(article=article, user=user).exists():
        NewsViews.objects.create(article=article, user=user, viewed_at=timezone.now())
    views_count = NewsViews.objects.filter(article=article).count()
    likes_count = NewsLikes.objects.filter(article=article).count()
    user_liked = (
        NewsLikes.objects.filter(article=article, user=user).exists() if user else False
    )
    if (
        request.method == "POST"
        and request.user.is_authenticated
        and "like" in request.POST
    ):
        if not user_liked:
            NewsLikes.objects.create(
                article=article, user=request.user, created_at=timezone.now()
            )
            likes_count += 1
            user_liked = True
    return render(
        request,
        "accounts/news_detail.html",
        {
            "article": article,
            "views_count": views_count,
            "likes_count": likes_count,
            "user_liked": user_liked,
        },
    )
@login_required
def create_news(request):
    if not request.user.is_authenticated or (request.user.role.role_name or "").lower() != "moderator":
        messages.error(request, "У вас нет прав для этого действия.")
        return redirect("news")
    if request.method == "POST":
        form = NewsForm(request.POST, request.FILES)
        if form.is_valid():
            article = NewsArticles(
                title=form.cleaned_data["title"],
                content=form.cleaned_data["content"],
                author=request.user,
                published_at=timezone.now(),
                updated_at=timezone.now(),
                tags="Администрация",
            )
            image = form.cleaned_data.get("image")
            if image:
                image_id = str(uuid.uuid4())
                file_path = f"news/{image_id}_{image.name}"
                default_storage.save(file_path, image)
                article.image_url = file_path
            article.save()
            messages.success(request, "Новость успешно создана!")
            return redirect("news")
    else:
        form = NewsForm()
    return render(request, "accounts/create_news.html", {"form": form})
def delete_news(request, article_id):
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Неверный метод запроса"})
    if not request.user.is_authenticated or (request.user.role.role_name or "").lower() != "moderator":
        return JsonResponse(
            {"success": False, "error": "У вас нет прав для этого действия."}
        )
    article = get_object_or_404(NewsArticles, article_id=article_id)
    if article.image_url:
        try:
            default_storage.delete(article.image_url)
        except Exception as e:
            logger.error(f"Ошибка при удалении картинки новости {article_id}: {str(e)}")
    article.delete()
    return JsonResponse({"success": True})
@login_required
def cosmochat(request):
    if not request.user.is_authenticated:
        if request.headers.get("x-requested-with") == "XMLHttpRequest":
            return JsonResponse(
                {"success": False, "error": "Требуется авторизация"}, status=401
            )
        messages.error(
            request, "Пожалуйста, войдите в систему, чтобы получить доступ к чату."
        )
        return redirect("login")
    chats = (
        ChatConversations.objects.filter(chatparticipants__user=request.user)
        .prefetch_related(
            "chatparticipants_set__user"
        )
        .annotate(
            latest_message_time=Max("messages__created_at")
        )
        .order_by(F("latest_message_time").desc(nulls_last=True), "-updated_at")
    )
    for chat in chats:
        if chat.is_group_chat:
            chat.display_name = chat.name
            chat.display_avatar = None
        else:
            other_participant = None
            for p in chat.chatparticipants_set.all():
                if p.user_id != request.user.user_id:
                    other_participant = p
                    break
            if other_participant and other_participant.user:
                user_profile = other_participant.user
                chat.display_name = f"{user_profile.first_name or ''} {user_profile.last_name or ''}".strip()
                chat.display_avatar = user_profile.get_profile_picture_url()
            else:
                chat.display_name = "Удаленный чат"
                chat.display_avatar = None
    search_form = UserSearchForm(request.GET)
    users = Users.objects.all()
    if search_form.is_valid():
        query = search_form.cleaned_data.get("query", "")
        roles = search_form.cleaned_data.get("roles", [])
        if query:
            users = users.filter(
                Q(email__icontains=query)
                | Q(first_name__icontains=query)
                | Q(last_name__icontains=query)
            )
        if roles:
            users = users.filter(role__role_name__in=roles)
    users = users.exclude(user_id=request.user.user_id)
    chat_id = request.GET.get("chat_id")
    if chat_id:
        chat = ChatConversations.objects.filter(conversation_id=chat_id).first()
        if chat:
            participant_ids = chat.chatparticipants_set.values_list(
                "user_id", flat=True
            )
            users = users.exclude(user_id__in=participant_ids)
        else:
            if request.headers.get("x-requested-with") == "XMLHttpRequest":
                return JsonResponse(
                    {"success": False, "error": "Чат не найден"}, status=404
                )
    for user in users[:5]:
        profile_url = (
            user.get_profile_picture_url() if user.profile_picture_url else "None"
        )
        logger.info(
            f"Cosmochat User ID: {user.user_id}, Profile Picture URL: {user.profile_picture_url}, Generated URL: {profile_url}"
        )
    for chat in chats[:5]:
        participants = chat.chatparticipants_set.all()
        participant_info = [
            f"ID: {p.user.user_id}, Picture: {p.user.get_profile_picture_url() or 'None'}"
            for p in participants
            if p.user and p.user != request.user
        ]
        logger.info(
            f"Chat ID: {chat.conversation_id}, Participants (excluding self): {participant_info}"
        )
    message_form = MessageForm()
    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        users_data = [
            {
                "user_id": user.user_id,
                "name": f"{user.first_name} {user.last_name}",
                "role": user.role.role_name if user.role else "Система",
            }
            for user in users
        ]
        return JsonResponse({"users": users_data})
    return render(
        request,
        "accounts/cosmochat.html",
        {
            "search_form": search_form,
            "users": users,
            "chats": chats,
            "message_form": message_form,
        },
    )

def chat_list(request):
    """API endpoint для получения списка чатов пользователя"""
    if not request.user.is_authenticated:
        return JsonResponse({"success": False, "error": "Требуется авторизация"}, status=401)
    
    try:
        chats = (
            ChatConversations.objects.filter(chatparticipants__user=request.user)
            .prefetch_related(
                "chatparticipants_set__user",
                "messages_set"
            )
            .annotate(
                latest_message_time=Max("messages__created_at"),
                has_messages=Count("messages"),
                chat_created_at=F("created_at")
            )
            .order_by(
                "-has_messages",  # Сначала чаты с сообщениями
                F("latest_message_time").desc(nulls_last=True),  # Затем по времени последнего сообщения
                "-updated_at",  # И по времени обновления
                "-chat_created_at"  # Новые чаты сверху
            )
        )
        
        chats_data = []
        for chat in chats:
            if chat.is_group_chat:
                try:
                    chat_name = chat.name or "Групповой чат"
                except:
                    chat_name = "Групповой чат"
                
                try:
                    conversation_id = chat.conversation_id or 0
                except:
                    conversation_id = 0
                
                try:
                    created_at = chat.created_at.isoformat() if chat.created_at else None
                except:
                    created_at = None
                
                try:
                    updated_at = chat.updated_at.isoformat() if chat.updated_at else None
                except:
                    updated_at = None
                
                try:
                    has_messages = chat.has_messages or 0
                except:
                    has_messages = 0
                
                try:
                    latest_message_time = chat.latest_message_time.isoformat() if chat.latest_message_time else None
                except:
                    latest_message_time = None
                
                chat_data = {
                    "conversation_id": conversation_id,
                    "name": chat_name,
                    "is_group_chat": True,
                    "is_deleted": False,
                    "has_left": False,
                    "is_deal": False,
                    "latest_message": None,
                    "unread_count": 0,
                    "created_at": created_at,
                    "updated_at": updated_at,
                    "has_messages": has_messages,
                    "latest_message_time": latest_message_time
                }
            else:
                other_participant = None
                for p in chat.chatparticipants_set.all():
                    if p.user_id != request.user.user_id:
                        other_participant = p
                        break
                
                if other_participant and other_participant.user:
                    user_profile = other_participant.user
                    try:
                        user_name = f"{user_profile.first_name or ''} {user_profile.last_name or ''}".strip() or user_profile.email
                    except:
                        try:
                            user_id = user_profile.user_id or 0
                            user_name = user_profile.email or f"Пользователь {user_id}"
                        except:
                            try:
                                user_id = user_profile.user_id or 0
                            except:
                                user_id = 0
                            user_name = f"Пользователь {user_id}"
                    
                    try:
                        conversation_id = chat.conversation_id or 0
                    except:
                        conversation_id = 0
                    
                    try:
                        created_at = chat.created_at.isoformat() if chat.created_at else None
                    except:
                        created_at = None
                    
                    try:
                        updated_at = chat.updated_at.isoformat() if chat.updated_at else None
                    except:
                        updated_at = None
                    
                    try:
                        has_messages = chat.has_messages or 0
                    except:
                        has_messages = 0
                    
                    try:
                        latest_message_time = chat.latest_message_time.isoformat() if chat.latest_message_time else None
                    except:
                        latest_message_time = None
                    
                    chat_data = {
                        "conversation_id": conversation_id,
                        "name": user_name,
                        "is_group_chat": False,
                        "is_deleted": False,
                        "has_left": False,
                        "is_deal": False,
                        "latest_message": None,
                        "unread_count": 0,
                        "created_at": created_at,
                        "updated_at": updated_at,
                        "has_messages": has_messages,
                        "latest_message_time": latest_message_time
                    }
                else:
                    try:
                        conversation_id = chat.conversation_id or 0
                    except:
                        conversation_id = 0
                    
                    try:
                        created_at = chat.created_at.isoformat() if chat.created_at else None
                    except:
                        created_at = None
                    
                    try:
                        updated_at = chat.updated_at.isoformat() if chat.updated_at else None
                    except:
                        updated_at = None
                    
                    try:
                        has_messages = chat.has_messages or 0
                    except:
                        has_messages = 0
                    
                    try:
                        latest_message_time = chat.latest_message_time.isoformat() if chat.latest_message_time else None
                    except:
                        latest_message_time = None
                    
                    chat_data = {
                        "conversation_id": conversation_id,
                        "name": "Удаленный чат",
                        "is_group_chat": False,
                        "is_deleted": True,
                        "has_left": True,
                        "is_deal": False,
                        "latest_message": None,
                        "unread_count": 0,
                        "created_at": created_at,
                        "updated_at": updated_at,
                        "has_messages": has_messages,
                        "latest_message_time": latest_message_time
                    }
            
            # Получаем последнее сообщение
            latest_message = chat.messages_set.order_by('-created_at').first()
            if latest_message:
                try:
                    is_read = latest_message.is_read()
                except:
                    is_read = False
                
                try:
                    sender_name = f"{latest_message.sender.first_name} {latest_message.sender.last_name}" if latest_message.sender else "Неизвестно"
                except:
                    sender_name = "Неизвестно"
                
                try:
                    created_at = latest_message.created_at.strftime("%d.%m.%Y %H:%M") if latest_message.created_at else ""
                    created_at_time = latest_message.created_at.strftime("%H:%M") if latest_message.created_at else ""
                    created_at_date = latest_message.created_at.strftime("%d.%m") if latest_message.created_at else ""
                except:
                    created_at = ""
                    created_at_time = ""
                    created_at_date = ""
                
                try:
                    message_text = latest_message.message_text or ""
                except:
                    message_text = ""
                
                try:
                    message_id = latest_message.message_id or 0
                except:
                    message_id = 0
                
                try:
                    sender_id = latest_message.sender.user_id if latest_message.sender else None
                except:
                    sender_id = None
                
                chat_data["last_message"] = {
                    "message_id": message_id,
                    "message_text": message_text,
                    "sender_id": sender_id,
                    "sender_name": sender_name,
                    "created_at": created_at,
                    "created_at_time": created_at_time,
                    "created_at_date": created_at_date,
                    "is_read": is_read
                }
            
            # Подсчитываем непрочитанные сообщения
            try:
                unread_count = chat.messages_set.filter(
                    ~Q(sender=request.user),
                    is_read=False
                ).count()
            except:
                unread_count = 0
            chat_data["unread_count"] = unread_count
            
            # Добавляем информацию об участнике для личных чатов
            if not chat.is_group_chat:
                other_participant = None
                for p in chat.chatparticipants_set.all():
                    if p.user_id != request.user.user_id:
                        other_participant = p
                        break
                
                if other_participant and other_participant.user:
                    try:
                        profile_picture_url = other_participant.user.get_profile_picture_url()
                    except:
                        profile_picture_url = None
                    
                    try:
                        first_name = other_participant.user.first_name or ""
                        last_name = other_participant.user.last_name or ""
                    except:
                        first_name = ""
                        last_name = ""
                    
                    try:
                        user_id = other_participant.user.user_id or 0
                    except:
                        user_id = 0
                    
                    chat_data["participant"] = {
                        "user_id": user_id,
                        "first_name": first_name,
                        "last_name": last_name,
                        "profile_picture_url": profile_picture_url
                    }
            
            chats_data.append(chat_data)
        
        return JsonResponse({
            "success": True,
            "chats": chats_data
        })
        
    except Exception as e:
        import traceback
        logger.error(f"Ошибка при получении списка чатов: {e}")
        logger.error(f"Traceback: {traceback.format_exc()}")
        return JsonResponse({
            "success": False,
            "error": "Внутренняя ошибка сервера"
        }, status=500)

def get_chat_messages(request, chat_id):
    if not request.user.is_authenticated:
        return JsonResponse({"success": False, "error": "Требуется авторизация"})
    chat = get_object_or_404(ChatConversations, conversation_id=chat_id)
    if not chat.chatparticipants_set.filter(user=request.user).exists():
        return JsonResponse(
            {"success": False, "error": "У вас нет доступа к этому чату"}
        )
    since = request.GET.get("since")
    messages = chat.messages_set.all()
    if since:
        try:
            since_dt = dt.fromisoformat(since.replace("Z", "+00:00"))
            messages = messages.filter(created_at__gt=since_dt)
        except ValueError:
            return JsonResponse(
                {"success": False, "error": "Неверный формат параметра since"}
            )
    messages = messages.order_by("created_at")
    messages_data = [
        {
            "message_id": msg.message_id,
            "sender_id": msg.sender.user_id if msg.sender else None,
            "sender_name": (
                f"{msg.sender.first_name} {msg.sender.last_name}"
                if msg.sender
                else "Неизвестно"
            ),
            "message_text": msg.message_text,
            "created_at": (
                msg.created_at.strftime("%d.%m.%Y %H:%M") if msg.created_at else ""
            ),
            "created_at_iso": msg.created_at.isoformat() if msg.created_at else "",
            "is_read": msg.is_read(),
            "is_own": msg.sender == request.user if msg.sender else False,
        }
        for msg in messages
    ]
    participants = chat.get_participants()
    participants_data = [
        {
            "user_id": p.user.user_id,
            "name": f"{p.user.first_name} {p.user.last_name}",
            "role": p.user.role.role_name if p.user.role else "Неизвестно",
        }
        for p in participants
    ]
    return JsonResponse(
        {"success": True, "messages": messages_data, "participants": participants_data}
    )


@login_required
def send_message(request):
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Неверный метод запроса"})
    form = MessageForm(request.POST)
    if not form.is_valid():
        return JsonResponse({"success": False, "error": "Неверные данные формы"})
    chat_id = request.POST.get("chat_id")
    chat = get_object_or_404(ChatConversations, conversation_id=chat_id)
    if not chat.chatparticipants_set.filter(user=request.user).exists():
        return JsonResponse(
            {"success": False, "error": "У вас нет доступа к этому чату"}
        )
    message = Messages(
        conversation=chat,
        sender=request.user,
        message_text=form.cleaned_data["message_text"],
        status=MessageStatuses.objects.get(status_name="sent"),
        created_at=timezone.now(),
        updated_at=timezone.now(),
    )
    message.save()
    chat.updated_at = timezone.now()
    chat.save()
    return JsonResponse(
        {
            "success": True,
            "message": {
                "message_id": message.message_id,
                "sender_id": request.user.user_id,
                "sender_name": f"{request.user.first_name} {request.user.last_name}",
                "message_text": message.message_text,
                "created_at": message.created_at.strftime("%d.%m.%Y %H:%M"),
                "created_at_iso": message.created_at.isoformat(),
                "is_read": message.is_read(),
                "is_own": True,
            },
        }
    )


@login_required
def mark_messages_read(request, chat_id):
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Неверный метод запроса"})
    chat = get_object_or_404(ChatConversations, conversation_id=chat_id)
    if not chat.chatparticipants_set.filter(user=request.user).exists():
        return JsonResponse(
            {"success": False, "error": "У вас нет доступа к этому чату"}
        )
    read_status = MessageStatuses.objects.get(status_name="read")
    messages = chat.messages_set.filter(status__status_name="sent").exclude(
        sender=request.user
    )
    messages.update(status=read_status, updated_at=timezone.now())
    return JsonResponse({"success": True})


@login_required
def start_chat(request, user_id):
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Неверный метод запроса"})
    target_user = get_object_or_404(Users, user_id=user_id)
    if target_user == request.user:
        return JsonResponse(
            {"success": False, "error": "Нельзя создать чат с самим собой"}
        )
    existing_chat = (
        ChatConversations.objects.annotate(num_participants=Count("chatparticipants"))
        .filter(
            is_group_chat=False, num_participants=2, chatparticipants__user=request.user
        )
        .filter(chatparticipants__user=target_user)
        .first()
    )
    if existing_chat:
        return JsonResponse(
            {"success": True, "chat_id": existing_chat.conversation_id, "existed": True}
        )
    chat = ChatConversations.objects.create(
        name=f"Чат {request.user.first_name} и {target_user.first_name}",
        is_group_chat=False,
        created_at=timezone.now(),
        updated_at=timezone.now(),
    )
    ChatParticipants.objects.create(conversation=chat, user=request.user)
    ChatParticipants.objects.create(conversation=chat, user=target_user)
    chat_data = {
        "conversation_id": chat.conversation_id,
        "name": chat.name,
        "is_group_chat": chat.is_group_chat,
        "participant": {
            "user_id": target_user.user_id,
            "first_name": target_user.first_name,
            "profile_picture_url": target_user.get_profile_picture_url(),
        },
        "last_message": None,
        "unread_count": 0,
    }
    return JsonResponse({"success": True, "chat": chat_data, "existed": False})


@login_required
def add_participant(request, chat_id):
    logger.debug(
        f"Adding participant to chat {chat_id}, user_id: {request.POST.get('user_id')}"
    )
    if request.method != "POST":
        return JsonResponse(
            {"success": False, "error": "Неверный метод запроса"}, status=405
        )
    chat = get_object_or_404(ChatConversations, conversation_id=chat_id)
    if not chat.chatparticipants_set.filter(user=request.user).exists():
        return JsonResponse(
            {"success": False, "error": "У вас нет доступа к этому чату"}, status=403
        )
    user_id = request.POST.get("user_id")
    if not user_id:
        return JsonResponse(
            {"success": False, "error": "Не указан пользователь"}, status=400
        )
    try:
        new_user = Users.objects.get(user_id=user_id)
    except Users.DoesNotExist:
        logger.error(f"User {user_id} not found")
        return JsonResponse(
            {"success": False, "error": "Пользователь не найден"}, status=404
        )
    if chat.chatparticipants_set.filter(user=new_user).exists():
        return JsonResponse(
            {"success": False, "error": "Пользователь уже в чате"}, status=400
        )
    if not chat.is_group_chat:
        participants = chat.get_participants()
        if participants.count() >= 3:
            return JsonResponse(
                {"success": False, "error": "В личном чате максимум 3 участника"},
                status=400,
            )
        current_roles = {
            p.user.role.role_name.lower()
            for p in participants
            if p.user and p.user.role
        }
        if new_user.role and new_user.role.role_name.lower() in current_roles:
            return JsonResponse(
                {"success": False, "error": "Пользователь с такой ролью уже в чате"},
                status=400,
            )
    ChatParticipants.objects.create(conversation=chat, user=new_user)
    chat.updated_at = timezone.now()
    chat.save()
    logger.info(f"Добавлен участник {new_user.user_id} в чат {chat.conversation_id}")
    return JsonResponse(
        {
            "success": True,
            "new_participant": {
                "user_id": new_user.user_id,
                "name": f"{new_user.first_name} {new_user.last_name}",
                "role": new_user.role.role_name if new_user.role else "Система",
            },
        }
    )


@login_required
def available_users_for_chat(request, chat_id):
    chat = get_object_or_404(ChatConversations, conversation_id=chat_id)
    if not chat.chatparticipants_set.filter(user=request.user).exists():
        return JsonResponse(
            {"success": False, "error": "У вас нет доступа к этому чату"}, status=403
        )
    current_participant_ids = chat.chatparticipants_set.values_list(
        "user_id", flat=True
    )
    users = Users.objects.exclude(user_id__in=current_participant_ids).exclude(
        user_id=request.user.user_id
    )
    if chat.is_group_chat:
        users = users.exclude(role__role_name="moderator")
    else:
        current_roles = chat.chatparticipants_set.exclude(
            user=request.user
        ).values_list("user__role__role_name", flat=True)
        users = users.filter(
            role__role_name__in=["startuper", "investor", "moderator"]
        ).exclude(role__role_name__in=current_roles)
    users_data = [
        {
            "user_id": user.user_id,
            "name": f"{user.first_name} {user.last_name}",
            "role": user.role.role_name if user.role else "Неизвестно",
        }
        for user in users
    ]
    return JsonResponse({"success": True, "users": users_data})


@login_required
def leave_chat(request, chat_id):
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Неверный метод запроса"})
    chat = get_object_or_404(ChatConversations, conversation_id=chat_id)
    if not chat.chatparticipants_set.filter(user=request.user).exists():
        return JsonResponse(
            {"success": False, "error": "У вас нет доступа к этому чату"}
        )
    ChatParticipants.objects.filter(conversation=chat, user=request.user).delete()
    remaining_participants = chat.chatparticipants_set.all()
    if remaining_participants.exists():
        message = Messages(
            conversation=chat,
            sender=None,
            message_text=f"{request.user.first_name} {request.user.last_name} покинул чат",
            status=MessageStatuses.objects.get(status_name="sent"),
            created_at=timezone.now(),
            updated_at=timezone.now(),
        )
        message.save()
        chat.updated_at = timezone.now()
        chat.save()
    else:
        chat.delete()
        return JsonResponse({"success": True, "deleted": True})
    return JsonResponse({"success": True, "deleted": False})


def planetary_system(request):
    """
    Планетарная система - отображает стартапы как планеты на орбитах
    """
    directions_data = FIXED_CATEGORIES.copy()
    selected_direction_name = request.GET.get("direction", "All")
    logger.info(f"🪐 Планетарная система: выбрано направление '{selected_direction_name}'")
    startups_query = Startups.objects.filter(
        status="approved"
    ).select_related("direction", "owner").order_by("-created_at")
    if selected_direction_name != "All" and selected_direction_name != "Все":
        from django.db.models import Q
        direction_filter = Q()
        for category in FIXED_CATEGORIES:
            if category['original_name'] == selected_direction_name or category['direction_name'] == selected_direction_name:
                direction_filter |= Q(direction__direction_name=category['direction_name'])
        if direction_filter:
            startups_query = startups_query.filter(direction_filter)
    startups_list = list(startups_query)
    print(f"🚀 ПЛАНЕТАРНАЯ СИСТЕМА DEBUG:")
    print(f"🚀 Выбрано направление: '{selected_direction_name}'")
    print(f"🚀 Всего одобренных стартапов в БД: {Startups.objects.filter(status='approved').count()}")
    print(f"🚀 Загружено стартапов после фильтрации: {len(startups_list)}")
    if startups_list:
        for i, startup in enumerate(startups_list[:3]):
            print(f"🚀   {i+1}. {startup.title} - направление: {startup.direction.direction_name if startup.direction else 'Нет'}")
    logger.info(f"🪐 Загружено стартапов: {len(startups_list)}")
    selected_startups = []
    if len(startups_list) > 0:
        selected_startups = startups_list[:6]
    else:
        selected_startups = []
    planets_data = []
    for i, startup in enumerate(selected_startups):
        planet_image_url = None
        
        if startup.planet_image:
            planet_image_url = f"{settings.S3_PUBLIC_BASE_URL}/choosable_planets/{startup.planet_image}"
        
        if not planet_image_url:
            import random
            folder_choice = random.choice(['planets_round', 'planets_ring'])
            if folder_choice == 'planets_round':
                planet_image_num = (i % 15) + 1
                planet_image_url = f"/static/accounts/images/planetary_system/planets_round/{planet_image_num}.png"
            else:
                planet_image_num = (i % 6) + 1
                planet_image_url = f"/static/accounts/images/planetary_system/planets_ring/{planet_image_num}.png"
        
        direction_original = 'Не указано'
        if startup.direction:
            for cat in directions_data:
                if cat['direction_name'] == startup.direction.direction_name or cat['original_name'] == getattr(startup.direction, 'original_name', None):
                    direction_original = cat['original_name']
                    break
        planets_data.append({
            "id": startup.startup_id,
            "startup_id": startup.startup_id,
            "name": startup.title,
            "description": startup.short_description or startup.description[:200] if startup.description else "",
            "image": planet_image_url,
            "rating": startup.get_average_rating(),
            "voters_count": startup.total_voters,
            "comment_count": startup.comments.count(),
            "direction": direction_original,
            "funding_goal": f"{startup.funding_goal:,.0f} ₽".replace(",", " ") if startup.funding_goal else "Не указано",
            "valuation": f"{startup.valuation:,.0f} ₽".replace(",", " ") if startup.valuation else "Не указано",
            "investors": startup.get_investors_count(),
            "progress": startup.get_progress_percentage(),
            "investment_type": "Выкуп+инвестирование" if startup.both_mode else ("Только выкуп" if startup.only_buy else "Только инвестирование")
        })
    all_approved_startups = list(Startups.objects.filter(status="approved").select_related("direction", "owner").order_by("-created_at"))
    all_startups_data = []
    for idx, startup in enumerate(all_approved_startups):
        planet_image_url = None
        
        if startup.planet_image:
            planet_image_url = f"{settings.S3_PUBLIC_BASE_URL}/choosable_planets/{startup.planet_image}"
        
        if not planet_image_url:
            import random
            folder_choice = random.choice(['planets_round', 'planets_ring'])
            if folder_choice == 'planets_round':
                planet_image_num = (idx % 15) + 1
                planet_image_url = f"/static/accounts/images/planetary_system/planets_round/{planet_image_num}.png"
            else:
                planet_image_num = (idx % 6) + 1
                planet_image_url = f"/static/accounts/images/planetary_system/planets_ring/{planet_image_num}.png"
        
        direction_original = 'Не указано'
        if startup.direction:
            for cat in directions_data:
                if cat['direction_name'] == startup.direction.direction_name or cat['original_name'] == getattr(startup.direction, 'original_name', None):
                    direction_original = cat['original_name']
                    break
        all_startups_data.append({
            "id": startup.startup_id,
            "startup_id": startup.startup_id,
            "name": startup.title,
            "description": startup.short_description or startup.description[:200] if startup.description else "",
            "image": planet_image_url,
            "rating": startup.get_average_rating(),
            "voters_count": startup.total_voters,
            "comment_count": startup.comments.count(),
            "direction": direction_original,
            "funding_goal": f"{startup.funding_goal:,.0f} ₽".replace(",", " ") if startup.funding_goal else "Не указано",
            "valuation": f"{startup.valuation:,.0f} ₽".replace(",", " ") if startup.valuation else "Не указано",
            "investors": startup.get_investors_count(),
            "progress": startup.get_progress_percentage(),
            "investment_type": "Выкуп+инвестирование" if startup.both_mode else ("Только выкуп" if startup.only_buy else "Только инвестирование")
        })
    logo_data = {
        "image": "/static/accounts/images/logo.png"
    }
    print(f"🚀 ПЕРЕДАЕТСЯ В ШАБЛОН:")
    print(f"🚀 Планет для отображения: {len(planets_data)}")
    print(f"🚀 Всех стартапов для фильтрации: {len(all_startups_data)}")
    print(f"🚀 Направлений: {len(directions_data)}")
    print(f"🚀 Выбранная галактика: '{selected_direction_name}'")
    print(f"🚀 Первые 3 планеты: {[p.get('name', 'Нет названия') for p in planets_data[:3]]}")
    print(f"🚀 Переводы направлений: {[(d.get('original_name'), d.get('direction_name')) for d in directions_data[:5]]}")
    context = {
        "planets_data_json": json.dumps(planets_data, ensure_ascii=False),
        "directions_data_json": json.dumps(directions_data, ensure_ascii=False),
        "all_startups_data_json": json.dumps(all_startups_data, ensure_ascii=False),
        "logo_data": logo_data,
        "directions": directions_data,
        "selected_galaxy": selected_direction_name,
    }
    return render(request, "accounts/planetary_system.html", context)


@login_required
def my_startups(request):
    try:
        if request.user.role and request.user.role.role_name == 'startuper':
            user_startups_qs = (
                Startups.objects.filter(owner=request.user)
                .select_related("direction", "stage", "status_id")
                .prefetch_related("comments")
            )
        else:
            user_startups_qs = (
                Startups.objects.all()
                .select_related("direction", "stage", "status_id")
                .prefetch_related("comments")
            )
        total_user_startups_count = user_startups_qs.count()
        approved_startups_qs = user_startups_qs.filter(status="approved")
        financial_analytics_data = approved_startups_qs.aggregate(
            total_raised=Sum("amount_raised"),
            max_raised=Max("amount_raised"),
            approved_startups_count=Count("startup_id"),
        )
        approved_startups_count = financial_analytics_data.get(
            "approved_startups_count", 0
        )
        total_amount_raised = financial_analytics_data.get("total_raised") or Decimal(
            "0"
        )
        max_raised = financial_analytics_data.get("max_raised") or Decimal("0")
        
        startups_with_funding = approved_startups_qs.filter(amount_raised__gt=0)
        min_raised_data = startups_with_funding.aggregate(
            min_raised=Min("amount_raised")
        )
        min_raised = min_raised_data.get("min_raised") or Decimal("0")
        category_data_raw = (
            user_startups_qs.values("direction__direction_name")
            .annotate(category_count=Count("startup_id"))
            .order_by("-category_count")
        )
        investment_categories = []
        invested_category_data_dict = {}
        total_for_category_percentage = (
            total_user_startups_count if total_user_startups_count > 0 else 1
        )
        for cat_data in category_data_raw:
            percentage = 0
            category_count = cat_data.get("category_count")
            category_name = cat_data.get("direction__direction_name") or "Без категории"
            if category_count and total_for_category_percentage > 0:
                try:
                    percentage = round(
                        (int(category_count) / total_for_category_percentage) * 100
                    )
                    percentage = min(percentage, 100)
                except Exception as e:
                    logger.error(
                        f"Ошибка расчета процента (по количеству) для категории '{category_name}': {e}"
                    )
                    percentage = 0
            investment_categories.append(
                {
                    "name": category_name,
                    "percentage": percentage,
                }
            )
            invested_category_data_dict[category_name] = percentage
        current_year = timezone.now().year
        logger.info(
            f"[my_startups] Preparing chart data for user {request.user.email}, year: {current_year}"
        )
        monthly_data_direct = (
            approved_startups_qs.filter(
                updated_at__year=current_year, amount_raised__gt=0
            )
            .annotate(month=TruncMonth("updated_at"))
            .values("month")
            .annotate(monthly_total=Sum(Coalesce("amount_raised", Decimal(0))))
            .order_by("month")
        )
        month_labels = [
            "Янв",
            "Фев",
            "Мар",
            "Апр",
            "Май",
            "Июн",
            "Июл",
            "Авг",
            "Сен",
            "Окт",
            "Ноя",
            "Дек",
        ]
        monthly_totals = [0] * 12
        for data in monthly_data_direct:
            month_index = data["month"].month - 1
            if 0 <= month_index < 12:
                monthly_total_decimal = data.get(
                    "monthly_total", Decimal(0)
                ) or Decimal(0)
                monthly_totals[month_index] = float(monthly_total_decimal)
        logger.info(
            f"[my_startups] Preparing chart data for user {request.user.email}, year: {current_year}"
        )
        monthly_category_data_raw = (
            approved_startups_qs.filter(
                updated_at__year=current_year,
                amount_raised__gt=0,
                direction__isnull=False,
            )
            .annotate(month=TruncMonth("updated_at"))
            .values("month", "direction__direction_name")
            .annotate(monthly_category_total=Sum(Coalesce("amount_raised", Decimal(0))))
            .order_by("month", "direction__direction_name")
        )
        logger.info(
            f"[my_startups] Raw monthly category data from DB: {list(monthly_category_data_raw)}"
        )
        structured_monthly_data = collections.defaultdict(
            lambda: collections.defaultdict(float)
        )
        unique_categories = set()
        for data in monthly_category_data_raw:
            month_dt = data["month"]
            category_name = data["direction__direction_name"]
            amount = float(data.get("monthly_category_total", 0) or 0)
            month_key = month_dt.strftime("%Y-%m-01")
            structured_monthly_data[month_key][category_name] += amount
            unique_categories.add(category_name)
        sorted_categories = sorted(list(unique_categories))
        logger.info(
            f"[my_startups] Unique categories found for chart: {sorted_categories}"
        )
        chart_data_list = []
        start_date = datetime.date(current_year, 1, 1)
        for i in range(12):
            current_month_key = (start_date + relativedelta(months=i)).strftime(
                "%Y-%m-01"
            )
            month_data = {
                "month_key": current_month_key,
                "category_data": dict(structured_monthly_data[current_month_key]),
            }
            chart_data_list.append(month_data)
        logger.info(
            f"[my_startups] Final structured chart data list: {chart_data_list}"
        )
        try:
            all_directions_qs = Directions.objects.all().order_by("direction_name")
            all_directions_list = [
                {"direction_name": d.direction_name} for d in all_directions_qs
            ]
        except Exception as e:
            logger.error(f"Ошибка при получении всех направлений: {str(e)}")
            all_directions_list = []
        try:
            approved_startups_annotated = (
                approved_startups_qs.annotate(
                    average_rating=Avg(
                        models.ExpressionWrapper(
                            Coalesce(models.F("sum_votes"), 0)
                            * 1.0
                            / Coalesce(models.F("total_voters"), 1),
                            output_field=FloatField(),
                        ),
                        filter=models.Q(total_voters__gt=0),
                        default=0.0,
                    ),
                    comment_count=Count("comments"),
                )
                .annotate(average_rating=Coalesce("average_rating", 0.0))
                .order_by("-created_at")
            )
        except Exception as e:
            logger.error(f"Ошибка при получении одобренных стартапов: {str(e)}")
            approved_startups_annotated = []
        s3_client = boto3.client(
            "s3",
            aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
            aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
            endpoint_url=settings.AWS_S3_ENDPOINT_URL,
            region_name=settings.AWS_S3_REGION_NAME,
        )
        planetary_startups = []
        print(f"🚀 DEBUG: approved_startups_annotated count: {len(approved_startups_annotated)}")
        for idx, startup in enumerate(approved_startups_annotated, start=1):
            orbit_size = (idx * 100) + 150
            orbit_time = (idx * 10) + 40
            planet_size = 60
            planet_data = {
                "startup_id": startup.startup_id,
                "title": startup.title or "Без названия",
                "planet_image": startup.planet_image,
                "logo_urls": startup.logo_urls,
                "average_rating": startup.average_rating or 0,
                "rating": startup.average_rating or 0,
                "total_voters": startup.total_voters or 0,
                "comment_count": startup.comment_count or 0,
                "description": startup.description or "Описание отсутствует.",
                "short_description": startup.description or "Описание отсутствует.",
                "progress": startup.get_progress_percentage() or 0,
                "funding_goal": startup.funding_goal or 0,
                "amount_raised": startup.amount_raised or 0,
                "get_investors_count": startup.get_investors_count(),
                "direction": startup.direction.direction_name if startup.direction else "Не указано",
                "investment_type": "Не указано",
                "orbit_size": orbit_size,
                "orbit_time": orbit_time,
                "planet_size": planet_size,
            }
            planetary_startups.append(planet_data)
            print(f"🚀 DEBUG: Added planet data for startup {startup.startup_id}: {startup.title}")
        print(f"🚀 DEBUG: Total planetary_startups: {len(planetary_startups)}")
    except Exception as e:
        logger.error(f"Критическая ошибка в my_startups view: {e}", exc_info=True)
        messages.error(
            request, "Произошла ошибка при загрузке страницы ваших стартапов."
        )
        return redirect("profile")
    context = {
        "user_startups": approved_startups_annotated,
        "planetary_startups": planetary_startups,
        "total_investment": total_amount_raised,
        "startups_count": approved_startups_count,
        "max_investment": max_raised,
        "min_investment": min_raised,
        "investment_categories": investment_categories[
            :7
        ],
        "invested_category_data": invested_category_data_dict,
        "all_directions": all_directions_list,
        "month_labels": month_labels,
        "chart_monthly_category_data": chart_data_list,
        "chart_categories": sorted_categories,
        "startup_applications": user_startups_qs.order_by("-updated_at"),
    }
    context["planetary_startups_json"] = json.dumps(
        planetary_startups, cls=DjangoJSONEncoder, ensure_ascii=False
    )
    return render(request, "accounts/my_startups.html", context)


@login_required
def notifications_view(request):
    return render(request, "accounts/notifications.html")


@login_required
def create_group_chat(request):
    if request.method != "POST":
        return JsonResponse(
            {"success": False, "error": "Метод не разрешен."}, status=405
        )
    try:
        data = json.loads(request.body)
        chat_name = data.get("name", "").strip()
        user_ids = data.get("user_ids", [])
        if not chat_name:
            return JsonResponse(
                {"success": False, "error": "Название чата не может быть пустым."},
                status=400,
            )
        if not user_ids:
            return JsonResponse(
                {
                    "success": False,
                    "error": "Необходимо выбрать хотя бы одного участника.",
                },
                status=400,
            )
        try:
            participant_ids = list(set(int(uid) for uid in user_ids))
        except (ValueError, TypeError):
            return JsonResponse(
                {"success": False, "error": "Неверный формат ID пользователей."},
                status=400,
            )
        if request.user.user_id in participant_ids:
            participant_ids.remove(request.user.user_id)
        if not participant_ids:
            return JsonResponse(
                {
                    "success": False,
                    "error": "Нельзя создать групповой чат только с самим собой.",
                },
                status=400,
            )
        if Users.objects.filter(
            user_id__in=participant_ids, role__role_name="moderator"
        ).exists():
            return JsonResponse(
                {
                    "success": False,
                    "error": "Модераторы не могут быть добавлены в групповой чат.",
                },
                status=400,
            )
        with transaction.atomic():
            conversation = ChatConversations.objects.create(
                name=chat_name,
                is_group_chat=True,
                created_at=timezone.now(),
                updated_at=timezone.now(),
            )
            all_participant_users = [request.user]
            users_to_add = Users.objects.filter(user_id__in=participant_ids)
            all_participant_users.extend(list(users_to_add))
            if len(all_participant_users) != len(participant_ids) + 1:
                logger.error(
                    f"Не все пользователи найдены для создания чата. Передано ID: {participant_ids}"
                )
                raise Exception("Один или несколько пользователей не найдены.")
            participants_to_create = [
                ChatParticipants(conversation=conversation, user=user)
                for user in all_participant_users
            ]
            ChatParticipants.objects.bulk_create(participants_to_create)
        chat_data = {
            "conversation_id": conversation.conversation_id,
            "name": conversation.name,
            "is_group_chat": conversation.is_group_chat,
            "participant": None,
            "last_message": None,
            "unread_count": 0,
        }
        logger.info(
            f"Групповой чат создан: ID={conversation.conversation_id}, Название={chat_name}, Участников={len(all_participant_users)}"
        )
        return JsonResponse({"success": True, "chat": chat_data})
    except json.JSONDecodeError:
        logger.error("Неверный формат JSON в create_group_chat")
        return JsonResponse(
            {"success": False, "error": "Неверный формат данных (JSON)."}, status=400
        )
    except Exception as e:
        logger.error(f"Ошибка при создании группового чата: {str(e)}", exc_info=True)
        return JsonResponse(
            {"success": False, "error": "Внутренняя ошибка сервера."}, status=500
        )


@login_required
def support_page_view(request):
    is_moderator = (
        request.user.is_authenticated
        and request.user.role
        and request.user.role.role_name == "moderator"
    )
    context = {"is_moderator": is_moderator}
    return render(request, "accounts/support.html", context)


@login_required
def change_owner(request, startup_id):
    logger.info(f"Change owner request for startup {startup_id} by user {request.user.user_id}")
    
    if request.method != "POST":
        logger.warning(f"Invalid method {request.method} for change_owner")
        return JsonResponse({"success": False, "error": "Неверный метод запроса"})
    
    if not getattr(request.user, "role", None) or (request.user.role.role_name or "").lower() != "moderator":
        logger.warning(f"User {request.user.user_id} does not have moderator role")
        return JsonResponse(
            {"success": False, "error": "У вас нет прав для этого действия"}
        )
    
    try:
        startup = get_object_or_404(Startups, startup_id=startup_id)
        new_owner_id = request.POST.get("new_owner_id")
        
        if not new_owner_id:
            logger.error("No new_owner_id provided")
            return JsonResponse({"success": False, "error": "Не указан новый владелец"})
        
        new_owner = get_object_or_404(Users, user_id=new_owner_id)
        startup.owner = new_owner
        startup.save()
        
        logger.info(f"Successfully changed owner of startup {startup_id} to user {new_owner_id}")
        return JsonResponse({"success": True})
        
    except Exception as e:
        logger.error(f"Error changing owner for startup {startup_id}: {str(e)}")
        return JsonResponse({"success": False, "error": f"Ошибка при смене владельца: {str(e)}"})


@login_required
def get_investors(request, startup_id):
    logger.info(f"Get investors request for startup {startup_id} by user {request.user.user_id}")
    
    if not request.user.is_authenticated or (request.user.role.role_name or "").lower() != "moderator":
        logger.warning(f"User {request.user.user_id} does not have moderator role for get_investors")
        return JsonResponse({"error": "Доступ запрещен"}, status=403)
    
    try:
        startup = get_object_or_404(Startups, startup_id=startup_id)
        investors = InvestmentTransactions.objects.filter(startup=startup).select_related(
            "investor"
        )
        
        logger.info(f"Found {investors.count()} investment transactions for startup {startup_id}")
        
        investor_list = []
        for tx in investors:
            if tx.investor:
                investor_list.append(
                    {
                        "user_id": tx.investor.user_id,
                        "name": tx.investor.get_full_name() or tx.investor.email,
                        "amount": float(tx.amount),
                    }
                )
        
        html = render_to_string(
            "accounts/partials/_investors_list.html",
            {"investors": investor_list, "startup": startup, "user": request.user},
        )
        
        logger.info(f"Generated HTML for {len(investor_list)} investors")
        return JsonResponse({"html": html})
        
    except Exception as e:
        logger.error(f"Error getting investors for startup {startup_id}: {str(e)}")
        return JsonResponse({"error": f"Ошибка при получении списка инвесторов: {str(e)}"}, status=500)


@login_required
def add_investor(request, startup_id):
    logger.info(f"Add investor request for startup {startup_id} by user {request.user.user_id}")
    
    if not request.user.is_authenticated or (request.user.role.role_name or "").lower() != "moderator":
        logger.warning(f"User {request.user.user_id} does not have moderator role for add_investor")
        return JsonResponse({"error": "Доступ запрещен"}, status=403)
    
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            user_id = data.get("user_id")
            amount = Decimal(data.get("amount"))
            
            logger.info(f"Adding investor {user_id} with amount {amount} to startup {startup_id}")
            
            startup = get_object_or_404(Startups, startup_id=startup_id)
            user_to_invest = get_object_or_404(Users, user_id=user_id)
            
            if amount <= 0:
                logger.warning(f"Invalid amount {amount} for startup {startup_id}")
                return JsonResponse(
                    {"success": False, "error": "Сумма должна быть положительной."}
                )
            
            existing_tx = InvestmentTransactions.objects.filter(
                startup_id=startup_id, investor=user_to_invest
            ).first()
            
            if existing_tx:
                logger.info(f"Updating existing investment for user {user_id} in startup {startup_id}")
                existing_tx.amount = amount
                existing_tx.save()
            else:
                logger.info(f"Creating new investment for user {user_id} in startup {startup_id}")
                try:
                    investment_type_obj = TransactionTypes.objects.get(
                        type_name="investment"
                    )
                    InvestmentTransactions.objects.create(
                        investor=user_to_invest,
                        startup=startup,
                        amount=amount,
                        transaction_type=investment_type_obj,
                    )
                except TransactionTypes.DoesNotExist:
                    logger.error("Transaction type 'investment' not found")
                    return JsonResponse(
                        {"error": "Тип транзакции 'investment' не найден в системе."},
                        status=500,
                    )
            
            startup.amount_raised = startup.investmenttransactions_set.aggregate(
                total=Sum("amount")
            )["total"] or Decimal("0")
            startup.save(update_fields=["amount_raised"])
            new_investor_count = startup.get_investors_count()
            
            logger.info(f"Successfully added investor to startup {startup_id}. New amount: {startup.amount_raised}, investors: {new_investor_count}")
            
            return JsonResponse(
                {
                    "success": True,
                    "new_amount_raised": float(startup.amount_raised),
                    "new_investor_count": new_investor_count,
                }
            )
        except (json.JSONDecodeError, TypeError, ValueError) as e:
            logger.error(f"Data format error in add_investor: {str(e)}")
            return JsonResponse(
                {"error": f"Неверный формат данных: {str(e)}"}, status=400
            )
        except Exception as e:
            logger.error(f"Unexpected error in add_investor: {str(e)}")
            return JsonResponse(
                {"error": f"Внутренняя ошибка сервера: {str(e)}"}, status=500
            )
    
    logger.warning(f"Invalid method {request.method} for add_investor")
    return JsonResponse({"error": "Метод не поддерживается"}, status=405)


@login_required
def edit_investment(request, startup_id, user_id):
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Неверный метод запроса"})
    if not getattr(request.user, "role", None) or (request.user.role.role_name or "").lower() != "moderator":
        return JsonResponse(
            {"success": False, "error": "У вас нет прав для этого действия"}
        )
    startup = get_object_or_404(Startups, startup_id=startup_id)
    investor = get_object_or_404(Users, user_id=user_id)
    new_amount = Decimal(request.POST.get("amount", "0"))
    if new_amount <= 0:
        return JsonResponse({"success": False, "error": "Сумма должна быть больше 0"})
    transaction = get_object_or_404(
        InvestmentTransactions,
        startup=startup,
        investor=investor,
        transaction_status="completed",
    )
    old_amount = transaction.amount
    transaction.amount = new_amount
    transaction.updated_at = timezone.now()
    transaction.save()
    startup.amount_raised = (
        (startup.amount_raised or Decimal("0")) - old_amount + new_amount
    )
    startup.save()
    return JsonResponse({"success": True})


@login_required
def delete_investment(request, startup_id, user_id):
    logger.info(f"Delete investment request for startup {startup_id}, user {user_id} by user {request.user.user_id}")
    
    if not request.user.is_authenticated or (request.user.role.role_name or "").lower() != "moderator":
        logger.warning(f"User {request.user.user_id} does not have moderator role for delete_investment")
        return JsonResponse({"error": "Доступ запрещен"}, status=403)
    
    if request.method == "POST":
        with transaction.atomic():
            try:
                user_to_delete = get_object_or_404(Users, pk=user_id)
                tx = get_object_or_404(
                    InvestmentTransactions,
                    startup_id=startup_id,
                    investor=user_to_delete,
                )
                
                logger.info(f"Found investment transaction {tx.id} for deletion")
                
                startup = tx.startup
                tx.delete()
                
                new_total = startup.investmenttransactions_set.aggregate(
                    total=Sum("amount")
                )["total"] or Decimal("0")
                startup.amount_raised = new_total
                startup.save(update_fields=["amount_raised"])
                new_investor_count = startup.get_investors_count()
                
                logger.info(f"Successfully deleted investment. New total: {new_total}, investors: {new_investor_count}")
                
                return JsonResponse(
                    {
                        "success": True,
                        "new_amount_raised": float(startup.amount_raised),
                        "new_investor_count": new_investor_count,
                    }
                )
            except InvestmentTransactions.DoesNotExist:
                logger.warning(f"Investment transaction not found for startup {startup_id}, user {user_id}")
                return JsonResponse({"error": "Инвестиция не найдена"}, status=404)
            except Exception as e:
                logger.error(f"Ошибка при удалении инвестиции: {e}")
                return JsonResponse({"error": "Внутренняя ошибка сервера"}, status=500)
    
    logger.warning(f"Invalid method {request.method} for delete_investment")
    return JsonResponse({"error": "Неверный метод запроса"}, status=405)


@login_required
def support_orders_view(request):
    if (
        request.user.is_authenticated
        and request.user.role
        and request.user.role.role_name == "moderator"
    ):
        orders = SupportTicket.objects.all().order_by("-created_at")
        is_moderator = True
    else:
        orders = SupportTicket.objects.filter(user=request.user).order_by("-created_at")
        is_moderator = False
    
    # Добавляем пагинацию
    page_number = request.GET.get('page', 1)
    paginator = Paginator(orders, 10)  # 10 заявок на страницу
    page_obj = paginator.get_page(page_number)
    
    context = {
        "orders": page_obj,  # Используем page_obj вместо orders
        "page_obj": page_obj,
        "paginator": paginator,
        "is_moderator": is_moderator
    }
    return render(request, "accounts/support_orders.html", context)


@login_required
def support_ticket_detail(request, ticket_id):
    ticket = get_object_or_404(SupportTicket, pk=ticket_id)
    user = request.user
    is_moderator = (
        user.is_authenticated and user.role and user.role.role_name == "moderator"
    )
    if not (user == ticket.user or is_moderator):
        return HttpResponseForbidden("У вас нет доступа к этой заявке.")
    
    if is_moderator:
        all_tickets = SupportTicket.objects.all().order_by("-created_at")
    else:
        all_tickets = SupportTicket.objects.filter(user=user).order_by("-created_at")
    
    form = None
    if is_moderator:
        if request.method == "POST":
            form = ModeratorTicketForm(request.POST, instance=ticket)
            if form.is_valid():
                form.save()
                ticket.refresh_from_db()
                messages.success(request, "Заявка успешно обновлена.")
                return redirect("support_ticket_detail", ticket_id=ticket.ticket_id)
            else:
                print(f"DEBUG: Форма невалидна: {form.errors}")
        else:
            form = ModeratorTicketForm(instance=ticket)
    
    context = {
        "ticket": ticket,
        "form": form,
        "is_moderator": is_moderator,
        "all_tickets": all_tickets,
    }
    return render(request, "accounts/support_ticket_detail.html", context)


@login_required
def close_support_ticket(request, ticket_id):
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Неверный метод запроса"}, status=405)
    
    ticket = get_object_or_404(SupportTicket, pk=ticket_id)
    user = request.user
    
    is_moderator = (
        user.is_authenticated and user.role and user.role.role_name == "moderator"
    )
    
    if not (user == ticket.user or is_moderator):
        return JsonResponse({"success": False, "error": "У вас нет доступа к этой заявке"}, status=403)
    
    try:
        ticket.status = 'closed'
        ticket.save()
        return JsonResponse({"success": True})
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=500)


@login_required
def update_ticket_status(request, ticket_id):
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Неверный метод запроса"}, status=405)
    
    ticket = get_object_or_404(SupportTicket, pk=ticket_id)
    user = request.user
    
    is_moderator = (
        user.is_authenticated and user.role and user.role.role_name == "moderator"
    )
    
    if not is_moderator:
        return JsonResponse({"success": False, "error": "У вас нет прав для изменения статуса"}, status=403)
    
    try:
        import json
        data = json.loads(request.body)
        new_status = data.get('status')
        
        if new_status not in ['new', 'in_progress', 'closed']:
            return JsonResponse({"success": False, "error": "Неверный статус"}, status=400)
        
        ticket.status = new_status
        ticket.save()
        return JsonResponse({"success": True})
    except json.JSONDecodeError:
        return JsonResponse({"success": False, "error": "Неверный формат данных"}, status=400)
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=500)


@login_required
def support_contact_view(request):
    if request.method == "POST":
        form = SupportTicketForm(request.POST)
        if form.is_valid():
            ticket = form.save(commit=False)
            ticket.user = request.user
            ticket.save()
            try:
                logger.info(f"Dispatching Telegram for support ticket {ticket.ticket_id}")
                sent_ok = send_telegram_support_message(ticket)
                logger.info(f"Telegram dispatch result for ticket {ticket.ticket_id}: {sent_ok}")
            except Exception as e:
                logger.error(f"Unexpected error during Telegram dispatch for ticket {ticket.ticket_id}: {e}", exc_info=True)
            messages.success(
                request, "Ваше обращение успешно отправлено! Мы скоро с вами свяжемся."
            )
            return redirect("support_contact")
    else:
        form = SupportTicketForm()
    context = {"form": form}
    return render(request, "accounts/support_contact.html", context)


@login_required
def rename_chat(request, chat_id):
    if request.method != "POST":
        return JsonResponse(
            {"success": False, "error": "Неверный метод запроса"}, status=405
        )
    chat = get_object_or_404(ChatConversations, conversation_id=chat_id)
    if not chat.chatparticipants_set.filter(user=request.user).exists():
        return JsonResponse(
            {"success": False, "error": "У вас нет доступа к этому чату"}, status=403
        )
    try:
        data = json.loads(request.body)
        new_name = data.get("name", "").strip()
        if not new_name:
            return JsonResponse(
                {"success": False, "error": "Название не может быть пустым"}, status=400
            )
        with transaction.atomic():
            chat.name = new_name
            chat.updated_at = timezone.now()
            chat.save()
        logger.info(f"Чат {chat.conversation_id} переименован в {new_name}")
        return JsonResponse({"success": True, "chat_name": new_name})
    except json.JSONDecodeError:
        logger.error("Неверный формат JSON в rename_chat")
        return JsonResponse(
            {"success": False, "error": "Неверный формат данных"}, status=400
        )
    except Exception as e:
        logger.error(f"Ошибка при переименовании чата {chat_id}: {str(e)}")
        return JsonResponse(
            {"success": False, "error": f"Ошибка: {str(e)}"}, status=500
        )


@login_required
def available_users(request):
    users = Users.objects.exclude(user_id=request.user.user_id).exclude(
        role__role_name="moderator"
    )
    users_data = [
        {
            "user_id": user.user_id,
            "name": f"{user.first_name} {user.last_name}",
            "role": user.role.role_name if user.role else "unknown",
            "profile_picture_url": user.get_profile_picture_url() or "",
        }
        for user in users
    ]
    return JsonResponse({"success": True, "users": users_data})


@login_required
def find_or_create_chat(request, recipient_id):
    if request.method == "POST":
        recipient = get_object_or_404(Users, user_id=recipient_id)
        if request.user.user_id == recipient.user_id:
            return JsonResponse(
                {"error": "You cannot start a chat with yourself."}, status=400
            )
        user_chats = ChatConversations.objects.filter(
            is_group_chat=False, chatparticipants__user=request.user
        ).annotate(num_participants=Count("chatparticipants"))
        personal_chats = user_chats.filter(num_participants=2)
        chat = personal_chats.filter(chatparticipants__user=recipient).first()
        if not chat:
            chat = ChatConversations.objects.create(
                is_group_chat=False,
                created_at=timezone.now(),
                updated_at=timezone.now(),
            )
            ChatParticipants.objects.create(conversation=chat, user=request.user)
            ChatParticipants.objects.create(conversation=chat, user=recipient)
        chat_url = reverse("cosmochat") + f"?chat_id={chat.conversation_id}"
        return JsonResponse({"chat_url": chat_url})
    return JsonResponse({"error": "Invalid request method."}, status=405)
def get_user_rating_for_startup(user_id, startup_id):
    """
    // ... existing code ...
    """
    pass
def custom_404(request, exception):
    return render(request, "accounts/404.html", status=404)
@csrf_exempt
@require_POST
def telegram_webhook(request, token):
    from django.conf import settings
    bot_token = getattr(settings, 'TELEGRAM_BOT_TOKEN', None)
    if not bot_token:
        logger.error("TELEGRAM_BOT_TOKEN is not configured")
        return HttpResponse(status=500)
    if token != bot_token:
        logger.warning("Invalid token in webhook URL.")
        return HttpResponseForbidden("Invalid token")
    try:
        data = json.loads(request.body)
        logger.info(f"Webhook received data: {data}")
        if "callback_query" not in data:
            return HttpResponse(status=200)
        callback_query = data["callback_query"]
        callback_data = callback_query["data"]
        message = callback_query["message"]
        chat_id = message["chat"]["id"]
        message_id = message["message_id"]
        requests.post(
            f"https://api.telegram.org/bot{bot_token}/answerCallbackQuery",
            json={"callback_query_id": callback_query["id"]},
        )
        new_text = message.get("text", "")
        new_keyboard = None
        ticket = None
        if callback_data.startswith("close_ticket_"):
            ticket_id = int(callback_data.split("_")[2])
            ticket = SupportTicket.objects.filter(pk=ticket_id).first()
            if ticket:
                ticket.status = "closed"
                ticket.save(update_fields=["status"])
                status_line = "\n\n<b>✅ ЗАЯВКА ЗАКРЫТА</b>"
                if status_line not in new_text:
                    new_text += status_line
                new_keyboard = {
                    "inline_keyboard": [
                        [
                            {
                                "text": "↩️ Вернуть в работу",
                                "callback_data": f"reopen_ticket_{ticket.ticket_id}",
                            }
                        ]
                    ]
                }
                logger.info(f"Ticket {ticket_id} was closed via Telegram.")
        elif callback_data.startswith("reopen_ticket_"):
            ticket_id = int(callback_data.split("_")[2])
            ticket = SupportTicket.objects.filter(pk=ticket_id).first()
            if ticket:
                ticket.status = "new"
                ticket.save(update_fields=["status"])
                status_line = "\n\n<b>✅ ЗАЯВКА ЗАКРЫТА</b>"
                if new_text.endswith(status_line):
                    new_text = new_text[: -len(status_line)]
                new_keyboard = {
                    "inline_keyboard": [
                        [
                            {
                                "text": "✅ Исполнено",
                                "callback_data": f"close_ticket_{ticket.ticket_id}",
                            }
                        ]
                    ]
                }
                logger.info(f"Ticket {ticket_id} was reopened via Telegram.")
        if new_keyboard:
            payload = {
                "chat_id": chat_id,
                "message_id": message_id,
                "text": new_text,
                "parse_mode": "HTML",
                "reply_markup": new_keyboard,
            }
            requests.post(
                f"https://api.telegram.org/bot{bot_token}/editMessageText", json=payload
            )
        return HttpResponse(status=200)
    except json.JSONDecodeError:
        logger.error("Error decoding JSON from Telegram webhook.")
        return HttpResponse(status=400)
    except Exception as e:
        logger.error(f"Error processing Telegram webhook: {e}", exc_info=True)
        return HttpResponse(status=500)

@login_required
def download_startups_report(request):
    try:
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Стартапы"
        
        ws.merge_cells('A1:K1')
        title_cell = ws.cell(row=1, column=1, value="Отчет по стартапам")
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        ws.cell(row=2, column=1, value="Владелец").font = Font(bold=True)
        owner_name = ""
        if request.user.role and request.user.role.role_name == 'startuper':
            first_name = request.user.first_name or ""
            last_name = request.user.last_name or ""
            owner_name = f"{first_name} {last_name}".strip()
        else:
            owner_name = "Все стартапы"
        ws.cell(row=2, column=2, value=owner_name)
        
        headers = [
            "ID", "Название", "Статус", "Категория", "Стадия", 
            "Цель финансирования", "Собрано", "Рейтинг", "Количество инвесторов", "Список инвесторов", "Дата создания"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        if request.user.role and request.user.role.role_name == 'startuper':
            startups = Startups.objects.select_related('owner', 'direction', 'stage').filter(owner=request.user)
        else:
            startups = Startups.objects.select_related('owner', 'direction', 'stage').all()
        
        for row, startup in enumerate(startups, 5):
            try:
                try:
                    startup_id = startup.startup_id
                except Exception:
                    startup_id = row - 1
                ws.cell(row=row, column=1, value=startup_id)
                
                try:
                    title = startup.title or ""
                except Exception:
                    title = ""
                ws.cell(row=row, column=2, value=title)
                
                try:
                    status_display = startup.get_status_display()
                except Exception:
                    status_display = startup.status or "Неизвестен"
                ws.cell(row=row, column=3, value=status_display)
                
                try:
                    direction_name = startup.direction.direction_name if startup.direction else "Не указана"
                except Exception:
                    direction_name = "Не указана"
                ws.cell(row=row, column=4, value=direction_name)
                
                try:
                    stage_name = startup.stage.stage_name if startup.stage else "Не указана"
                except Exception:
                    stage_name = "Не указана"
                ws.cell(row=row, column=5, value=stage_name)
                
                try:
                    funding_goal = startup.funding_goal or 0
                except Exception:
                    funding_goal = 0
                ws.cell(row=row, column=6, value=funding_goal)
                
                try:
                    amount_raised = startup.amount_raised or 0
                except Exception:
                    amount_raised = 0
                ws.cell(row=row, column=7, value=amount_raised)
                
                try:
                    avg_rating = UserVotes.objects.filter(startup=startup).aggregate(Avg('rating'))['rating__avg']
                    ws.cell(row=row, column=8, value=round(avg_rating, 2) if avg_rating else 0)
                except Exception:
                    ws.cell(row=row, column=8, value=0)
                
                try:
                    investors_count = startup.get_investors_count()
                except Exception:
                    investors_count = 0
                ws.cell(row=row, column=9, value=investors_count)
                
                try:
                    investors = (
                        InvestmentTransactions.objects
                        .filter(startup=startup)
                        .select_related('investor')
                        .values_list('investor__first_name', 'investor__last_name')
                        .distinct()
                    )
                    investors_list = [
                        f"{first or ''} {last or ''}".strip() for first, last in investors if first or last
                    ]
                    cell = ws.cell(row=row, column=10, value="\n".join(investors_list))
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                except Exception:
                    cell = ws.cell(row=row, column=10, value="")
                    cell.alignment = Alignment(wrap_text=True, vertical='top')

                try:
                    created_date = startup.created_at.strftime("%d.%m.%Y") if startup.created_at else ""
                except Exception:
                    created_date = ""
                ws.cell(row=row, column=11, value=created_date)
            except Exception as e:
                logger.error(f"Ошибка при обработке стартапа {startup.startup_id}: {e}")
                continue
        
        # Устанавливаем ширину столбцов
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Устанавливаем минимальную высоту строк для корректного отображения переносов
        for row_num in range(5, len(startups) + 5):
            ws.row_dimensions[row_num].height = 50  # Устанавливаем высоту строки
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        report_date = dt.now().strftime('%Y-%m-%d')
        filename = f'Отчет по стартапам_{report_date}.xlsx'
        response['Content-Disposition'] = f'attachment; filename={filename}'
        wb.save(response)
        return response
    except Exception as e:
        logger.error(f"Ошибка при генерации отчета: {e}", exc_info=True)
        return HttpResponse("Ошибка при генерации отчета", status=500)

def approve_franchise(request, franchise_id):
    if not request.user.is_authenticated or (request.user.role.role_name or "").lower() != "moderator":
        messages.error(request, "У вас нет прав для этого действия.")
        return redirect("home")
    franchise = get_object_or_404(Franchises, franchise_id=franchise_id)
    if request.method == "POST":
        moderator_comment = request.POST.get("moderator_comment", "")
        franchise.moderator_comment = moderator_comment
        franchise.status = "approved"
        try:
            franchise.status_id = ReviewStatuses.objects.get(status_name="Approved")
        except ReviewStatuses.DoesNotExist:
            raise ValueError("Статус 'Approved' не найден в базе данных.")
        franchise.save()
        messages.success(request, "Франшиза одобрена.")
    return redirect("moderator_dashboard")


def reject_franchise(request, franchise_id):
    if not request.user.is_authenticated or (request.user.role.role_name or "").lower() != "moderator":
        messages.error(request, "У вас нет прав для этого действия.")
        return redirect("home")
    franchise = get_object_or_404(Franchises, franchise_id=franchise_id)
    if request.method == "POST":
        moderator_comment = request.POST.get("moderator_comment", "")
        franchise.moderator_comment = moderator_comment
        franchise.status = "rejected"
        try:
            franchise.status_id = ReviewStatuses.objects.get(status_name="Rejected")
        except ReviewStatuses.DoesNotExist:
            raise ValueError("Статус 'Rejected' не найден в базе данных.")
        franchise.save()
        messages.success(request, "Франшиза отклонена.")
    return redirect("moderator_dashboard")


def approve_agency(request, franchise_id):
    if not request.user.is_authenticated or (request.user.role.role_name or "").lower() != "moderator":
        messages.error(request, "У вас нет прав для этого действия.")
        return redirect("home")
    agency = get_object_or_404(Agencies, agency_id=franchise_id)
    if request.method == "POST":
        moderator_comment = request.POST.get("moderator_comment", "")
        agency.moderator_comment = moderator_comment
        agency.status = "approved"
        agency.save()
        messages.success(request, "Агентство одобрено.")
    return redirect("moderator_dashboard")


def reject_agency(request, franchise_id):
    if not request.user.is_authenticated or (request.user.role.role_name or "").lower() != "moderator":
        messages.error(request, "У вас нет прав для этого действия.")
        return redirect("home")
    agency = get_object_or_404(Agencies, agency_id=franchise_id)
    if request.method == "POST":
        moderator_comment = request.POST.get("moderator_comment", "")
        agency.moderator_comment = moderator_comment
        agency.status = "rejected"
        agency.save()
        messages.success(request, "Агентство отклонено.")
    return redirect("moderator_dashboard")


def approve_specialist(request, specialist_id):
    if not request.user.is_authenticated or (request.user.role.role_name or "").lower() != "moderator":
        messages.error(request, "У вас нет прав для этого действия.")
        return redirect("home")
    spec = get_object_or_404(Specialists, specialist_id=specialist_id)
    if request.method == "POST":
        moderator_comment = request.POST.get("moderator_comment", "")
        spec.moderator_comment = moderator_comment
        spec.status = "approved"
        spec.save()
        messages.success(request, "Специалист одобрен.")
    return redirect("moderator_dashboard")


def reject_specialist(request, specialist_id):
    if not request.user.is_authenticated or (request.user.role.role_name or "").lower() != "moderator":
        messages.error(request, "У вас нет прав для этого действия.")
        return redirect("home")
    spec = get_object_or_404(Specialists, specialist_id=specialist_id)
    if request.method == "POST":
        moderator_comment = request.POST.get("moderator_comment", "")
        spec.moderator_comment = moderator_comment
        spec.status = "rejected"
        spec.save()
        messages.success(request, "Специалист отклонен.")
    return redirect("moderator_dashboard")


@login_required
def vote_franchise(request, franchise_id):
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Неверный метод запроса"})
    franchise = get_object_or_404(Franchises, franchise_id=franchise_id)
    rating = int(request.POST.get("rating", 0))
    if not 1 <= rating <= 5:
        return JsonResponse(
            {"success": False, "error": "Недопустимое значение рейтинга"}
        )
    if UserVotes.objects.filter(user=request.user, franchise=franchise).exists():
        return JsonResponse(
            {"success": False, "error": "Вы уже голосовали за эту франшизу"}
        )
    UserVotes.objects.create(
        user=request.user, franchise=franchise, rating=rating, created_at=timezone.now()
    )
    franchise.total_voters += 1
    franchise.sum_votes += rating
    franchise.save()
    average_rating = (
        franchise.sum_votes / franchise.total_voters if franchise.total_voters > 0 else 0
    )
    return JsonResponse({"success": True, "average_rating": average_rating})


@login_required
def vote_agency(request, franchise_id):
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Неверный метод запроса"})
    agency = get_object_or_404(Agencies, agency_id=franchise_id)
    rating = int(request.POST.get("rating", 0))
    if not 1 <= rating <= 5:
        return JsonResponse({"success": False, "error": "Недопустимое значение рейтинга"})
    if AgencyVotes.objects.filter(user=request.user, agency=agency).exists():
        return JsonResponse({"success": False, "error": "Вы уже голосовали за это агентство"})
    AgencyVotes.objects.create(user=request.user, agency=agency, rating=rating, created_at=timezone.now())
    agency.total_voters += 1
    agency.sum_votes += rating
    agency.save()
    average_rating = agency.sum_votes / agency.total_voters if agency.total_voters > 0 else 0
    return JsonResponse({"success": True, "average_rating": average_rating})

@login_required
def vote_specialist(request, specialist_id):
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Неверный метод запроса"})
    specialist = get_object_or_404(Specialists, specialist_id=specialist_id)
    rating = int(request.POST.get("rating", 0))
    if not 1 <= rating <= 5:
        return JsonResponse({"success": False, "error": "Недопустимое значение рейтинга"})
    if SpecialistVotes.objects.filter(user=request.user, specialist=specialist).exists():
        return JsonResponse({"success": False, "error": "Вы уже голосовали за этого специалиста"})
    SpecialistVotes.objects.create(user=request.user, specialist=specialist, rating=rating, created_at=timezone.now())
    specialist.total_voters += 1
    specialist.sum_votes += rating
    specialist.save()
    average_rating = specialist.sum_votes / specialist.total_voters if specialist.total_voters > 0 else 0
    return JsonResponse({"success": True, "average_rating": average_rating})

def load_similar_franchises(request, franchise_id: int):
    try:
        franchise = get_object_or_404(Franchises, franchise_id=franchise_id)
        similar_franchises = (
            Franchises.objects.filter(
                direction=franchise.direction,
                status="approved",
            )
            .exclude(franchise_id=franchise_id)
            .order_by("?")[:4]
        )

        context = {
            "similar_franchises": similar_franchises,
        }
        if similar_franchises.count() < 4:
            return HttpResponse("")
        return render(request, "accounts/partials/_similar_franchise_cards.html", context)
    except Exception as e:
        logger.error(f"Ошибка при загрузке похожих франшиз: {e}")
        return JsonResponse({"error": "Ошибка при загрузке похожих франшиз"}, status=500)

@login_required
def load_similar_agencies(request, franchise_id: int):
    try:
        agency = get_object_or_404(Agencies, agency_id=franchise_id)
        if agency.customization_data and "agency_category" in agency.customization_data:
            similar_qs = (
                Agencies.objects.filter(
                    customization_data__agency_category=agency.customization_data.get("agency_category"),
                    status="approved",
                )
                .exclude(agency_id=agency.agency_id)
                .order_by("?")[:4]
            )
        else:
            similar_qs = Agencies.objects.filter(status="approved").exclude(agency_id=agency.agency_id).order_by("?")[:4]
        html = render_to_string(
            "accounts/partials/_similar_agency_cards.html",
            {"similar_franchises": similar_qs, "request": request},
        )
        return HttpResponse(html)
    except Exception as e:
        logger.error(f"Ошибка при загрузке похожих агентств: {e}")
        return JsonResponse({"error": "Ошибка при загрузке похожих агентств"}, status=500)

@login_required
def load_similar_specialists(request, specialist_id: int):
    try:
        specialist = get_object_or_404(Specialists, specialist_id=specialist_id)
        if specialist.customization_data and "specialist_category" in specialist.customization_data:
            similar_qs = (
                Specialists.objects.filter(
                    customization_data__specialist_category=specialist.customization_data.get("specialist_category"),
                    status="approved",
                )
                .exclude(specialist_id=specialist.specialist_id)
                .order_by("?")[:4]
            )
        else:
            similar_qs = Specialists.objects.filter(status="approved").exclude(specialist_id=specialist.specialist_id).order_by("?")[:4]
        html = render_to_string(
            "accounts/partials/_similar_specialist_cards.html",
            {"similar_specialists": similar_qs, "request": request},
        )
        return HttpResponse(html)
    except Exception as e:
        logger.error(f"Ошибка при загрузке похожих специалистов: {e}")
        return JsonResponse({"error": "Ошибка при загрузке похожих специалистов"}, status=500)

@login_required
def delete_startup_comment(request, comment_id: int):
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Недопустимый метод"}, status=405)
    if not hasattr(request.user, "role") or (request.user.role.role_name or "") != "moderator":
        return JsonResponse({"success": False, "error": "Нет прав"}, status=403)
    comment = get_object_or_404(Comments, pk=comment_id)
    comment.delete()
    return JsonResponse({"success": True})

@login_required
def delete_franchise_comment(request, comment_id: int):
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Недопустимый метод"}, status=405)
    if not hasattr(request.user, "role") or (request.user.role.role_name or "") != "moderator":
        return JsonResponse({"success": False, "error": "Нет прав"}, status=403)
    from .models import FranchiseComments
    comment = get_object_or_404(FranchiseComments, pk=comment_id)
    comment.delete()
    return JsonResponse({"success": True})

@login_required
def delete_agency_comment(request, comment_id: int):
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Недопустимый метод"}, status=405)
    if not hasattr(request.user, "role") or (request.user.role.role_name or "") != "moderator":
        return JsonResponse({"success": False, "error": "Нет прав"}, status=403)
    comment = get_object_or_404(AgencyComments, pk=comment_id)
    comment.delete()
    return JsonResponse({"success": True})

@login_required
def delete_specialist_comment(request, comment_id: int):
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Недопустимый метод"}, status=405)
    if not hasattr(request.user, "role") or (request.user.role.role_name or "") != "moderator":
        return JsonResponse({"success": False, "error": "Нет прав"}, status=403)
    comment = get_object_or_404(SpecialistComments, pk=comment_id)
    comment.delete()
    return JsonResponse({"success": True})


@login_required
def edit_franchise(request, franchise_id):
    franchise = get_object_or_404(Franchises, franchise_id=franchise_id)
    if request.user != franchise.owner and request.user.role.role_name != 'moderator':
        messages.error(request, "У вас нет прав для редактирования этой франшизы.")
        return redirect("franchise_detail", franchise_id=franchise_id)
    
    if request.method == "POST":
        franchise.title = request.POST.get("title", franchise.title)
        franchise.description = request.POST.get("description", franchise.description)
        franchise.short_description = request.POST.get("short_description", franchise.short_description)
        franchise.investment_size = request.POST.get("investment_size", franchise.investment_size)
        franchise.franchise_cost = request.POST.get("franchise_cost", franchise.franchise_cost)
        franchise.profit_calculation = request.POST.get("profit_calculation", franchise.profit_calculation)
        franchise.terms = request.POST.get("terms", franchise.terms)
        franchise.additional_info = request.POST.get("additional_info", franchise.additional_info)
        franchise.own_businesses_count = request.POST.get("own_businesses_count", franchise.own_businesses_count)
        franchise.franchise_businesses_count = request.POST.get("franchise_businesses_count", franchise.franchise_businesses_count)
        
        if 'logo' in request.FILES:
            franchise.logo = request.FILES['logo']
        
        franchise.save()
        messages.success(request, "Франшиза успешно обновлена.")
        return redirect("franchise_detail", franchise_id=franchise_id)
    
    context = {
        'franchise': franchise,
    }
    return render(request, 'accounts/edit_franchise.html', context)


@login_required
def change_owner_franchise(request, franchise_id):
    if not request.user.is_authenticated or request.user.role.role_name != 'moderator':
        return JsonResponse({'success': False, 'error': 'Недостаточно прав'})
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Неверный метод запроса'})
    
    try:
        franchise = get_object_or_404(Franchises, franchise_id=franchise_id)
        new_owner_id = request.POST.get('new_owner_id')
        
        if not new_owner_id:
            return JsonResponse({'success': False, 'error': 'ID нового владельца не указан'})
        
        new_owner = get_object_or_404(Users, user_id=new_owner_id)
        franchise.owner = new_owner
        franchise.save()
        
        return JsonResponse({'success': True, 'message': 'Владелец франшизы изменен'})
    except Exception as e:
        logger.error(f"Ошибка при смене владельца франшизы: {e}")
        return JsonResponse({'success': False, 'error': 'Ошибка при смене владельца'})


@login_required
def get_investors_franchise(request, franchise_id):
    if not request.user.is_authenticated or request.user.role.role_name != 'moderator':
        return JsonResponse({'success': False, 'error': 'Недостаточно прав'})
    
    try:
        franchise = get_object_or_404(Franchises, franchise_id=franchise_id)
        investors = InvestmentTransactions.objects.filter(franchise=franchise).select_related('investor')
        
        investors_data = []
        for transaction in investors:
            investors_data.append({
                'user_id': transaction.investor.user_id,
                'name': f"{transaction.investor.first_name or ''} {transaction.investor.last_name or ''}".strip(),
                'amount': float(transaction.amount),
                'date': transaction.created_at.strftime('%d.%m.%Y')
            })
        
        return JsonResponse({'success': True, 'investors': investors_data})
    except Exception as e:
        logger.error(f"Ошибка при получении инвесторов франшизы: {e}")
        return JsonResponse({'success': False, 'error': 'Ошибка при получении данных'})


@login_required
def add_investor_franchise(request, franchise_id):
    if not request.user.is_authenticated or request.user.role.role_name != 'moderator':
        return JsonResponse({'success': False, 'error': 'Недостаточно прав'})
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Неверный метод запроса'})
    
    try:
        franchise = get_object_or_404(Franchises, franchise_id=franchise_id)
        investor_id = request.POST.get('investor_id')
        amount = request.POST.get('amount')
        
        if not investor_id or not amount:
            return JsonResponse({'success': False, 'error': 'Не указаны ID инвестора или сумма'})
        
        investor = get_object_or_404(Users, user_id=investor_id)
        amount_decimal = Decimal(amount)
        
        transaction = InvestmentTransactions(
            franchise=franchise,
            investor=investor,
            amount=amount_decimal,
            transaction_type=TransactionTypes.objects.get(type_name="investment"),
            transaction_status="completed",
            payment_method=PaymentMethods.objects.get(method_name="default"),
            created_at=timezone.now(),
            updated_at=timezone.now(),
        )
        transaction.save()
        
        return JsonResponse({'success': True, 'message': 'Инвестор добавлен'})
    except Exception as e:
        logger.error(f"Ошибка при добавлении инвестора франшизы: {e}")
        return JsonResponse({'success': False, 'error': 'Ошибка при добавлении инвестора'})


@login_required
def edit_investment_franchise(request, franchise_id, user_id):
    if not request.user.is_authenticated or request.user.role.role_name != 'moderator':
        return JsonResponse({'success': False, 'error': 'Недостаточно прав'})
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Неверный метод запроса'})
    
    try:
        franchise = get_object_or_404(Franchises, franchise_id=franchise_id)
        investor = get_object_or_404(Users, user_id=user_id)
        new_amount = request.POST.get('amount')
        
        if not new_amount:
            return JsonResponse({'success': False, 'error': 'Не указана сумма'})
        
        transaction = InvestmentTransactions.objects.get(franchise=franchise, investor=investor)
        transaction.amount = Decimal(new_amount)
        transaction.save()
        
        return JsonResponse({'success': True, 'message': 'Инвестиция обновлена'})
    except InvestmentTransactions.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Инвестиция не найдена'})
    except Exception as e:
        logger.error(f"Ошибка при редактировании инвестиции франшизы: {e}")
        return JsonResponse({'success': False, 'error': 'Ошибка при обновлении'})


@login_required
def delete_investment_franchise(request, franchise_id, user_id):
    if not request.user.is_authenticated or request.user.role.role_name != 'moderator':
        return JsonResponse({'success': False, 'error': 'Недостаточно прав'})
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Неверный метод запроса'})
    
    try:
        franchise = get_object_or_404(Franchises, franchise_id=franchise_id)
        investor = get_object_or_404(Users, user_id=user_id)
        
        transaction = InvestmentTransactions.objects.get(franchise=franchise, investor=investor)
        transaction.delete()
        
        return JsonResponse({'success': True, 'message': 'Инвестиция удалена'})
    except InvestmentTransactions.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Инвестиция не найдена'})
    except Exception as e:
        logger.error(f"Ошибка при удалении инвестиции франшизы: {e}")
        return JsonResponse({'success': False, 'error': 'Ошибка при удалении'})


